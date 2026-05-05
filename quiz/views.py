import json
import random
import pandas as pd
import os
import urllib.parse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from datetime import timedelta
from io import BytesIO
from django.views.decorators.clickjacking import xframe_options_exempt
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync


# Django Core
from django.core.mail import EmailMessage, send_mail
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.conf import settings
from django.db import transaction
from django.forms import inlineformset_factory

# Django Auth & Decorators
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.cache import cache_control
from django.contrib.auth.models import User
from django.contrib.auth import get_user_model

# [핵심] DB Aggregation & Functions
from django.db.models import (
    Avg, Count, Q, Max, Min, F, Case, When, Value, 
    CharField, Window, IntegerField, FloatField
)
from django.db.models.functions import DenseRank, Coalesce, Cast

# Local Apps & Models
from attendance.models import DailySchedule, ScheduleRequest
from .utils import calculate_tag_stats

# accounts 앱 모델 (중복 제거됨)
from accounts.models import (
    Profile, Badge, EvaluationRecord, EvaluationCategory, 
    ManagerEvaluation, Cohort, Company, Process, 
    ProcessAccessRequest, FinalAssessment, PartLeader
)

# quiz 앱 모델
from .models import (
    Quiz, Question, Choice, TestResult, UserAnswer, 
    QuizAttempt, ExamSheet, Tag, StudentLog, Notification, 
    QuizResult, StudentAnswer, ReferenceLink
)

# Forms
from .forms import (
    EvaluationForm, TraineeFilterForm, QuizForm, 
    QuestionForm, StudentLogForm, ChoiceForm
)

def is_process_manager(user, target_profile):
    """
    요청자(user)가 관리자(Superuser)이거나, 
    대상 교육생(target_profile)과 '같은 공정의 매니저(교수)'인지 확인합니다.
    """
    # 1. 최고 관리자는 프리패스
    if user.is_superuser:
        return True
    
    # 2. 매니저(교수)인 경우: 본인의 공정과 학생의 공정이 같은지 확인
    if hasattr(user, 'profile') and user.profile.is_manager:
        if user.profile.process == target_profile.process:
            return True
            
    return False

# 1. '마이 페이지'
@login_required
def my_page(request):
    """
    마이페이지 (홈)
    - 매니저 강제 리다이렉트 제거됨 (누구나 접근 가능)
    - 시험 불합격 시 면담 요청 상태(예정/완료) 체크 로직 포함
    - [수정 완료] 템플릿 경로를 실제 파일 위치인 'quiz/my_page.html'로 변경
    """
    user = request.user
    
    # 1. 프로필 가져오기 (없으면 생성)
    profile, created = Profile.objects.get_or_create(user=user)

    # -------------------------------------------------------
    # [1] 진행 중인 시험 (결재 대기/승인됨)
    # -------------------------------------------------------
    pending_attempts = QuizAttempt.objects.filter(
        user=user, 
        status__in=['대기중', '승인됨']
    )

    # -------------------------------------------------------
    # [2] 시험 결과 + 면담 상태 데이터 가공 (핵심 로직)
    # -------------------------------------------------------
    # 최근 5개 결과 조회
    raw_results = TestResult.objects.filter(user=user).select_related('quiz').order_by('-completed_at')[:5]
    enhanced_results = []

    for result in raw_results:
        counseling_status = None
        
        if not result.is_pass:
            # ★ [수정됨] 자동 생성된 시험 불합격(exam_fail) 잠금 로그의 상태를 확인!
            fail_log = StudentLog.objects.filter(
                profile=profile,
                log_type='exam_fail',
                related_quiz=result.quiz
            ).last()

            if fail_log and not fail_log.is_resolved:
                counseling_status = '대기중' # 잠겨 있음 -> 매니저 호출 대기 중
            else:
                counseling_status = '완료'   # 매니저가 조치 완료(잠금 해제) 함
        
        # 템플릿에서 사용할 데이터 구조 만들기
        enhanced_results.append({
            'result': result,
            'counseling_status': counseling_status
        })

    # -------------------------------------------------------
    # [3] 배지 & 최근 피드백 (평가 로그)
    # -------------------------------------------------------
    latest_badges = profile.badges.all().order_by('-id')[:3]
    
    latest_evaluations = StudentLog.objects.filter(
        profile=profile
    ).order_by('-created_at')[:3]
    
    # -------------------------------------------------------
    # [4] 통계 데이터 (옵션: 필요 시 사용)
    # -------------------------------------------------------
    # 전체 결과 재조회 (통계용)
    all_results = TestResult.objects.filter(user=user)
    total_tests = all_results.count()
    pass_count = all_results.filter(is_pass=True).count()
    
    avg_score = 0
    if total_tests > 0:
        total_score_sum = sum(r.score for r in all_results)
        avg_score = round(total_score_sum / total_tests, 1)

    context = {
        'profile': profile,
        'pending_attempts': pending_attempts,
        'enhanced_results': enhanced_results,   # [핵심] 상태 포함 결과 리스트
        'latest_badges': latest_badges,
        'latest_evaluations': latest_evaluations,
        'stats': {
            'total': total_tests,
            'passed': pass_count,
            'avg_score': avg_score,
        }
    }
    
    # [수정됨] 실제 파일이 존재하는 경로('quiz/my_page.html')로 변경했습니다.
    return render(request, 'quiz/my_page.html', context)


# [신규] 학생이 모달에서 면담 요청/사유를 작성하면 저장하는 함수
@login_required
@require_POST
def student_create_counseling_log(request):
    """
    교육생 면담/상담 요청 함수
    - 시험 성적 문의 -> 매니저의 '교육생 상세 정보(성적표)' 페이지로 이동
    - 기타/경고 문의 -> 매니저의 '로그 관리(특이사항)' 페이지로 이동
    """
    User = get_user_model() 

    try:
        # 1. 데이터 수신
        quiz_title = request.POST.get('quiz_title')
        score = request.POST.get('score')
        ref_log_type = request.POST.get('ref_log_type') 
        user_reason = request.POST.get('reason', '') 

        final_reason = ""
        noti_summary = ""
        target_url = "" # [핵심] 알림 클릭 시 이동할 URL 변수

        # [Case 1] 시험 불합격 원클릭 요청
        if quiz_title:
            final_reason = f"[면담 요청] '{quiz_title}' 시험 불합격 ({score}점)\n- 재시험 및 학습 상담 요청"
            noti_summary = f"'{quiz_title}' 불합격 면담 요청"
            # 시험 관련이므로 '성적 상세 페이지'로 이동
            target_url = reverse('quiz:manager_trainee_detail', args=[request.user.profile.id])
        
        # [Case 2] 특정 기록(경고 등) 상담 요청
        elif ref_log_type:
            final_reason = f"[상담 요청] 관련 기록: {ref_log_type}\n\n[내용]\n{user_reason}"
            noti_summary = f"특이사항({ref_log_type}) 관련 상담"
            # 기록 관련이므로 '로그 관리 페이지'로 이동
            target_url = reverse('quiz:manage_student_logs', args=[request.user.profile.id])
            
        # [Case 3] 일반 직접 작성
        elif user_reason:
            final_reason = user_reason
            # 내용이 길면 말줄임표 처리
            short_reason = (user_reason[:12] + '...') if len(user_reason) > 12 else user_reason
            noti_summary = f"면담 요청: {short_reason}"
            # 일반 상담이므로 '로그 관리 페이지'로 이동
            target_url = reverse('quiz:manage_student_logs', args=[request.user.profile.id])
            
        else:
            messages.error(request, "요청 내용이 없습니다.")
            return redirect('quiz:my_page')

        # 2. DB에 로그(StudentLog) 저장
        StudentLog.objects.create(
            profile=request.user.profile,
            recorder=request.user, 
            log_type='counseling',
            reason=final_reason,
            is_resolved=False 
        )
        
        # 3. 관리자 알림 발송
        managers = User.objects.filter(is_staff=True)
        
        for manager in managers:
            Notification.objects.create(
                recipient=manager,       
                sender=request.user,     
                message=f"📢 [{request.user.profile.name}] {noti_summary}",        
                notification_type='counseling', 
                related_url=target_url   # [핵심] 위에서 결정된 URL 저장
            )
        
        messages.success(request, "면담 요청이 전송되었습니다.")
        
    except Exception as e:
        messages.error(request, f"오류 발생: {e}")
    
    return redirect('quiz:my_page')

from django.db.models import Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Quiz, QuizAttempt, StudentLog, TestResult

# =======================================================
# [1] 헬퍼 함수: 퀴즈 리스트에 상태(잠금, 완료, 대기 등)를 붙여주는 함수
# =======================================================
def process_quiz_list(quiz_list, user):
    """
    퀴즈 쿼리셋(목록)을 받아, 각 퀴즈 객체에 현재 유저의 상태(user_status)와
    불합격 잠금 로그(blocking_log) 등을 매핑하여 반환합니다.
    """
    user_groups = user.groups.values_list('id', flat=True)

    for quiz in quiz_list:
        quiz.user_status = None
        quiz.action_id = None
        quiz.is_pass = False
        quiz.blocking_log = None  # [초기화] 잠금 로그 필드

        # -----------------------------------------------------------
        # [0] 잠금(Penalty) 상태 우선 확인 (입구컷 로직 연동)
        # -----------------------------------------------------------
        if hasattr(user, 'profile'):
            quiz.blocking_log = StudentLog.objects.filter(
                profile__user=user,
                related_quiz=quiz,
                log_type='exam_fail', # 시험 불합격 제재만 확인
                is_resolved=False     # 해결되지 않음 = 잠금 상태
            ).last()

        # 1. 최근 결과 확인
        latest_result = TestResult.objects.filter(user=user, quiz=quiz).order_by('-completed_at').first()
        
        # 2. 요청 상태 (개인 요청: 대기중 or 승인됨)
        active_individual_attempt = QuizAttempt.objects.filter(
            user=user, quiz=quiz, 
            assignment_type=QuizAttempt.AssignmentType.INDIVIDUAL,
            status__in=['대기중', '승인됨'],
            testresult__isnull=True
        ).first()

        if active_individual_attempt:
            quiz.user_status = active_individual_attempt.status
            quiz.action_id = active_individual_attempt.id
            continue

        # 3. 그룹 상태 확인
        is_individually_assigned = quiz.allowed_users.filter(id=user.id).exists()
        is_group_assigned = quiz.allowed_groups.filter(id__in=user_groups).exists()
        
        if is_group_assigned and not is_individually_assigned:
            # 그룹 할당이 되어있지만, 이미 완료한 기록이 있는지 확인
            completed_group_attempt = TestResult.objects.filter(
                user=user, quiz=quiz, 
                attempt__assignment_type=QuizAttempt.AssignmentType.GROUP
            ).exists()
            if not completed_group_attempt:
                quiz.user_status = '그룹 응시 가능'
                quiz.action_id = quiz.id
                continue
        
        # 4. 완료됨 확인
        if latest_result:
            quiz.user_status = '완료됨'
            quiz.action_id = latest_result.id
            quiz.is_pass = latest_result.is_pass
            continue
        
        # 5. 기본 상태 (아무 기록 없음 -> 요청 가능)
        quiz.user_status = '요청 가능'
        quiz.action_id = quiz.id
        
    return quiz_list


# =======================================================
# [2] 메인 뷰: 대시보드 Index (실제 작동 함수)
# =======================================================
@login_required
def index(request):
    """
    대시보드 메인 페이지 (교육생 센터 홈)
    """
    user = request.user
    user_groups = user.groups.all()
    
    # 사용자 프로필 및 공정 정보 가져오기
    user_process = None
    if hasattr(user, 'profile') and user.profile.process:
        user_process = user.profile.process

    # -------------------------------------------------------
    # [A] 공통 과목 (Common) 쿼리
    # -------------------------------------------------------
    all_common_quizzes = Quiz.objects.filter(
        category=Quiz.Category.COMMON,
        is_published=True
    ).distinct()

    # -------------------------------------------------------
    # [B] 권한 쿼리 (내 그룹 or 나에게 직접 할당)
    # -------------------------------------------------------
    permission_query = Q(allowed_groups__in=user_groups) | Q(allowed_users=user)

    # -------------------------------------------------------
    # [C] '나의 공정' 퀴즈 목록 쿼리
    # -------------------------------------------------------
    my_process_condition = Q(related_process=user_process) | permission_query
    
    if user_process is None:
        my_process_condition = permission_query

    my_process_quizzes_list = Quiz.objects.filter(
        Q(category=Quiz.Category.PROCESS) & 
        (my_process_condition) &
        Q(is_published=True)
    ).distinct()

    # -------------------------------------------------------
    # [D] '기타 공정' 퀴즈 목록 쿼리
    # -------------------------------------------------------
    # 전체 프로세스 퀴즈 중 '나의 공정 퀴즈'를 제외한 나머지
    other_process_quizzes_list = Quiz.objects.filter(
        category=Quiz.Category.PROCESS,
        is_published=True
    ).exclude(
        id__in=my_process_quizzes_list.values('id')
    ).distinct()

    # -------------------------------------------------------
    # 안전(Safety) & 기타(Etc) 과목 쿼리 추가
    # -------------------------------------------------------
    safety_quizzes_list = Quiz.objects.filter(
        category=Quiz.Category.SAFETY,
        is_published=True
    ).distinct()

    etc_quizzes_list = Quiz.objects.filter(
        category=Quiz.Category.ETC,
        is_published=True
    ).distinct()

# -------------------------------------------------------
    # [E] 합격 여부 & 잠금 해제 조건 체크 (분리 적용)
    # -------------------------------------------------------
    
    # 1. 공통 과목 체크
    # (A) "합격" 여부 (수료증/통계용 -> 80점 이상)
    all_common_passed = False
    passed_common_count = TestResult.objects.filter(
        user=user, quiz__in=all_common_quizzes, is_pass=True
    ).values('quiz').distinct().count()
    
    if all_common_quizzes.count() > 0 and passed_common_count >= all_common_quizzes.count():
        all_common_passed = True
    elif all_common_quizzes.count() == 0:
        all_common_passed = True

    # (B) "잠금 해제" 여부 (다음 단계 오픈용 -> 응시만 하면 OK)
    all_common_unlocked = False
    attempted_common_count = TestResult.objects.filter(
        user=user, quiz__in=all_common_quizzes
    ).values('quiz').distinct().count()

    if all_common_quizzes.count() > 0 and attempted_common_count >= all_common_quizzes.count():
        all_common_unlocked = True
    elif all_common_quizzes.count() == 0:
        all_common_unlocked = True


    # 2. 내 공정 과목 체크
    # (A) "합격" 여부
    all_my_process_passed = False
    passed_my_process_count = TestResult.objects.filter(
        user=user, quiz__in=my_process_quizzes_list, is_pass=True
    ).values('quiz').distinct().count()
    
    if my_process_quizzes_list.count() > 0 and passed_my_process_count >= my_process_quizzes_list.count():
        all_my_process_passed = True
    elif my_process_quizzes_list.count() == 0:
        all_my_process_passed = True

    # (B) "잠금 해제" 여부
    all_my_process_unlocked = False
    attempted_my_process_count = TestResult.objects.filter(
        user=user, quiz__in=my_process_quizzes_list
    ).values('quiz').distinct().count()

    if my_process_quizzes_list.count() > 0 and attempted_my_process_count >= my_process_quizzes_list.count():
        all_my_process_unlocked = True
    elif my_process_quizzes_list.count() == 0:
        all_my_process_unlocked = True

    # -------------------------------------------------------
    # [F] 상태 매핑 (위에 있는 헬퍼 함수 호출)
    # -------------------------------------------------------
    common_quizzes = process_quiz_list(all_common_quizzes, user)
    my_process_quizzes = process_quiz_list(my_process_quizzes_list, user)
    other_process_quizzes = process_quiz_list(other_process_quizzes_list, user)
    safety_quizzes = process_quiz_list(safety_quizzes_list, user)
    etc_quizzes = process_quiz_list(etc_quizzes_list, user)

    # -------------------------------------------------------
    # [G] 배지 표시 여부 (진행중인 건이 있는지 체크)
    # -------------------------------------------------------
    my_process_has_override = any(quiz.user_status in ['승인됨', '대기중'] for quiz in my_process_quizzes)
    other_process_has_override = any(quiz.user_status in ['승인됨', '대기중'] for quiz in other_process_quizzes)
    safety_has_override = any(quiz.user_status in ['승인됨', '대기중'] for quiz in safety_quizzes)
    etc_has_override = any(quiz.user_status in ['승인됨', '대기중'] for quiz in etc_quizzes)

    context = {
        'common_quizzes': common_quizzes,
        'my_process_quizzes': my_process_quizzes,
        'other_process_quizzes': other_process_quizzes,
        
        'safety_quizzes': safety_quizzes,
        'etc_quizzes': etc_quizzes,
        'safety_has_override': safety_has_override,
        'etc_has_override': etc_has_override,
        
        'all_common_passed': all_common_passed,
        'all_my_process_passed': all_my_process_passed,
        
        'my_process_has_override': my_process_has_override,
        'other_process_has_override': other_process_has_override,
        
        'profile': getattr(user, 'profile', None),

        'all_common_unlocked': all_common_unlocked,
        'all_my_process_unlocked': all_my_process_unlocked,
    }

    return render(request, 'quiz/index.html', context)

    
@login_required
def request_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)

    # ===============================================================
    # 입구컷: 잠금 상태라면 요청 자체를 차단
    # ===============================================================
    if hasattr(request.user, 'profile'):
        blocking_log = StudentLog.objects.filter(
            profile__user=request.user,
            related_quiz=quiz,
            log_type='exam_fail',
            is_resolved=False
        ).last()

        if blocking_log:
            if blocking_log.stage == 1:
                messages.error(request, "🚫 [1차 불합격] 재응시가 잠금되었습니다. 담당 매니저와 면담이 필요합니다.")
            elif blocking_log.stage == 2:
                messages.error(request, "🚫 [2차 불합격] 재응시가 잠금되었습니다. 파트장(PL) 면담이 필요합니다.")
            else:
                messages.error(request, "🚫 불합격 패널티로 인해 요청할 수 없습니다.")
            return redirect('quiz:index')

    # 기존 로직: 대기중이거나 승인된 요청이 있는지 확인
    existing_attempt = QuizAttempt.objects.filter(
        user=request.user, 
        quiz=quiz, 
        status__in=['대기중', '승인됨']
    ).first()

    if existing_attempt:
        if existing_attempt.status == '승인됨':
            messages.info(request, f"이미 승인된 '{quiz.title}' 시험이 있습니다. 바로 응시해주세요.")
        else:
            messages.warning(request, f"이미 '{quiz.title}' 시험 요청이 대기 중입니다.")
    else:
        
        QuizAttempt.objects.create(
            user=request.user, 
            quiz=quiz, 
            assignment_type=QuizAttempt.AssignmentType.INDIVIDUAL,
            status='대기중' 
        )
        # 새로운 요청 생성
        from quiz.models import Notification
        from django.urls import reverse
        
        # 1. 해당 학생의 공정을 담당하는 매니저 찾기
        target_managers = User.objects.filter(
            profile__is_manager=True, 
            profile__process=request.user.profile.process
        )
        # 2. 담당 매니저가 없으면 최고관리자에게 발송
        if not target_managers.exists():
            target_managers = User.objects.filter(is_superuser=True)
            
        for manager in target_managers:
            Notification.objects.create(
                recipient=manager,
                sender=request.user,
                message=f"📝 [시험 요청] {request.user.profile.name}님이 '{quiz.title}' 응시를 요청했습니다.",
                related_url=reverse('quiz:manager_exam_requests'), # 클릭 시 승인 대기함으로 이동
                icon='bi-pencil-square',
                notification_type='exam'
            )
        
        broadcast_realtime_notification(manager.id)
        # =========================================================

        messages.success(request, f"'{quiz.title}' 시험 응시를 요청했습니다. 관리자의 승인을 기다려 주세요.")
    
    return redirect('quiz:index')

@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def take_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)

    # =========================================================
    # [1] 보안 및 입구컷 (기존 로직 100% 유지)
    # =========================================================
    
    # [디버깅]
    print(f"\n[DEBUG] 입구컷 검사 시작 - 사용자: {request.user}, 퀴즈: {quiz.title}")

    # 1. 프로필 존재 여부 확인
    if not hasattr(request.user, 'profile'):
        print("[DEBUG] 차단됨: 프로필 없음")
        messages.error(request, "계정 프로필이 설정되지 않아 응시할 수 없습니다. 관리자에게 문의하세요.")
        return redirect('quiz:index')

    # 2. 미해결된 '시험 불합격' 제재 확인 (Fail-Closed)
    blocking_log = StudentLog.objects.filter(
        profile__user=request.user,
        related_quiz=quiz,
        log_type='exam_fail',
        is_resolved=False
    ).last()

    if blocking_log:
        print(f"[DEBUG] 차단됨! 로그 발견 (ID: {blocking_log.id}, 단계: {blocking_log.stage})")
        if blocking_log.stage == 1:
            messages.error(request, "🚫 [1차 불합격] 재응시가 잠금되었습니다. 담당 매니저와 면담이 필요합니다.")
        elif blocking_log.stage == 2:
            messages.error(request, "🚫 [2차 불합격] 재응시가 잠금되었습니다. 파트장(PL) 면담이 필요합니다.")
        elif blocking_log.stage >= 3:
            messages.error(request, "🚫 [3차 불합격] 퇴소 기준에 도달하여 더 이상 응시할 수 없습니다.")
        else:
            messages.error(request, "🚫 불합격 패널티로 인해 응시가 제한되었습니다.")
        return redirect('quiz:index')
    else:
        print("[DEBUG] 통과: 차단할 불합격 로그가 없습니다.")


    # =========================================================
    # [2] 시험 준비 (유령 청소 + 좀비 방지 패치 적용)
    # =========================================================
    
    # 1. [패치] 입장 시, 중복된 '진행중' 시험지가 있다면 최신 것만 남기고 삭제 (0점 오류 원인 제거)
    if request.method == 'GET':
        ghosts = QuizAttempt.objects.filter(user=request.user, quiz=quiz, status='진행중').order_by('-id')
        if ghosts.count() > 1:
            print(f"[DEBUG] 유령 시험지 {ghosts.count()-1}개 청소")
            for g in ghosts[1:]:
                g.delete()

    # 2. Attempt ID 가져오기
    attempt_id = None
    # [핵심] '진짜 제출(finish)'일 때만 POST 데이터를 신뢰
    if request.method == 'POST' and request.POST.get('action') == 'finish':
        attempt_id = request.POST.get('custom_attempt_id')
    
    if not attempt_id:
        attempt_id = request.session.get('attempt_id')

    # 3. [패치] 완료된 좀비 세션 확인 사살
    if attempt_id:
        zombie_check = QuizAttempt.objects.filter(pk=attempt_id).first()
        if not zombie_check or zombie_check.status == '완료됨':
            print(f"[DEBUG] 완료된 좀비 세션(ID: {attempt_id}) 발견 -> 폐기")
            attempt_id = None
            request.session.pop('attempt_id', None)
            request.session.pop('quiz_questions', None)

    # 4. 문제 ID 로드
    question_ids = request.session.get('quiz_questions')

    # [세션 복구 로직] (기존 기능 유지)
    if request.method == 'POST' and not question_ids:
        recovered_ids = []
        for key in request.POST.keys():
            if key.startswith('question_'):
                try: recovered_ids.append(int(key.replace('question_', '')))
                except: pass
        if recovered_ids:
            question_ids = recovered_ids
            request.session['quiz_questions'] = question_ids

    # 5. Attempt 생성/로드
    if not attempt_id:
        # POST 요청인데 ID가 없으면 세션 만료
        if request.method == 'POST':
            messages.error(request, "시험 세션이 만료되었습니다. 다시 시도해주세요.")
            return redirect('quiz:index')

        ongoing = QuizAttempt.objects.filter(user=request.user, quiz=quiz, status='진행중').last()
        if ongoing:
            attempt_id = ongoing.id
            if not question_ids: 
                try: question_ids = list(quiz.question_set.values_list('id', flat=True))
                except: question_ids = list(quiz.questions.values_list('id', flat=True))
        else:
            # 새 시험지 생성
            new_att = QuizAttempt.objects.create(user=request.user, quiz=quiz, status='진행중')
            attempt_id = new_att.id
            
            try: all_ids = list(quiz.question_set.values_list('id', flat=True))
            except: all_ids = list(quiz.questions.values_list('id', flat=True))
            
            random.shuffle(all_ids)
            limit = quiz.question_count if quiz.question_count else 25
            question_ids = all_ids[:limit]
        
        request.session['attempt_id'] = attempt_id
        request.session['quiz_questions'] = question_ids
    
    attempt = get_object_or_404(QuizAttempt, pk=attempt_id)

    # 이미 완료된 시험지 접근 방지
    if attempt.status == '완료됨':
        last = TestResult.objects.filter(attempt=attempt).last()
        return redirect('quiz:exam_result', result_id=last.id) if last else redirect('quiz:index')

    # 문제 객체 로드
    target_questions = []
    if question_ids:
        qs = Question.objects.filter(pk__in=question_ids)
        q_dict = {q.id: q for q in qs}
        for qid in question_ids:
            if qid in q_dict: target_questions.append(q_dict[qid])
    else:
        try: target_questions = list(quiz.question_set.all())[:25]
        except: target_questions = list(quiz.questions.all())[:25]


    # =========================================================
    # [3] 제출 및 채점 (중복 방지 + 오류 수정 패치 완료)
    # =========================================================
    if request.method == 'POST' and request.POST.get('action') == 'finish':
        
        with transaction.atomic():
            attempt.refresh_from_db()
            if attempt.status == '완료됨': 
                return redirect('quiz:exam_result', result_id=TestResult.objects.filter(attempt=attempt).first().id)

            earned_score = 0.0
            answers_to_save = []
            
            # [패치] 배점 필드 안전 확인 (getattr)
            total_assigned = sum(getattr(q, 'score', 0) for q in target_questions)
            use_assigned = total_assigned > 0
            default_score = 100 / len(target_questions) if target_questions else 0

            for q in target_questions:
                user_val = request.POST.get(f'question_{q.id}', '')
                user_list = request.POST.getlist(f'question_{q.id}')
                
                save_text = ""
                sel_obj = None
                is_correct = False
                
                # [패치] 안전한 배점 가져오기
                q_score = getattr(q, 'score', 0) if use_assigned else default_score

                # --- 채점 로직 (기존 유지) ---
                if q.question_type in ['multiple_select', '다중선택', '다중']:
                    if user_list:
                        sub_ids = set(int(x) for x in user_list if x.isdigit())
                        cor_ids = set(q.choice_set.filter(is_correct=True).values_list('id', flat=True))
                        if sub_ids == cor_ids and sub_ids: is_correct = True
                        save_text = ", ".join([c.choice_text for c in Choice.objects.filter(id__in=sub_ids)])
                
                elif q.question_type in ['multiple_choice', 'true_false', '객관식', 'OX']:
                    val = str(user_val).strip()
                    if val.isdigit():
                        try:
                            c = Choice.objects.get(pk=val)
                            sel_obj = c; save_text = c.choice_text
                            if c.is_correct: is_correct = True
                        except: pass
                    else:
                        save_text = val
                        if q.choice_set.filter(is_correct=True, choice_text__iexact=val).exists(): is_correct = True
                
                else: 
                    save_text = str(user_val).strip()
                    
                    if save_text:
                        # 1. 정답 정규화 함수 (특수문자 제거)
                        def normalize_text(text):
                            if not text: return ""
                            text = str(text).lower()
                            # ★ 콤마(,)를 포함하여 띄어쓰기, 마침표 등을 전부 지우고 비교합니다!
                            for char in ['.', ' ', '-', '_', '/', '(', ')', ',']: 
                                text = text.replace(char, '')
                            return text.strip()

                        # 사용자 입력 정규화
                        user_norm = normalize_text(save_text)
                        
                        # [비교] Choice(복수 정답)들과 1:1 비교 (이제 split 안 합니다!)
                        choices = q.choice_set.filter(is_correct=True)
                        for c in choices:
                            if normalize_text(c.choice_text) == user_norm:
                                is_correct = True
                                break
                        
                        # DB에 저장된 정답 가져오기
                        ans_field = getattr(q, 'answer', None)
                        
                        # [비교 1] answer 필드와 비교 (콤마 쪼개기 적용!)
                        if ans_field:
                            # ★ 콤마 대신 파이프(|)로 쪼개서 리스트로 만듦
                            possible_answers = [x.strip() for x in ans_field.split('|')]
                            
                            for possible in possible_answers:
                                # 쪼갠 것들 중 하나라도 맞으면 정답!
                                if normalize_text(possible) == user_norm:
                                    is_correct = True
                                    break
                        
                        # [비교 2] Choice(복수 정답)들과 비교 (기존 유지)
                        if not is_correct:
                            choices = q.choice_set.filter(is_correct=True)
                            for c in choices:
                                # ★ 여기도 콤마 대신 파이프(|)로 변경!
                                c_answers = [x.strip() for x in c.choice_text.split('|')]
                                for c_ans in c_answers:
                                    if normalize_text(c_ans) == user_norm:
                                        is_correct = True
                                        break
                                if is_correct: break

                # [공통] 점수 합산 및 저장 데이터 수집
                if is_correct: earned_score += q_score
                answers_to_save.append({'q':q, 'text':save_text, 'sel':sel_obj, 'is_c':is_correct})

            # ---------------------------------------------------------
            # [루프 종료] 최종 점수 계산 및 결과 저장
            # ---------------------------------------------------------
            final_score = min(int(round(earned_score)), 100)
            is_pass = final_score >= quiz.pass_score

            # ★ [패치] create 대신 update_or_create 사용 (중복 결과 생성 100% 방지)
            tr, created = TestResult.objects.update_or_create(
                attempt=attempt, 
                defaults={
                    'user': request.user,
                    'quiz': quiz,
                    'score': final_score,
                    'is_pass': is_pass,
                    'completed_at': timezone.now()
                }
            )
            
            # 기존 답안 삭제 후 재생성 (덮어쓰기)
            UserAnswer.objects.filter(test_result=tr).delete()
            for item in answers_to_save:
                UserAnswer.objects.create(
                    test_result=tr, question=item['q'], short_answer_text=item['text'], 
                    selected_choice=item['sel'], is_correct=item['is_c']
                )
            
            attempt.status = '완료됨'
            attempt.result = tr
            attempt.score = final_score
            attempt.save()

            # 세션 정리
            request.session.pop('quiz_questions', None)
            request.session.pop('attempt_id', None)

            # 통계 갱신 (에러 방지 처리)
            try: update_student_stats_force(request.user.profile)
            except: pass


            # =========================================================
            # [4] 불합격 시 3단계 제재 실행 (기존 코드 100% 복구)
            # =========================================================
            if not is_pass:

                if attempt.quiz.category in ['safety', 'etc']:
                    messages.warning(request, f"불합격입니다. ({final_score}점) 하지만 이 과목은 패널티가 없습니다.")
                    # 여기서 바로 리턴하거나, 아래 로직을 else로 감싸야 합니다.
                    return redirect('quiz:exam_result', result_id=tr.id)
                
                last_pass = TestResult.objects.filter(
                    user=request.user, quiz=quiz, is_pass=True
                ).order_by('-completed_at').first()
                
                fail_query = TestResult.objects.filter(
                    user=request.user, quiz=quiz, is_pass=False
                )
                if last_pass:
                    fail_query = fail_query.filter(completed_at__gt=last_pass.completed_at)
                
                fail_count = fail_query.count()

                if hasattr(request.user, 'profile'):
                    profile = request.user.profile
                    
                    # -----------------------------------------------------
                    # ★ [신규 자동 알림 함수] 최고 관리자 + 해당 공정 매니저에게 발송
                    # -----------------------------------------------------
                    def send_auto_fail_noti():
                        from django.urls import reverse
                        target_url = reverse('quiz:manager_trainee_detail', args=[profile.id])
                        cat_str = "공통" if quiz.category == 'common' else "공정"
                        
                        msg = f"🚨 {profile.name}님 {cat_str} '{quiz.title}' 시험 {final_score}점으로 재시험 전 면담이 필요합니다."
                        
                        # 1. 최고 관리자 전원 수집 (set을 이용해 중복 방지)
                        receivers = set(User.objects.filter(is_superuser=True))
                        
                        # 2. 해당 학생과 같은 공정을 담당하는 매니저 수집
                        if profile.process:
                            managers = User.objects.filter(
                                is_staff=True, 
                                profile__is_manager=True, 
                                profile__process=profile.process
                            )
                            receivers.update(managers)
                            
                        # 3. 수집된 인원들에게 알림 생성
                        for recv in receivers:
                            Notification.objects.create(
                                recipient=recv,
                                sender=request.user,
                                message=msg,
                                notification_type='counseling',
                                related_url=target_url
                            )
                    # -----------------------------------------------------

                    if fail_count == 1:
                        # 1차: 잠금 및 자동 알림
                        if not StudentLog.objects.filter(profile=profile, related_quiz=quiz, log_type='exam_fail', stage=1, is_resolved=False).exists():
                            StudentLog.objects.create(
                                profile=profile,
                                log_type='exam_fail',
                                reason=f"[{quiz.title}] 1차 불합격 - 재응시 잠금 (면담 필요)",
                                related_quiz=quiz,
                                stage=1,
                                is_resolved=False
                            )
                            send_auto_fail_noti() # ★ 제출 즉시 알림 빵!
                        messages.error(request, "1차 불합격입니다. 매니저에게 면담 요청 알림이 자동 전송되었습니다.")

                    elif fail_count == 2:
                        # 2차: 잠금 및 자동 알림
                        if not StudentLog.objects.filter(profile=profile, related_quiz=quiz, log_type='exam_fail', stage=2, is_resolved=False).exists():
                            StudentLog.objects.create(
                                profile=profile,
                                log_type='exam_fail',
                                reason=f"[{quiz.title}] 2차 불합격 - PL 면담 필요 (메일 발송됨)",
                                related_quiz=quiz,
                                stage=2,
                                is_resolved=False
                            )
                            send_auto_fail_noti() # ★ 제출 즉시 알림 빵!
                        messages.error(request, "2차 불합격입니다. 매니저/PL에게 면담 요청 알림이 자동 전송되었습니다.")

                    elif fail_count >= 3:
                        # 3차: 계정 차단 및 확실한 퇴소 처리
                        if not StudentLog.objects.filter(profile=profile, related_quiz=quiz, log_type='exam_fail', stage=3, is_resolved=False).exists():
                            reason_text = f"[시스템 자동] '{quiz.title}' 시험 3회 불합격으로 인한 퇴소"
                            StudentLog.objects.create(
                                profile=profile,
                                log_type='exam_fail',
                                reason=reason_text, # 구체적인 사유 기록
                                related_quiz=quiz,
                                stage=3,
                                is_resolved=False
                            )
                            # 계정 차단
                            request.user.is_active = False
                            request.user.save()
                            
                            # ★ [핵심 추가] 프로필 상태를 확실하게 '퇴소'로 변경
                            profile.status = 'dropout'
                            profile.save()
                            
                        messages.error(request, "3차 불합격으로 퇴소 처리 및 계정이 비활성화되었습니다.")
            else:
                messages.success(request, f"합격입니다! 점수: {final_score}점")

            return redirect('quiz:exam_result', result_id=tr.id)

    # =========================================================
    # [5] 화면 렌더링 (GET)
    # =========================================================
    for q in target_questions:
        choices = list(q.choice_set.all())
        random.shuffle(choices)
        q.shuffled_choices = choices

    context = {
        'quiz': quiz,
        'questions': target_questions,
        'attempt': attempt,
        'start_time': attempt.started_at.isoformat() if attempt.started_at else timezone.now().isoformat(),
        'is_in_test_mode': True,
    }

    return render(request, 'quiz/take_quiz.html', context)


# ==================================================================
# 2. 데이터 저장 처리 (Save) - 수정됨
# ==================================================================
@staff_member_required
@require_POST
def bulk_add_sheet_save(request):
    try:
        body = json.loads(request.body)
        quiz_id = body.get('quiz_id')
        raw_data = body.get('data', []) # [row, row, ...] 형식의 리스트

        if not quiz_id:
            return JsonResponse({'status': 'error', 'message': '선택된 시험(Quiz)이 없습니다.'})

        try:
            target_quiz = Quiz.objects.get(id=quiz_id)
        except Quiz.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '존재하지 않는 시험입니다.'})

        success_count = 0

        # 데이터 정합성을 위해 트랜잭션 사용 (중간에 에러나면 롤백)
        with transaction.atomic():
            for row in raw_data:
                # 데이터 포맷: [0:문제, 1:유형, 2:난이도, 3:태그, 4~7:보기, 8:정답]
                
                # 1. 문제 내용 (없으면 건너뜀)
                question_text = str(row[0] or '').strip()
                if not question_text:
                    continue 

                # 2. 유형(Type) 판별 [핵심 수정: 한글 -> 영어 코드 변환]
                # 템플릿(html)에서 사용하는 영어 코드와 일치시켜야 합니다.
                raw_type = str(row[1] or '').strip()
                q_type = 'multiple_choice' # 기본값

                if '주관식' in raw_type or '단답' in raw_type:
                    q_type = 'short_answer'
                elif 'OX' in raw_type.upper():
                    q_type = 'true_false'
                elif '다중' in raw_type:
                    q_type = 'multiple_select'
                # 그 외는 'multiple_choice' (객관식)

                # 3. 난이도 처리 (영어 코드로 변환 권장)
                raw_diff = str(row[2] or '중').strip()
                difficulty = 'medium'
                if '하' in raw_diff: difficulty = 'easy'
                elif '상' in raw_diff: difficulty = 'hard'

                # 4. 문제 생성
                # created_by 필드는 모델에 없다고 하셔서 제외했습니다.
                new_question = Question.objects.create(
                    question_text=question_text,
                    question_type=q_type,
                    difficulty=difficulty
                )

                # 시험에 문제 연결
                target_quiz.questions.add(new_question)

                # 5. 태그 처리
                tags_str = str(row[3] or '').strip()
                if tags_str:
                    # 쉼표나 공백으로 구분하여 태그 저장
                    for tag_name in tags_str.replace(',', ' ').split():
                        if tag_name.strip():
                            tag, _ = Tag.objects.get_or_create(name=tag_name.strip())
                            new_question.tags.add(tag)

                # 6. 정답 및 보기 처리 [핵심 수정 구간]
                raw_answer = str(row[8] or '').strip() # 정답 칸 데이터

                # (A) 주관식 처리
                if q_type == 'short_answer':
                    # 보기 1~4는 무시하고, '정답' 칸의 텍스트를 정답으로 저장
                    if raw_answer:
                        Choice.objects.create(
                            question=new_question, 
                            choice_text=raw_answer, 
                            is_correct=True
                        )

                # (B) OX 퀴즈 처리
                elif q_type == 'true_false':
                    user_ans = raw_answer.upper()
                    # O/X 보기 자동 생성 및 정답 체크
                    Choice.objects.create(question=new_question, choice_text='O', is_correct=(user_ans == 'O'))
                    Choice.objects.create(question=new_question, choice_text='X', is_correct=(user_ans == 'X'))

                # (C) 객관식 / 다중선택 처리
                else:
                    # 보기 데이터 가져오기 (row[4] ~ row[7])
                    options = [row[4], row[5], row[6], row[7]]
                    
                    # 정답 번호/텍스트 파싱
                    # 예: "1" 또는 "1,3" 또는 "사과"
                    correct_indices = [] # 정답인 보기의 인덱스(1~4) 저장용
                    correct_texts = []   # 정답인 텍스트 저장용

                    # 콤마로 구분된 정답 처리 (다중선택 대응)
                    parts = [p.strip() for p in raw_answer.replace(' ', '').split(',') if p.strip()]
                    
                    for p in parts:
                        if p.isdigit():
                            correct_indices.append(int(p))
                        else:
                            correct_texts.append(p)

                    # 보기 생성 루프
                    for idx, opt_raw in enumerate(options, start=1):
                        opt_text = str(opt_raw or '').strip()
                        if opt_text:
                            is_correct = False
                            
                            # 번호로 정답 체크 (예: 정답칸에 '1' 입력 시 1번 보기 정답)
                            if idx in correct_indices:
                                is_correct = True
                            # 텍스트로 정답 체크 (예: 정답칸에 '사과' 입력 시 내용이 '사과'인 보기 정답)
                            elif opt_text in correct_texts: 
                                is_correct = True
                            # 또는 입력된 정답 텍스트와 보기가 정확히 일치하는 경우
                            elif raw_answer == opt_text:
                                is_correct = True

                            Choice.objects.create(
                                question=new_question, 
                                choice_text=opt_text, 
                                is_correct=is_correct
                            )

                success_count += 1

        return JsonResponse({'status': 'success', 'count': success_count})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'})
    except Exception as e:
        print(f"Bulk Save Error: {e}")
        return JsonResponse({'status': 'error', 'message': f'오류 발생: {str(e)}'})

# =========================================================
# [2] 퀴즈 결과 처리 (quiz_results)
# =========================================================
@login_required
def quiz_results(request):
    # 세션에서 데이터 로드
    question_ids = request.session.get('quiz_questions', [])
    user_answers = request.session.get('user_answers', {})
    attempt_id = request.session.get('attempt_id')

    if not question_ids or not attempt_id:
        messages.error(request, "제출된 답안이 없습니다.")
        return redirect('quiz:index')

    attempt = get_object_or_404(QuizAttempt, pk=attempt_id)
    
    # 중복 제출 방지
    if attempt.status == '완료됨':
        return redirect('quiz:my_results_index')

    profile, created = Profile.objects.get_or_create(user=request.user)
    badges_before = set(profile.badges.values_list('id', flat=True))

    # 문제 순서대로 가져오기 (채점 정확성)
    questions_dict = {q.id: q for q in Question.objects.filter(pk__in=question_ids)}
    
    correct_answers = 0
    results_data = []

    # TestResult 가져오기 (start_quiz에서 만들었으므로 get)
    # 혹시 없으면 create (안전장치)
    test_result, _ = TestResult.objects.get_or_create(
        attempt=attempt, 
        defaults={'user': request.user, 'quiz': attempt.quiz, 'score': 0, 'is_pass': False}
    )

    for q_id in question_ids:
        question = questions_dict.get(q_id)
        if not question: continue

        q_id_str = str(q_id)
        user_answer = user_answers.get(q_id_str)
        is_correct = False
        selected_choice = None
        short_answer_text = None

        try:
            if question.question_type == '객관식':
                if user_answer:
                    selected_choice = Choice.objects.get(pk=user_answer)
                    if selected_choice.is_correct:
                        is_correct = True
            
            elif question.question_type == '다중선택':
                # 로직 유지
                correct_ids = set(question.choice_set.filter(is_correct=True).values_list('id', flat=True))
                user_ids = set(user_answer if isinstance(user_answer, list) else [])
                if correct_ids and correct_ids == user_ids:
                    is_correct = True
                short_answer_text = ", ".join(map(str, user_ids))

            elif question.question_type.startswith('주관식'):
                possible = question.choice_set.filter(is_correct=True).values_list('choice_text', flat=True)
                user_text = str(user_answer).strip().lower() if user_answer else ""
                short_answer_text = user_answer
                for ans in possible:
                    if user_text == ans.strip().lower():
                        is_correct = True
                        break
        except Exception:
            pass

        if is_correct:
            correct_answers += 1
        
        # 상세 답안 저장
        UserAnswer.objects.create(
            test_result=test_result,
            question=question,
            selected_choice=selected_choice,
            short_answer_text=short_answer_text,
            is_correct=is_correct
        )

        results_data.append({
            'question': question,
            'selected_choice': selected_choice,
            'short_answer_text': short_answer_text,
            'is_correct': is_correct
        })
    
    # 점수 계산
    total_questions = len(question_ids)
    score = int((correct_answers / total_questions) * 100) if total_questions > 0 else 0
    is_pass = (score >= 80)
    
    # 결과 업데이트
    test_result.score = score
    test_result.is_pass = is_pass
    test_result.save()

    # 3회 탈락 시 잠금 로직
    if not is_pass:
        fail_count = TestResult.objects.filter(user=request.user, quiz=attempt.quiz, is_pass=False).count()
        if fail_count >= 3:
            request.user.profile.status = 'counseling'
            request.user.profile.save()
            messages.warning(request, "⛔ 3회 불합격하여 계정이 '면담 필요' 상태로 전환되었습니다.")

    # 뱃지 부여 (함수 호출 주석 처리 또는 import 필요)
    try:
        from .utils import award_badges # 필요시 import 위치 조정
        award_badges(request.user, test_result)
    except ImportError:
        pass # award_badges 함수가 없으면 패스

    # [핵심 수정] Attempt 상태를 반드시 '완료됨'으로 변경해야 재응시가 꼬이지 않음
    attempt.status = '완료됨'
    attempt.save()

    # 뱃지 알림 계산
    profile.refresh_from_db()
    badges_after = set(profile.badges.values_list('id', flat=True))
    new_badge_ids = badges_after - badges_before
    newly_awarded_badges = Badge.objects.filter(id__in=new_badge_ids)

    # 2회 불합격 시 메일 발송 로직 (기존 유지)
    if not is_pass:
        fail_cnt_mail = TestResult.objects.filter(user=request.user, quiz=attempt.quiz, is_pass=False).count()
        if fail_cnt_mail == 2:
            try:
                failed_attempts = TestResult.objects.filter(user=request.user, quiz=attempt.quiz, is_pass=False).order_by('completed_at')
                if failed_attempts.count() >= 2:
                    d_fmt = '%Y-%m-%d %H:%M'
                    d1 = f"{failed_attempts[0].completed_at.strftime(d_fmt)} / {failed_attempts[0].score}점"
                    d2 = f"{failed_attempts[1].completed_at.strftime(d_fmt)} / {failed_attempts[1].score}점"
                    
                    if hasattr(request.user, 'profile') and request.user.profile.pl and request.user.profile.pl.email:
                        pl = request.user.profile.pl
                        subject = f"[CBT 경고] 교육생 면담 요청: {profile.name}"
                        message = (
                            f"{pl.name}님,\n\n"
                            f"귀하의 담당 교육생 {profile.name}이(가) '{attempt.quiz.title}' 시험에서 2회 불합격했습니다.\n\n"
                            f"1차: {d1}\n2차: {d2}\n\n면담 및 지도가 필요합니다."
                        )
                        send_mail(subject, message, os.environ.get('EMAIL_HOST_USER'), [pl.email], fail_silently=True)
            except Exception as e:
                print(f"Mail Error: {e}")

    context = {
        'results_data': results_data,
        'score': score,
        'total_questions': total_questions,
        'correct_answers': correct_answers,
        'newly_awarded_badges': newly_awarded_badges,
        'test_result': test_result,
        'is_pass': is_pass,
    }

    # 세션 정리 (다음 시험을 위해)
    request.session.pop('quiz_questions', None)
    request.session.pop('user_answers', None)
    request.session.pop('attempt_id', None)
    request.session.pop('current_test_result_id', None)

    return render(request, 'quiz/quiz_results.html', context)


# =========================================================
# [3] 엑셀 파일 업로드 (upload_quiz)
# =========================================================
@login_required
def upload_quiz(request):
    if not request.user.is_staff:
        return redirect('quiz:index')

    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file).fillna('')
            
            error_count = 0
            success_count = 0

            for index, row in df.iterrows():
                q_type_excel = row['question_type']
                q_type_db = q_type_excel

                if q_type_excel == '주관식':
                    q_type_db = '주관식 (단일정답)'
                
                allowed_types = ['객관식', '다중선택', '주관식 (단일정답)', '주관식 (복수정답)']
                if q_type_db not in allowed_types:
                    messages.error(request, f"업로드 실패 (행 {index + 2}): 잘못된 유형입니다.")
                    error_count += 1
                    continue
                
                quiz, created = Quiz.objects.get_or_create(title=row['quiz_title'])
                
                # [핵심 수정 1] Question 생성 시 'quiz' 인자 제거
                question = Question.objects.create(
                    question_text=row['question_text'],
                    question_type=q_type_db,
                    difficulty=row['difficulty']
                )
                
                # [핵심 수정 2] 생성 후 M2M 관계 설정
                question.quizzes.add(quiz)

                if row['tags']:
                    tag_names = [tag.strip() for tag in str(row['tags']).split(',') if tag.strip()]
                    for tag_name in tag_names:
                        tag, created = Tag.objects.get_or_create(name=tag_name)
                        question.tags.add(tag)

                if q_type_db in ['객관식', '다중선택', '주관식 (복수정답)']:
                    for col in df.columns:
                        if str(col).startswith('correct_choice') and row[col]:
                            Choice.objects.create(question=question, choice_text=row[col], is_correct=True)
                    
                    if q_type_db in ['객관식', '다중선택']:
                        for col in df.columns:
                            if str(col).startswith('other_choice') and row[col]:
                                Choice.objects.create(question=question, choice_text=row[col], is_correct=False)
                
                elif q_type_db == '주관식 (단일정답)':
                    if row['correct_choice']:
                        Choice.objects.create(question=question, choice_text=row['correct_choice'], is_correct=True)

                success_count += 1
            
            if success_count > 0:
                messages.success(request, f"{success_count}개의 문제가 성공적으로 업로드되었습니다.")
            if error_count > 0:
                messages.warning(request, f"{error_count}개의 문제는 오류로 인해 건너뛰었습니다.")

        except Exception as e:
            messages.error(request, f"업로드 중 오류가 발생했습니다: {e}")

        return redirect('quiz:upload_quiz')

    return render(request, 'quiz/upload_quiz.html')

@login_required
def my_results_index(request):
    quizzes_taken = Quiz.objects.filter(testresult__user=request.user).distinct()
    context = {'quizzes_taken': quizzes_taken}
    return render(request, 'quiz/my_results_index.html', context)

@login_required
def my_results_by_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    all_results = TestResult.objects.filter(user=request.user, quiz=quiz).order_by('-completed_at')
    
    for result in all_results:
        newer_attempts_count = TestResult.objects.filter(
            user=request.user, quiz=result.quiz, completed_at__gt=result.completed_at
        ).count()
        total_attempts_for_quiz = TestResult.objects.filter(user=request.user, quiz=result.quiz).count()
        result.attempt_number = total_attempts_for_quiz - newer_attempts_count
        
    sorted_results = sorted(list(all_results), key=lambda r: r.completed_at, reverse=True)
    
    paginator = Paginator(sorted_results, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'quiz': quiz,
        'page_obj': page_obj
    }
    return render(request, 'quiz/my_results_list.html', context)

@login_required
def result_detail(request, result_id):
    """
    [수정] 결과 상세 및 오답 노트
    - 10점인데 만점이라고 뜨는 오류 수정
    - HTML이 'user_answers'를 찾든 'incorrect_answers'를 찾든 다 작동하게 처리
    """
    result = get_object_or_404(TestResult, pk=result_id)

    # 1. 권한 체크 (관리자 OK, 본인 OK)
    if not request.user.is_staff and result.user != request.user:
        raise Http404("권한이 없습니다.")

    # 2. [핵심] '오답'인 것만 따로 필터링해서 가져옵니다.
    # (점수가 10점이면 is_correct=False인 데이터가 분명히 존재함)
    only_incorrect = UserAnswer.objects.filter(test_result=result, is_correct=False).select_related('question')

    # 3. 전체 답안도 필요할 수 있으니 가져옵니다.
    all_answers = UserAnswer.objects.filter(test_result=result).select_related('question')

    context = {
        'result': result,
        'quiz': result.quiz,
        
        # ★ [치트키] HTML이 뭘 원할지 몰라서 다 넣어줍니다.
        # 만약 HTML이 "오답 목록"을 원하면 이게 들어갑니다.
        'incorrect_answers': only_incorrect, 
        
        # 만약 HTML이 "답안 목록"을 변수로 쓴다면, 오답만 보여주도록 이걸 연결합니다.
        # (상세 오답 확인 페이지이므로 틀린 것만 보여주는 게 맞습니다)
        'user_answers': only_incorrect, 
        
        # 혹시 전체 리스트가 필요하면 이걸 씁니다.
        'all_answers': all_answers,
    }
    
    return render(request, 'quiz/result_detail.html', context)

@login_required
def start_quiz(request, attempt_id):
    # 1. 본인 확인 및 정보 가져오기
    attempt = get_object_or_404(QuizAttempt, pk=attempt_id, user=request.user)
    quiz = attempt.quiz
    profile = request.user.profile

    # [Step 1] 3시간 유효시간 체크
    if attempt.requested_at and (timezone.now() > attempt.requested_at + timedelta(hours=3)):
        attempt.delete()  # 권한 회수
        messages.error(request, "⏳ 시험 응시 유효시간(3시간)이 초과되어 취소되었습니다. 다시 신청해주세요.")
        return redirect('quiz:index')

    # [Step 2] 계정 잠금(Lock) 및 3차 제한 검사
    if profile.status in ['counseling', 'dropout']:
        messages.error(request, "⛔ 계정이 잠겨있어 시험을 시작할 수 없습니다. 매니저 면담이 필요합니다.")
        return redirect('quiz:index')

    # 3차 탈락 여부 확인
    fail_count = TestResult.objects.filter(user=request.user, quiz=quiz, is_pass=False).count()
    if fail_count >= 3:
        if profile.status == 'attending':
            profile.status = 'counseling'
            profile.save()
        messages.error(request, f"⛔ '{quiz.title}' 시험에 3회 불합격하여 응시가 제한됩니다. 매니저 면담 후 해제 가능합니다.")
        return redirect('quiz:index')

    # [Step 3] 상태 체크
    if attempt.status == '완료됨':
        existing_result = TestResult.objects.filter(attempt=attempt).last()
        if existing_result:
            return redirect('quiz:result_detail', result_id=existing_result.id)
        else:
            messages.info(request, "이미 완료된 시험입니다.")
            return redirect('quiz:my_results_index')

    if attempt.status != '승인됨':
        messages.error(request, "아직 승인되지 않았거나 유효하지 않은 시험입니다.")
        return redirect('quiz:index')

    # [문제 출제 로직]
    # *중요* TestResult 미리 생성 (기존 로직 유지)
    test_result, created = TestResult.objects.get_or_create(
        user=request.user,
        quiz=quiz,
        attempt=attempt,
        defaults={'score': 0, 'is_pass': False}
    )

    final_questions = []

    # 1. [지정 문제 세트] 방식
    if quiz.generation_method == 'fixed' and quiz.exam_sheet:
        final_questions = list(quiz.exam_sheet.questions.all())
    
    # 2. [랜덤 출제] 방식
    else:
        loop_targets = []
        target_tags = None
        
        if quiz.generation_method == 'random_tag':
            target_tags = quiz.required_tags.all()
            if not target_tags.exists():
                messages.error(request, "설정된 태그가 없습니다. 관리자에게 문의하세요.")
                return redirect('quiz:index')
            loop_targets = list(target_tags)
        else:
            loop_targets = ['ALL'] 

        # 난이도별 분배 로직
        total_slots = 25 
        count = len(loop_targets)
        if count == 0: count = 1
            
        base_quota = total_slots // count
        remainder = total_slots % count

        for i, target in enumerate(loop_targets):
            this_quota = base_quota + (1 if i < remainder else 0)

            if target == 'ALL':
                base_qs = quiz.questions.all()
            else:
                base_qs = Question.objects.filter(tags=target)

            pool_h = list(base_qs.filter(difficulty='상'))
            pool_m = list(base_qs.filter(difficulty='중'))
            pool_l = list(base_qs.filter(difficulty='하'))
            
            random.shuffle(pool_h)
            random.shuffle(pool_m)
            random.shuffle(pool_l)

            target_h = int(this_quota * 0.32) 
            target_l = int(this_quota * 0.32) 
            target_m = this_quota - target_h - target_l 

            selected_in_loop = []

            # A. [상]
            picked_h = pool_h[:target_h]
            selected_in_loop.extend(picked_h)
            target_m += (target_h - len(picked_h))

            # B. [하]
            picked_l = pool_l[:target_l]
            selected_in_loop.extend(picked_l)
            target_m += (target_l - len(picked_l))

            # C. [중]
            picked_m = pool_m[:target_m]
            selected_in_loop.extend(picked_m)
            missing_m = target_m - len(picked_m)

            if missing_m > 0:
                remaining_l = pool_l[len(picked_l):]
                fallback_l = remaining_l[:missing_m]
                selected_in_loop.extend(fallback_l)
                
                still_missing = missing_m - len(fallback_l)
                if still_missing > 0:
                    remaining_h = pool_h[len(picked_h):]
                    fallback_h = remaining_h[:still_missing]
                    selected_in_loop.extend(fallback_h)
            
            final_questions.extend(selected_in_loop)
            
        # 25개 미달 시 채우기
        if len(final_questions) < 25:
            needed = 25 - len(final_questions)
            current_ids = [q.id for q in final_questions]
            
            if quiz.generation_method == 'random_tag' and target_tags:
                extra_pool = list(Question.objects.filter(tags__in=target_tags).exclude(id__in=current_ids).distinct())
            else:
                extra_pool = list(quiz.questions.exclude(id__in=current_ids))
            
            random.shuffle(extra_pool)
            final_questions.extend(extra_pool[:needed])

    random.shuffle(final_questions)
    
    if not final_questions:
        messages.error(request, "출제할 문제가 없습니다.")
        return redirect('quiz:index')

    # ----------------------------------------------------------
    # [세션 저장] - 오류 수정 핵심 부분
    # ----------------------------------------------------------
    # 기존 세션 정리
    keys_to_clear = ['quiz_questions', 'attempt_id', 'user_answers', 'current_test_result_id']
    for key in keys_to_clear:
        if key in request.session:
            del request.session[key]

    # [수정] 복잡한 동적 키 대신 고정 키 사용 (take_quiz와 일치시킴)
    request.session['quiz_questions'] = [q.id for q in final_questions]
    request.session['attempt_id'] = attempt.id
    request.session['current_test_result_id'] = test_result.id # 필요시 사용
    request.session['user_answers'] = {}

    return redirect('quiz:take_quiz', quiz_id=quiz.id)

@login_required
def submit_quiz(request):
    attempt_id = request.session.get('attempt_id')
    if attempt_id:
        attempt = QuizAttempt.objects.get(pk=attempt_id)
        if attempt.status != '완료됨':
            attempt.status = '완료됨'
            attempt.save()
    return redirect('quiz:quiz_results')

@login_required
def my_incorrect_answers_index(request):
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다. (관리자 전용)")
        return redirect('quiz:index') # 또는 'dashboard'
    
    incorrect_answers = UserAnswer.objects.filter(test_result__user=request.user, is_correct=False)
    quizzes_with_incorrects = Quiz.objects.filter(question__useranswer__in=incorrect_answers).distinct()
    context = {'quizzes_with_incorrects': quizzes_with_incorrects}
    return render(request, 'quiz/my_incorrect_answers_index.html', context)

@login_required
def my_incorrect_answers_by_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    incorrect_answers = UserAnswer.objects.filter(
        test_result__user=request.user, 
        question__quiz=quiz,
        is_correct=False
    )
    incorrect_question_ids = incorrect_answers.values_list('question', flat=True).distinct()
    incorrect_questions = Question.objects.filter(pk__in=incorrect_question_ids)
    context = {'quiz': quiz, 'incorrect_questions': incorrect_questions}
    return render(request, 'quiz/incorrect_answers_list.html', context)

@login_required
def approve_attempt(request, attempt_id):
    # 1. 관리자 권한 확인
    if not request.user.is_staff:
        messages.error(request, "권한이 없습니다.")
        return redirect('quiz:dashboard')

    attempt = get_object_or_404(QuizAttempt, pk=attempt_id)
    
    # [수정] 여기서 target_profile을 정의해줘야 에러가 안 납니다!
    target_profile = attempt.user.profile
    
    # 2. [핵심] 매니저의 공정과 교육생의 공정 비교 (최고 관리자는 제외)
    # 이제 target_profile 변수가 정의되었으므로 에러가 나지 않습니다.
    if not is_process_manager(request.user, target_profile):
        messages.error(request, f"🚫 본인 담당 공정({target_profile.process})의 교육생만 승인할 수 있습니다.")
        return redirect('quiz:manager_exam_requests')

    # 3. 승인 처리
    attempt.status = '승인됨'
    attempt.save()
    messages.success(request, f"{target_profile.name}님의 시험 요청을 승인했습니다.")
    
    return redirect('quiz:manager_exam_requests')

@login_required
def dashboard(request):
    """
    [Final Upgrade] 교육생 성적 관리 대시보드
    - 3단 분할 점수판 (1차/2차/3차)
    - 회사/공정/기수 다중 필터링 적용
    - 안전/기타 과목 제외 및 정렬 로직
    - ★ 핵심 추가: 본인 공정 외 '권한 승인된 타 공정' 조회 기능
    """
    user = request.user
    
    # 1. 권한 체크 (스태프이거나, 매니저/PL 프로필 보유자)
    if not (user.is_staff or (hasattr(user, 'profile') and (user.profile.is_manager or user.profile.is_pl))):
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    # 2. 필터링 옵션 데이터 조회
    companies = Company.objects.all()
    processes = Process.objects.all()
    all_cohorts = Cohort.objects.all().order_by('-start_date')

    # 3. 검색 조건 받기
    sel_company = request.GET.get('company')
    sel_process = request.GET.get('process')
    sel_cohort = request.GET.get('cohort')
    search_query = request.GET.get('q', '')

    # -----------------------------------------------------------
    # [핵심 로직] 조회 가능한 공정인지 권한 검증 (Permission Check)
    # -----------------------------------------------------------
    my_process = None
    if hasattr(user, 'profile'):
        my_process = user.profile.process

    # 실제로 데이터를 조회할 기준 공정 (None이면 전체-관리자용)
    viewing_process = None 

    if user.is_superuser:
        # 관리자는 선택한 공정 자유롭게 조회
        if sel_process:
            viewing_process = Process.objects.get(id=sel_process)
    else:
        # 매니저는 기본적으로 내 공정
        viewing_process = my_process
        
        # 만약 필터에서 '다른 공정'을 선택했다면? -> 티켓 확인
        if sel_process and str(sel_process) != str(my_process.id if my_process else ''):
            # 승인된 티켓이 있는지 확인
            has_ticket = ProcessAccessRequest.objects.filter(
                requester=user,
                status='approved',
                expires_at__gte=timezone.now()
            ).filter(
                Q(target_process_id=sel_process) | Q(target_process__isnull=True)
            ).exists()
            
            if has_ticket:
                viewing_process = Process.objects.get(id=sel_process)
            else:
                # 티켓 없으면 내 공정으로 강제 리셋 + 경고
                messages.warning(request, "⛔ 해당 공정 조회 권한이 없습니다. (권한 요청 필요)")
                # sel_process 값도 리셋해야 HTML selectbox가 제대로 표시됨
                sel_process = str(my_process.id) if my_process else ''
                viewing_process = my_process

    # -----------------------------------------------------------
    # 4. 표시할 시험 헤더 (공통 + 조회 중인 공정)
    # -----------------------------------------------------------
    header_quizzes = Quiz.objects.filter(category__in=['common', 'safety', 'etc'])

    # 선택된 공정이 있다면 해당 공정 과목도 추가로 합칩니다.
    if viewing_process:
        header_quizzes = header_quizzes | Quiz.objects.filter(related_process=viewing_process)
    
    # 카테고리 순서대로 예쁘게 정렬 (1.공통 -> 2.공정 -> 3.안전 -> 4.기타)
    header_quizzes = header_quizzes.distinct().order_by(
        Case(
            When(category='common', then=Value(1)),
            When(category='process', then=Value(2)),
            When(category='safety', then=Value(3)),
            When(category='etc', then=Value(4)),
            default=Value(5),
            output_field=IntegerField()
        ), 'title'
    )

    # -----------------------------------------------------------
    # 5. 기수별 데이터 루프
    # -----------------------------------------------------------
    if sel_cohort:
        target_cohorts = all_cohorts.filter(id=sel_cohort)
    else:
        target_cohorts = all_cohorts
    
    dashboard_data = []
    total_risk_count = 0

    for cohort in target_cohorts:
        # 기본 쿼리셋 (스태프 제외)
        profiles = Profile.objects.filter(cohort=cohort, user__is_staff=False).select_related('company', 'process')
        
        # [필터링] viewing_process(검증된 공정) 기준
        if viewing_process: 
            profiles = profiles.filter(process=viewing_process)
            
        if sel_company:
            profiles = profiles.filter(company_id=sel_company)
        
        if search_query:
            profiles = profiles.filter(Q(name__icontains=search_query) | Q(employee_id__icontains=search_query))

        # 데이터가 없으면 건너뜀
        if not profiles.exists():
            continue

        student_data = []
        cohort_1st_scores = [] 

        for p in profiles:
            student_scores_map = [] 
            
            total_1st_score = 0
            count_1st = 0

            for quiz in header_quizzes:
                # 해당 퀴즈의 기록 (시간순, 최대 3개)
                attempts = TestResult.objects.filter(user=p.user, quiz=quiz).order_by('completed_at')[:3]
                
                attempt_list = [None, None, None]
                
                if attempts.exists():
                    # 1. 1차 점수 집계 (평균용)
                    s1 = attempts[0].score
                    
                    # ★★★ [핵심 수정] '공통(common)'과 '공정(process)' 과목만 평균에 합산합니다! ★★★
                    # (안전, 기타 과목은 화면에는 나오지만 1차 평균 점수에는 영향을 주지 않음)
                    if quiz.category in ['common', 'process']:
                        total_1st_score += s1
                        count_1st += 1
                    
                    # 2. 1~3차 상세 데이터 채우기 (기존 유지)
                    for i, att in enumerate(attempts):
                        attempt_list[i] = {
                            'score': att.score,
                            'is_pass': att.is_pass
                        }
                
                # ★ 템플릿에서 파스텔 색상을 입히기 위해 카테고리 정보도 같이 넘겨줍니다.
                student_scores_map.append({
                    'quiz_id': quiz.id,
                    'category': quiz.category, # <-- 이 줄 추가!
                    'attempts': attempt_list 
                })

            # 평균 계산 (1차 점수 기준)
            avg_1st = round(total_1st_score / count_1st, 1) if count_1st > 0 else 0
            if count_1st > 0:
                cohort_1st_scores.append(avg_1st)

            ## 위험군 판별
            is_risk = False
            
            # ★ [핵심 수정] 퇴소(dropout)나 수료(completed) 상태인 학생은 집중 관리 대상에서 제외!
            if p.status not in ['dropout', 'completed']:
                if (avg_1st > 0 and avg_1st < 60) or p.warning_count >= 2:
                    is_risk = True
                    total_risk_count += 1

            student_data.append({
                'profile': p,
                'scores_map': student_scores_map,
                'avg_score': avg_1st,
                'is_risk': is_risk,
                'warning_count': p.warning_count
            })

        # 기수 평균 계산
        cohort_avg = round(sum(cohort_1st_scores) / len(cohort_1st_scores), 1) if cohort_1st_scores else 0
        
        # 최종 데이터 추가
        dashboard_data.append({
            'cohort': cohort,
            'students': student_data,
            'cohort_avg': cohort_avg,
            'risk_count': sum(1 for s in student_data if s['is_risk']),
            'total_students': len(student_data)
        })

    # 필터 적용 여부 확인 (UI 표시용)
    is_filtered = True if (sel_company or sel_process or sel_cohort or search_query) else False

    # 6. 컨텍스트 구성
    context = {
        'process_name': viewing_process.name if viewing_process else "전체 공정",
        'header_quizzes': header_quizzes,
        'dashboard_data': dashboard_data,
        'total_risk_count': total_risk_count,
        
        # 필터링 옵션 전달
        'companies': companies,
        'processes': processes,
        'cohorts': all_cohorts,
        
        # 현재 선택된 필터값 유지
        'sel_company': int(sel_company) if sel_company else '',
        'sel_process': int(sel_process) if sel_process else '', # HTML selectbox에서 selected 처리용
        'sel_cohort': int(sel_cohort) if sel_cohort else '',
        'search_query': search_query,
        'is_filtered': is_filtered,
    }
    
    return render(request, 'quiz/manager/new_dashboard.html', context)

    # 6. 컨텍스트 구성
    context = {
        'process_name': target_process.name if target_process else "전체 공정",
        'header_quizzes': header_quizzes,
        'dashboard_data': dashboard_data,
        'total_risk_count': total_risk_count,
        
        
        # 필터링 옵션 전달 (검색바 유지용)
        'companies': companies,
        'processes': processes,
        'cohorts': all_cohorts,
        'sel_company': int(sel_company) if sel_company else '',
        'sel_process': int(sel_process) if sel_process else '',
        'sel_cohort': int(sel_cohort) if sel_cohort else '',
        'search_query': search_query,

        'is_filtered': is_filtered,
    }
    
    return render(request, 'quiz/manager/new_dashboard.html', context)

@login_required
def student_analysis_detail(request, profile_id):
    """
    [Upgrade Final] 교육생 심층 분석
    - 기능 1: 태그별 정답률 분석 (Radar Chart)
    - 기능 2: AI 맞춤형 학습 처방 (Prescription)
    - 기능 3: 시험 이력 리스트 (오답노트 바로가기 연결용)
    """
    if not request.user.is_staff:
        return redirect('quiz:index')
    
    target_profile = get_object_or_404(Profile, pk=profile_id)
    
    # 1. [보안] 내 공정인지 확인 (관리자는 프리패스)
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.process != target_profile.process:
            messages.error(request, "타 공정 교육생 정보는 조회할 수 없습니다.")
            return redirect('quiz:manager_dashboard')

    # 2. [태그 분석] Radar Chart용 데이터 생성
    # 해당 학생이 푼 모든 문제의 정답 여부를 태그별로 집계
    user_answers = UserAnswer.objects.filter(test_result__user=target_profile.user).select_related('question')
    tag_stats = {} 

    for ua in user_answers:
        for tag in ua.question.tags.all():
            if tag.name not in tag_stats:
                tag_stats[tag.name] = {'total': 0, 'correct': 0}
            
            tag_stats[tag.name]['total'] += 1
            if ua.is_correct:
                tag_stats[tag.name]['correct'] += 1
    
    analysis_data = []
    for tag_name, stat in tag_stats.items():
        if stat['total'] > 0:
            rate = (stat['correct'] / stat['total']) * 100
        else:
            rate = 0
        analysis_data.append({
            'tag': tag_name, 
            'rate': round(rate, 1), 
            'count': stat['total']
        })
    
    # 정답률 낮은 순(취약점 우선) 정렬
    analysis_data.sort(key=lambda x: x['rate'])

    # 3. [AI 학습 처방] 맞춤형 코멘트 생성
    weak_tags = [item['tag'] for item in analysis_data if item['rate'] < 60]
    prescription = "취약점이 발견되지 않았습니다. 우수한 성취도를 보이고 있습니다."
    
    if weak_tags:
        tags_str = ', '.join(weak_tags[:3]) # 상위 3개만 표시
        prescription = f"'{tags_str}' 태그의 정답률이 저조합니다. 관련 문제 은행 풀이를 권장합니다."

    # 4. [시험 이력] 최근 응시한 시험 리스트 (오답노트 바로가기용)
    exam_history = TestResult.objects.filter(
        user=target_profile.user
    ).select_related('quiz').order_by('-completed_at')

    context = {
        'profile': target_profile,
        'tag_analysis': analysis_data,   # 차트/표 데이터
        'prescription': prescription,    # AI 처방 멘트
        'exam_history': exam_history,    # 시험 리스트 (오답 버튼용)
    }
    return render(request, 'quiz/manager/student_analysis.html', context)

@login_required
def personal_dashboard(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    summary_data = []
    quizzes_taken = Quiz.objects.filter(testresult__user=request.user).distinct()

    for quiz in quizzes_taken:
        results_for_quiz = TestResult.objects.filter(user=request.user, quiz=quiz)
        first_attempt = results_for_quiz.order_by('completed_at').first()
        first_score = first_attempt.score if first_attempt else None
        avg_score = results_for_quiz.aggregate(Avg('score'))['score__avg']
        max_score = results_for_quiz.aggregate(Max('score'))['score__max']
        attempts = results_for_quiz.count()

        summary_data.append({
            'title': quiz.title,
            'first_score': first_score,
            'avg_score': avg_score,
            'max_score': max_score,
            'attempts': attempts,
        })

    total_attempts = TestResult.objects.filter(user=request.user).count()
    overall_average_score = TestResult.objects.filter(user=request.user).aggregate(Avg('score'))['score__avg']

    context = {
        'total_attempts': total_attempts,
        'overall_average_score': overall_average_score,
        'summary_data': summary_data,
        'user_badges': profile.badges.all(), 
    }
    return render(request, 'quiz/personal_dashboard.html', context)

@login_required
def start_group_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    attempt = QuizAttempt.objects.create(
        user=request.user,
        quiz=quiz,
        status=QuizAttempt.Status.APPROVED,
        assignment_type=QuizAttempt.AssignmentType.GROUP
    )
    return redirect('quiz:start_quiz', attempt_id=attempt.id)

@login_required
def export_student_data(request):
    """
    교육생의 종합 데이터(성적, 평가, 특이사항, 근태)를 엑셀로 생성하여 
    브라우저에서 바로 다운로드하는 뷰
    """
    if not request.user.is_staff:
        return redirect('quiz:index')

    target_process_id = request.GET.get('process_id')
    
    # 1. 대상 프로필 조회 (성능 최적화)
    profiles = Profile.objects.select_related(
        'user', 'cohort', 'company', 'process', 'pl', 'final_assessment'
    ).prefetch_related(
        'user__testresult_set', 
        'badges', 
        'managerevaluation_set__selected_items', 
        'student_logs', 
        'dailyschedule_set__work_type'
    ).order_by('cohort__start_date', 'user__username')

    # 2. 권한 필터링 (기존의 티켓 로직 복원)
    my_process = None
    if hasattr(request.user, 'profile') and request.user.profile.process:
        my_process = request.user.profile.process

    if request.user.is_superuser:
        # 관리자는 선택한 공정 또는 전체 다운로드 가능
        if target_process_id and target_process_id != 'ALL':
            profiles = profiles.filter(process_id=target_process_id)
    else:
        # 매니저는 본인 공정만 가능 (또는 티켓 보유 시)
        if not my_process:
            messages.error(request, "본인 공정 정보가 없어 작업을 수행할 수 없습니다.")
            return redirect('quiz:dashboard')

        if target_process_id == 'ALL':
            # 전체 다운로드 권한 확인 (Global Ticket)
            global_ticket = ProcessAccessRequest.objects.filter(
                requester=request.user, target_process__isnull=True, status='approved'
            ).first()
            
            if global_ticket:
                global_ticket.status = 'expired' # 티켓 사용 처리
                global_ticket.save()
                # 필터링 없이 전체 profiles 다운로드
            else:
                messages.error(request, "⛔ 전체 데이터 다운로드 권한이 없습니다.")
                return redirect('quiz:dashboard')

        elif not target_process_id or str(target_process_id) == str(my_process.id):
            # 본인 공정 다운로드 (기본)
            profiles = profiles.filter(process=my_process)
            
        else:
            # 타 공정 티켓 확인 (Specific Ticket)
            access_ticket = ProcessAccessRequest.objects.filter(
                requester=request.user, target_process_id=target_process_id, status='approved'
            ).first()
            
            if access_ticket:
                profiles = profiles.filter(process_id=target_process_id)
                access_ticket.status = 'expired' # 티켓 사용 처리
                access_ticket.save()
            else:
                messages.error(request, "⛔ 해당 공정 접근 권한이 없습니다.")
                return redirect('quiz:dashboard')

    # 3. 엑셀 데이터 생성
    all_quizzes = Quiz.objects.all().order_by('title')
    data_list = []

    for profile in profiles:
        # 기본 정보
        row = {
            'ID': profile.user.username, 
            '이름': profile.name, 
            '사번': profile.employee_id,
            '기수': profile.cohort.name if profile.cohort else '-',
            '공정': profile.process.name if profile.process else '-',
            '상태': profile.get_status_display(),
            '누적 경고': profile.warning_count,
        }

        # 시험 점수 (1~3차)
        results = sorted(list(profile.user.testresult_set.all()), key=lambda x: x.completed_at)
        quiz_map = {}
        for r in results:
            if r.quiz_id not in quiz_map: quiz_map[r.quiz_id] = []
            quiz_map[r.quiz_id].append(r.score)
            
        for q in all_quizzes:
            atts = quiz_map.get(q.id, [])
            row[f"[{q.title}] 1차"] = atts[0] if len(atts) > 0 else '-'
            row[f"[{q.title}] 2차"] = atts[1] if len(atts) > 1 else '-'
            row[f"[{q.title}] 3차"] = atts[2] if len(atts) > 2 else '-'

        # 종합 평가
        fa = getattr(profile, 'final_assessment', None)
        row.update({
            '시험 평균 (85%)': fa.exam_avg_score if fa else 0,
            '실습 점수 (5%)': fa.practice_score if fa else 0,
            '수학태도 점수 (10%)': fa.attitude_score if fa else 0, 
            '최종 환산 점수': fa.final_score if fa else '-',
            '석차': fa.rank if fa else '-',
            '매니저 종합 의견': fa.manager_comment if fa else '-',
        })

        # 체크리스트
        last_eval = profile.managerevaluation_set.last()
        row['체크리스트'] = "\n".join([i.description for i in last_eval.selected_items.all()]) if last_eval else ""

        # 특이사항/경고 이력
        logs = profile.student_logs.all().order_by('created_at')
        log_txt = ""
        for l in logs:
            log_txt += f"[{l.created_at.date()}] {l.get_log_type_display()}: {l.reason}"
            if l.action_taken: log_txt += f" (조치: {l.action_taken})"
            log_txt += "\n"
        row['특이사항 이력'] = log_txt

        # 근태 요약
        schedules = profile.dailyschedule_set.all()
        w = schedules.filter(work_type__deduction=0).count()
        l = schedules.filter(work_type__deduction=1.0).count()
        row['근태'] = f"출근:{w} / 연차:{l}"
        
        data_list.append(row)

    # 4. 파일 생성 및 다운로드 (Direct Download)
    try:
        if not data_list:
            messages.warning(request, "다운로드할 데이터가 없습니다.")
            return redirect('quiz:manager_dashboard')

        df = pd.DataFrame(data_list)
        excel_file = BytesIO()

        # XlsxWriter 엔진 사용 (서식 적용)
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='종합_데이터')
            
            workbook = writer.book
            worksheet = writer.sheets['종합_데이터']
            
            # 셀 서식 (줄바꿈 및 정렬)
            format_wrap = workbook.add_format({'text_wrap': True, 'valign': 'top'})
            
            # 컬럼 너비 자동 조정
            for idx, col in enumerate(df.columns):
                if col in ['특이사항 이력', '체크리스트', '매니저의견']:
                    worksheet.set_column(idx, idx, 50, format_wrap)
                else:
                    worksheet.set_column(idx, idx, 15)
        
        # 파일 포인터 초기화
        excel_file.seek(0)

        # 파일명 설정 (한글 깨짐 방지)
        target_name = "전체"
        if target_process_id and target_process_id != 'ALL':
            try: target_name = Process.objects.get(pk=target_process_id).name
            except: pass
        elif my_process and not request.user.is_superuser:
            target_name = my_process.name

        filename = f"{target_name}_FullData_{timezone.now().strftime('%Y%m%d')}.xlsx"
        encoded_filename = urllib.parse.quote(filename)

        # HTTP 응답 생성 (다운로드 트리거)
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        messages.success(request, "엑셀 다운로드가 시작되었습니다.")
        return response

    except Exception as e:
        messages.error(request, f"엑셀 생성 중 오류 발생: {str(e)}")
        return redirect('quiz:manager_dashboard')

    # 3. 엑셀 데이터 생성 시작
    all_quizzes = Quiz.objects.all().order_by('title')
    data_list = []

    for profile in profiles:
        # (A) 기본 정보
        row_data = {
            '사용자 ID': profile.user.username,
            '이름': profile.name,
            '이메일': profile.user.email,
            '사번': profile.employee_id,
            '기수': profile.cohort.name if profile.cohort else '-',
            '소속 회사': profile.company.name if profile.company else '-',
            '공정': profile.process.name if profile.process else '-',
            '라인': profile.line if profile.line else '-',
            '담당 PL': profile.pl.name if profile.pl else '-',
            '상태': profile.get_status_display(),
        }

        # (B) 시험 점수 (1차, 2차, 3차)
        test_results = sorted(list(profile.user.testresult_set.all()), key=lambda x: x.completed_at)
        quiz_map = {}
        for res in test_results:
            if res.quiz_id not in quiz_map: quiz_map[res.quiz_id] = []
            quiz_map[res.quiz_id].append(res.score)
        
        for quiz in all_quizzes:
            attempts = quiz_map.get(quiz.id, [])
            row_data[f"[{quiz.title}] 1차"] = attempts[0] if len(attempts) > 0 else '-'
            row_data[f"[{quiz.title}] 2차"] = attempts[1] if len(attempts) > 1 else '-'
            row_data[f"[{quiz.title}] 3차"] = attempts[2] if len(attempts) > 2 else '-'

        # (C) 종합 평가 데이터 (FinalAssessment)
        fa = getattr(profile, 'final_assessment', None)
        row_data.update({
            '시험 평균': fa.exam_avg_score if fa else 0,
            '실습 점수': fa.practice_score if fa else 0,
            '노트 점수': fa.note_score if fa else 0,
            '태도 점수': fa.attitude_score if fa else 0,
            '최종 환산 점수': fa.final_score if fa else '-',
            '석차': fa.rank if fa else '-',
            '매니저 종합 의견': fa.manager_comment if fa else '-',
        })

        # (D) 체크리스트 평가 (ManagerEvaluation)
        # 가장 최근 평가서 1개를 가져옴
        last_eval = profile.managerevaluation_set.order_by('-created_at').first()
        checklist_str = ""
        if last_eval:
            items = last_eval.selected_items.all()
            # 엑셀 셀 하나에 줄바꿈으로 넣기 위해 join 사용
            checklist_str = "\n".join([f"[{'긍정' if item.is_positive else '부정'}] {item.description}" for item in items])
        row_data['체크리스트 평가'] = checklist_str

        # (E) 특이사항/경고 이력 (StudentLog)
        logs = profile.student_logs.all().order_by('created_at')
        log_str = ""
        for log in logs:
            log_str += f"[{log.created_at.strftime('%Y-%m-%d')}] {log.get_log_type_display()}: {log.reason}\n"
        row_data['특이사항/경고 이력'] = log_str

        # (F) 근태 요약 (DailySchedule)
        # WorkType의 deduction(차감) 값을 기준으로 카운트
        schedules = profile.dailyschedule_set.all()
        
        work_cnt = schedules.filter(work_type__deduction=0).count() # 정상출근
        leave_cnt = schedules.filter(work_type__deduction=1.0).count() # 연차
        half_cnt = schedules.filter(work_type__deduction=0.5).count() # 반차
        
        row_data['근태 요약'] = f"출근:{work_cnt} / 연차:{leave_cnt} / 반차:{half_cnt}"
        
        # (G) 뱃지 정보
        badge_count = profile.badges.count()
        badge_list = ", ".join([b.name for b in profile.badges.all()])
        row_data['획득 뱃지 수'] = badge_count
        row_data['뱃지 목록'] = badge_list

        data_list.append(row_data)

    # 4. 엑셀 파일 생성 및 발송
    try:
        if not data_list:
            messages.warning(request, "다운로드할 데이터가 없습니다.")
            return redirect('quiz:manager_dashboard')

        df = pd.DataFrame(data_list)
        excel_file = BytesIO()
        
        # XlsxWriter 엔진 사용 (서식 적용을 위해)
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='종합_데이터')
            
            workbook = writer.book
            worksheet = writer.sheets['종합_데이터']
            
            # 셀 줄바꿈 포맷 (특이사항 등이 길어질 수 있으므로)
            format_wrap = workbook.add_format({'text_wrap': True, 'valign': 'top'})
            
            # 컬럼 너비 자동 조정 (대략적으로 설정)
            for idx, col in enumerate(df.columns):
                if col in ['특이사항/경고 이력', '체크리스트 평가', '매니저 종합 의견']:
                    worksheet.set_column(idx, idx, 50, format_wrap) # 너비 50 & 줄바꿈
                elif col in ['사용자 ID', '이름', '이메일']:
                    worksheet.set_column(idx, idx, 20)
                else:
                    worksheet.set_column(idx, idx, 12)

        excel_file.seek(0)

        # 파일명 설정
        target_name = "전체"
        if target_process_id and target_process_id != 'ALL':
            try: target_name = Process.objects.get(pk=target_process_id).name
            except: pass
        elif my_process and not request.user.is_superuser:
            target_name = my_process.name

        subject = f"[보안] {request.user.profile.name}님 요청 데이터 ({target_name})"
        body = (
            f"요청하신 교육생 데이터입니다.\n"
            f"요청자: {request.user.profile.name}\n"
            f"대상 공정: {target_name}\n\n"
            f"* 포함 내역: 기본정보, 시험성적(1~3차), 종합평가(점수/석차), 체크리스트, 특이사항/경고 이력, 근태 요약, 뱃지 현황"
        )
        
        email = EmailMessage(
            subject, body, settings.EMAIL_HOST_USER, [request.user.email]
        )
        filename = f"{target_name}_FullData_{timezone.now().strftime('%Y%m%d')}.xlsx"
        email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()
        
        messages.success(request, f"✅ 상세 데이터가 포함된 엑셀 파일이 '{request.user.email}'로 발송되었습니다.")

    except Exception as e:
        print(f"Mail Error: {e}")
        messages.error(request, f"메일 발송 중 오류가 발생했습니다: {str(e)}")

    return redirect('quiz:manager_dashboard')



def get_pl_dashboard_data(pl_user):
    """
    담당 파트장(PL)의 교육생 명단을 가져와 가로형(피벗 테이블) 성적 데이터를 생성합니다.
    """
    try:
        part_leader_obj = PartLeader.objects.get(email=pl_user.email)
    except PartLeader.DoesNotExist:
        return []

    trainees = Profile.objects.filter(
        pl=part_leader_obj,
        user__is_superuser=False, # 슈퍼유저 제외
        is_manager=False,         # 매니저 제외
        is_pl=False               # PL 제외
    ).order_by('name').select_related('user', 'cohort', 'process')
    all_quizzes = Quiz.objects.all().order_by('title')
    
    data_list = []
    
    for trainee_profile in trainees:
        row = {
            'name': trainee_profile.name,
            'status': trainee_profile.get_status_display(),
            'cohort': trainee_profile.cohort.name if trainee_profile.cohort else '-',
        }
        
        results = trainee_profile.user.testresult_set.all().order_by('completed_at')
        
        for quiz in all_quizzes:
            quiz_attempts = results.filter(quiz=quiz)
            
            # 1차, 2차, 3차 점수 추출 (Horizontal Columns)
            score_1 = quiz_attempts[0].score if quiz_attempts.count() >= 1 else '-'
            score_2 = quiz_attempts[1].score if quiz_attempts.count() >= 2 else '-'
            score_3 = quiz_attempts[2].score if quiz_attempts.count() >= 3 else '-'
            
            row[f'{quiz.title}_1차'] = score_1
            row[f'{quiz.title}_2차'] = score_2
            row[f'{quiz.title}_3차'] = score_3
            
        data_list.append(row)
        
    return data_list

def process_expired_dropouts():
    """승인된 퇴사 요청 중 예정일이 어제 날짜인 인원들을 일괄 퇴소 처리"""
    from accounts.models import DropOutRequest
    from django.utils import timezone
    today = timezone.now().date()
    
    expired_requests = DropOutRequest.objects.filter(
        status='approved',
        drop_date__lt=today
    ).exclude(trainee__status='dropout')
    
    for req in expired_requests:
        trainee = req.trainee
        trainee.status = 'dropout'
        trainee.save()

@login_required
def manager_dashboard(request):
    process_expired_dropouts() 
    
    from accounts.models import Profile, Cohort, Process, Company, ProcessAccessRequest
    from quiz.models import QuizAttempt, StudentLog 
    from django.db.models import Q, Case, When, Value, IntegerField # Q 등 필수 임포트 확인
    
    user = request.user
    user_profile = getattr(user, 'profile', None)

    # 0. 권한 체크
    if not (user.is_staff or (user_profile and (user_profile.is_manager or user_profile.is_pl))):
        from django.contrib import messages
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    # 카운트 1~6 (기존 로직 유지)
    signup_pending_count = Profile.objects.filter(is_approved=False).count()
    
    exam_q = Q(status='대기중')
    if not user.is_superuser and user_profile and user_profile.process:
        exam_q &= Q(user__profile__process=user_profile.process)
    exam_pending_count = QuizAttempt.objects.filter(exam_q).count()

    risk_q = Q(status='counseling')
    if not user.is_superuser and user_profile and user_profile.process:
        risk_q &= Q(process=user_profile.process)
    risk_count = Profile.objects.filter(risk_q).count()

    access_req_count = 0
    if user.is_superuser:
        access_req_count = ProcessAccessRequest.objects.filter(status='pending').count()
    elif user_profile and user_profile.process:
        access_req_count = ProcessAccessRequest.objects.filter(
            target_process=user_profile.process, status='pending'
        ).count()

    schedule_pending_count = 0
    if user_profile:
        try:
            from attendance.models import ScheduleRequest
            if user.is_superuser:
                schedule_pending_count = ScheduleRequest.objects.filter(status='pending').count()
            elif user_profile.is_manager:
                schedule_pending_count = ScheduleRequest.objects.filter(
                    requester__process=user_profile.process, status='pending'
                ).exclude(requester=user_profile).count()
        except ImportError:
            pass

    counseling_q = Q(log_type='counseling', is_resolved=False)
    if not user.is_superuser and user_profile and user_profile.process:
        counseling_q &= Q(profile__process=user_profile.process)
    counseling_count = StudentLog.objects.filter(counseling_q).count()

    # --- 드릴다운 필터용 데이터 추출 및 전달 ---
    active_profiles = Profile.objects.filter(
        is_manager=False, is_pl=False
    ).exclude(status__in=['completed', 'dropout']).exclude(user__is_staff=True)
    
    filter_cohorts = Cohort.objects.filter(
    profile__in=active_profiles,
    is_manual_exam_allowed=True  # 관리자가 허용한 기수만!
).distinct()
    filter_processes = Process.objects.filter(profile__in=active_profiles).distinct()
    filter_companies = Company.objects.filter(profile__in=active_profiles).distinct()

    active_trainees = active_profiles.annotate(
        is_priority=Case(
            When(cohort=user_profile.cohort if user_profile else None, process=user_profile.process if user_profile else None, then=Value(2)),
            When(cohort=user_profile.cohort if user_profile else None, then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        )
    ).order_by('-is_priority', 'name')

    # ========================================================
    # ★ [수정됨] 스마트 비서: 공정별 위기 현황 데이터 생성 (통합 징계 기준 적용)
    # ========================================================
    process_summary = []
    
    # 1. 대상 공정 설정 (최고 관리자는 전체, 일반 매니저는 본인 공정만)
    if user.is_superuser:
        target_processes = filter_processes 
    elif user_profile and user_profile.process:
        target_processes = [user_profile.process]
    else:
        target_processes = []

    # 2. 공정별로 학생들을 위험/주의/정상으로 정확하게 분류
    for p in target_processes:
        p_students = active_trainees.filter(process=p)
        
        danger_list, caution_list, normal_list = [], [], []
        
        for student in p_students:
            # ★ 1. 시험 과락 횟수 및 과목명 추출
            passed_quiz_ids = TestResult.objects.filter(
                user=student.user, is_pass=True
            ).values_list('quiz_id', flat=True)

            # (B) 불합격 기록을 찾되, "이미 합격한 시험"의 불합격 기록은 완벽하게 무시(exclude)합니다!
            recent_fails = TestResult.objects.filter(
                user=student.user, is_pass=False
            ).exclude(
                quiz_id__in=passed_quiz_ids # 합격한 시험은 불합격 카운트에서 뺌!
            ).values('quiz__title').annotate(fail_cnt=Count('id'))
        
            is_danger_exam = False
            is_caution_exam = False
            fail_reasons = [] # 어떤 시험을 떨어졌는지 텍스트 저장용
            
            for fail in recent_fails:
                title = fail['quiz__title']
                cnt = fail['fail_cnt']
                # 긴 제목 줄이기 (예: [공통] 반도체 기초 -> 반도체 기초)
                short_title = title.split(']')[-1].strip() if ']' in title else title
                
                if cnt >= 2:
                    is_danger_exam = True
                    # ✅ '과락'을 '재시험'으로 변경!
                    fail_reasons.append(f"{short_title} {cnt}회 재시험")
                elif cnt == 1:
                    is_caution_exam = True
                    # ✅ '과락'을 '재시험'으로 변경!
                    fail_reasons.append(f"{short_title} 1차 재시험")
                    
            # ★ 2. 미해결된 잠금 확인
            has_locked_fail = StudentLog.objects.filter(profile=student, log_type='exam_fail', is_resolved=False).exists()

            # -----------------------------------------------
            # 🔴 위험(Danger) 분류
            # -----------------------------------------------
            if student.warning_count >= 3 or is_danger_exam or student.status == 'counseling':
                reasons = []
                if student.warning_count > 0: reasons.append(f"경고 {student.warning_count}회")
                if is_danger_exam: reasons.extend(fail_reasons)
                if student.status == 'counseling' and not reasons: reasons.append("면담필요(잠금)")
                
                student.risk_reason = " / ".join(reasons) 
                danger_list.append(student)
            
            # -----------------------------------------------
            # 🟡 주의(Caution) 분류
            # -----------------------------------------------
            elif student.warning_count in [1, 2] or is_caution_exam or has_locked_fail or student.status == 'caution':
                reasons = []
                if student.warning_count > 0: reasons.append(f"경고 {student.warning_count}회")
                if is_caution_exam: reasons.extend(fail_reasons)
                if has_locked_fail and not is_caution_exam: reasons.append("시험 잠금상태")
                if student.status == 'caution' and not reasons: reasons.append("주의요망")
                
                student.risk_reason = " / ".join(reasons) 
                caution_list.append(student)
            
            # -----------------------------------------------
            # 🟢 정상(Normal) 분류
            # -----------------------------------------------
            else:
                normal_list.append(student)
                
        # 공정에 학생이 있거나, 매니저 본인 공정이면 리스트에 추가
        if p_students.exists() or not user.is_superuser: 
            process_summary.append({
                'process_name': p.name if p else "공정 미지정",
                'danger': danger_list,
                'caution': caution_list,
                'normal': normal_list,
                'danger_count': len(danger_list),
                'caution_count': len(caution_list),
                'normal_count': len(normal_list),
                'total_count': len(p_students)
            })
    # ========================================================
    
    reference_links = ReferenceLink.objects.all()

    common_quizzes = Quiz.objects.filter(category='common', is_published=True).order_by('-created_at')

    context = {
        'signup_pending_count': signup_pending_count,
        'exam_pending_count': exam_pending_count,
        'risk_count': risk_count,
        'access_req_count': access_req_count,
        'schedule_pending_count': schedule_pending_count,
        'counseling_count': counseling_count,
        'active_trainees': active_trainees,
        'filter_cohorts': filter_cohorts,
        'filter_processes': filter_processes,
        'filter_companies': filter_companies,
        'process_summary': process_summary, # ★ 템플릿으로 전달!
        'process_summary': process_summary,
        'reference_links': reference_links,
        'common_quizzes': common_quizzes,
    }
    
    return render(request, 'quiz/manager/dashboard_main.html', context)





@login_required
def manager_trainee_list(request):
    if not request.user.is_staff: return redirect('quiz:index')

    today = timezone.now().date()
    
    # =========================================================
    # 1. 기간 만료자 자동 비활성화 (상태는 건드리지 않음)
    # =========================================================
    expired_profiles = Profile.objects.filter(
        cohort__end_date__lt=today, 
        user__is_active=True
    ).exclude(
        Q(user__is_superuser=True) | Q(user__is_staff=True) | 
        Q(is_manager=True) | Q(is_pl=True)
    ).select_related('user')

    if expired_profiles.exists():
        expired_user_ids = expired_profiles.values_list('user_id', flat=True)
        # 계정 로그인만 차단시킴 (상태는 아직 '재직중'으로 냅둠)
        User.objects.filter(id__in=expired_user_ids).update(is_active=False)
        print(f"⚠️ [System] 기간 만료된 교육생 {len(expired_user_ids)}명의 로그인을 차단했습니다.")

    # =========================================================
    # [신규 추가] 1-5. 평가 대기(미수료) 인원 계산
    # (기수는 끝났는데 상태가 아직 수료/퇴소가 아닌 사람들)
    # =========================================================
    pending_eval_count = Profile.objects.filter(
        cohort__end_date__lt=today,
        status__in=['attending', 'caution', 'counseling']
    ).exclude(user__is_superuser=True).exclude(is_manager=True).count()

    # =========================================================
    # 2. 리스트 조회 및 필터링
    # =========================================================
    active_cohort = Cohort.objects.filter(start_date__lte=today, end_date__gte=today).first()
    default_cohort_id = active_cohort.id if active_cohort else ''

    data = request.GET.copy()
    if 'cohort' not in data and default_cohort_id:
        data['cohort'] = default_cohort_id

    form = TraineeFilterForm(data)
    
    profiles = Profile.objects.select_related('user', 'cohort', 'process').exclude(
        user__is_superuser=True, is_manager=True
    ).order_by('cohort__start_date', 'name')

    if form.is_valid():
        if form.cleaned_data['cohort']: profiles = profiles.filter(cohort=form.cleaned_data['cohort'])
        if form.cleaned_data['process']: profiles = profiles.filter(process=form.cleaned_data['process'])
        
        status_val = form.cleaned_data['status']
        if status_val:
            if status_val == 'inactive': profiles = profiles.filter(is_approved=False)
            else: profiles = profiles.filter(status=status_val, is_approved=True)

        if form.cleaned_data['search']:
            q = form.cleaned_data['search']
            profiles = profiles.filter(
                Q(name__icontains=q)|Q(employee_id__icontains=q)|Q(user__username__icontains=q)
            )


    # URL에 ?filter=pending_eval 이 있으면 수료 대기자만 필터링!
    if request.GET.get('filter') == 'pending_eval':
        profiles = profiles.filter(
            cohort__end_date__lt=today,
            status__in=['attending', 'caution', 'counseling']
        )

    paginator = Paginator(profiles, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    pending_profiles = Profile.objects.filter(is_approved=False).select_related('user').order_by('-user__date_joined')

    return render(request, 'quiz/manager/trainee_list.html', {
        'form': form, 
        'profiles': page_obj, 
        'pending_users': pending_profiles,
        'total_count': profiles.count(),
        'pending_eval_count': pending_eval_count # ★ 알림용 변수 전달
    })

    # 4. 페이지네이션
    paginator = Paginator(profiles, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    # 5. [핵심 수정] 가입 대기자 목록 (Pending Users)
    # 기존: User.objects.filter(is_active=False) -> 이제는 모두 active=True이므로 0명 나옴
    # 변경: Profile.objects.filter(is_approved=False) -> 승인 안 된 사람 조회
    pending_profiles = Profile.objects.filter(is_approved=False).select_related('user').order_by('-user__date_joined')

    return render(request, 'quiz/manager/trainee_list.html', {
        'form': form, 
        'profiles': page_obj, 
        'pending_users': pending_profiles, # 템플릿 변수명 호환을 위해 pending_users 유지 (내용은 프로필 리스트)
        'total_count': profiles.count()
    })

# =========================================================
# 1. 교육생 상세 정보 (신호등/Stepper 포함)
# =========================================================
@login_required
def manager_trainee_detail(request, profile_id):
    """
    [교육생 상세] 시험 진행 프로세스 뷰 (최종 수정본)
    - 메인(공통/전공)과 기타(안전/기타)를 동일한 상세 포맷(1차/2차/3차)으로 가공
    """
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    profile = get_object_or_404(Profile, pk=profile_id)
    student = profile.user

    is_my_student = is_process_manager(request.user, profile)

    # ★ [보안 2] 타 공정 열람 티켓(권한)이 있는지 확인
    from accounts.models import ProcessAccessRequest
    has_ticket = False
    if not is_my_student and not request.user.is_superuser:
        has_ticket = ProcessAccessRequest.objects.filter(
            requester=request.user,
            status='approved',
            expires_at__gte=timezone.now()
        ).filter(
            Q(target_process=profile.process) | Q(target_process__isnull=True)
        ).exists()

    # 상세 조회 가능 여부 = 내 학생이거나, 타공정 열람 티켓이 있거나!
    can_view_detail = is_my_student or has_ticket

    # ---------------------------------------------------------
    # [1] 퀴즈 목록 분리 조회
    # ---------------------------------------------------------
    # 1. 메인 (공통 + 내 전공)
    main_quizzes = Quiz.objects.filter(
        Q(category='common') | Q(related_process=profile.process)
    ).distinct().order_by('category', 'title')

    # 2. 기타 (안전 + 기타)
    etc_quizzes = Quiz.objects.filter(
        category__in=['safety', 'etc']
    ).distinct().order_by('category', 'title')

    # ---------------------------------------------------------
    # [2] 데이터 가공 헬퍼 함수 (핵심: 이걸로 양쪽 다 똑같이 만듦)
    # ---------------------------------------------------------
    def make_process_list(quiz_qs):
        result_list = []
        for quiz in quiz_qs:
            history = TestResult.objects.filter(user=student, quiz=quiz).order_by('completed_at')
            attempts = list(history)
            count = len(attempts)
            last_result = attempts[-1] if count > 0 else None
            
            status = 'not_taken'
            score = '-'
            date = None
            if last_result:
                score = f"{last_result.score}점"
                date = last_result.completed_at
                status = 'pass' if last_result.is_pass else 'fail'

            # 잠금 여부
            is_locked = False
            if status == 'fail':
                is_locked = StudentLog.objects.filter(
                    profile=profile, related_quiz=quiz, log_type='exam_fail', is_resolved=False
                ).exists()

            # 쪽지 로그
            quiz_logs = StudentLog.objects.filter(
                profile=profile, related_quiz=quiz
            ).select_related('recorder').order_by('-created_at')

            result_list.append({
                'quiz': quiz,
                'status': status,      
                'score': score,
                'date': date,
                'try_1': attempts[0] if count >= 1 else None, 
                'try_2': attempts[1] if count >= 2 else None, 
                'try_3': attempts[2] if count >= 3 else None, 
                'is_locked': is_locked,
                'logs': quiz_logs,
            })
        return result_list

    # ---------------------------------------------------------
    # [3] 데이터 생성 및 전달
    # ---------------------------------------------------------
    exam_process_list = make_process_list(main_quizzes) # 상단 메인
    etc_process_list = make_process_list(etc_quizzes)   # 하단 기타 (동일 포맷)

    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')
    results = TestResult.objects.filter(user=student).order_by('-completed_at')
    badges = getattr(profile, 'badges', None)
    if badges: badges = badges.all()

    return render(request, 'quiz/manager/trainee_detail.html', {
        'profile': profile,
        'exam_process_list': exam_process_list,
        'etc_process_list': etc_process_list, # 이름 변경됨 (etc_list -> etc_process_list)
        'logs': logs,
        'results': results,
        'badges': badges,
        'is_my_student': is_my_student,
        'can_view_detail': can_view_detail,
    })

# =========================================================
# 2. AJAX 로그 저장 (모달 창에서 '저장' 클릭 시 호출)
# =========================================================
@login_required
@require_POST
def final_log_saver(request, profile_id):
    try:
        # [디버깅] 요청 확인
        print(f"\n🔥🔥🔥 [최종 함수 실행] ID: {profile_id} 요청 도착! 🔥🔥🔥")
        
        profile = get_object_or_404(Profile, pk=profile_id)
        
        # 데이터 수신
        content = request.POST.get('content')
        opinion = request.POST.get('opinion')
        is_passed = request.POST.get('is_passed') == 'on'
        quiz_id = request.POST.get('quiz_id')
        raw_stage = request.POST.get('stage')

        print(f"👉 받은 데이터 - QuizID: {quiz_id}, Stage: {raw_stage}")

        if not quiz_id:
            print("❌ 실패: Quiz ID가 없습니다.")
            return JsonResponse({'status': 'error', 'message': 'Quiz ID 누락'})

        related_quiz = get_object_or_404(Quiz, pk=quiz_id)
        
        try:
            stage = int(raw_stage)
        except:
            stage = 1

        # [DB 저장]
        # 로그 타입은 빨간 테두리 표시를 위해 'warning'을 유지하되, 
        # 실제 경고 카운트는 아래에서 올리지 않도록 제어합니다.
        StudentLog.objects.create(
            profile=profile, 
            recorder=request.user,
            log_type='counseling' if is_passed else 'warning',
            reason=content, 
            action_taken=opinion, 
            is_resolved=is_passed,
            related_quiz=related_quiz,  # ★ 시험 정보 연결 필수
            stage=stage
        )

        # [상태 업데이트]
        if is_passed:
            # 합격/조치완료 시 -> '상담필요' 상태였다면 '재직중'으로 복구
            if profile.status == 'counseling': 
                profile.status = 'attending'
        
        # ⛔ [삭제됨] 아래 코드를 지워서 시험 불합격 시 경고 카운트가 올라가지 않게 함
        # else:
        #     profile.warning_count += 1
        
        profile.save()
        
        return JsonResponse({'status': 'success', 'message': '✅ 상담 내용이 저장되었습니다.'})

    except Exception as e:
        print(f"❌ 에러: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)})


# =========================================================
# 3. 특이사항/경고/징계 관리 페이지 (히스토리 탭의 폼 처리)
# =========================================================
@login_required
def manage_student_logs(request, profile_id):
    """
    [관리자용] 교육생 특이사항/로그 관리 페이지
    - 기능: 시험 프로세스 조회, 경고/면담 기록 저장, 잠금 해제
    """
    # 1. 권한 체크
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    profile = get_object_or_404(Profile, pk=profile_id)
    student = profile.user

    # ========================================================
    # [1] View Data: 시험 진행 프로세스 표 데이터 생성
    # ========================================================
    target_quizzes = Quiz.objects.filter(
        Q(category='common') | Q(related_process=profile.process)
    ).distinct().order_by('category', 'title')

    exam_process_list = []
    for quiz in target_quizzes:
        # created_at -> completed_at (수정완료)
        history = TestResult.objects.filter(user=student, quiz=quiz).order_by('completed_at')
        attempts = list(history)
        count = len(attempts)
        last_result = attempts[-1] if count > 0 else None
        
        status = 'not_taken'
        score = '-'
        date = None # 변수명 통일
        
        if last_result:
            score = f"{last_result.score}점"
            date = last_result.completed_at
            status = 'pass' if last_result.is_pass else 'fail'

        # 잠금 여부 확인
        is_locked = False
        if status == 'fail':
            is_locked = StudentLog.objects.filter(
                profile=profile, related_quiz=quiz, log_type='exam_fail', is_resolved=False
            ).exists()
        
        # 쪽지 기능용 로그
        quiz_logs = StudentLog.objects.filter(profile=profile, related_quiz=quiz).order_by('-created_at')

        exam_process_list.append({
            'quiz': quiz,
            'status': status,
            'score': score,
            'date': date,
            'try_1': attempts[0] if count >= 1 else None,
            'try_2': attempts[1] if count >= 2 else None,
            'try_3': attempts[2] if count >= 3 else None,
            'is_locked': is_locked,
            'logs': quiz_logs
        })

    # ========================================================
    # [2] View Data: 잠긴 시험 로그 (드롭다운 선택용)
    # ========================================================
    locked_logs = StudentLog.objects.filter(
        profile=profile,
        log_type='exam_fail',
        is_resolved=False
    ).select_related('related_quiz').order_by('-created_at')

    # ========================================================
    # [3] POST 요청 처리: 로그 저장 및 로직 실행
    # ========================================================
    if request.method == 'POST':
        log_type = request.POST.get('log_type')
        reason = request.POST.get('reason')
        action_taken = request.POST.get('action_taken')
        
        is_unlocked = request.POST.get('resolve_lock') == 'on'
        pl_check = request.POST.get('pl_check') == 'on'
        related_quiz_id = request.POST.get('related_quiz_id')
        related_quiz = get_object_or_404(Quiz, pk=related_quiz_id) if related_quiz_id else None

        try:
            with transaction.atomic():
                # (1) 로그 생성
                new_log = StudentLog.objects.create(
                    profile=profile,
                    recorder=request.user,
                    log_type=log_type,
                    reason=reason,
                    action_taken=action_taken,
                    related_quiz=related_quiz,
                    is_resolved=is_unlocked,
                    created_at=timezone.now()
                )

                # (2) 경고 누적 로직
                if log_type == 'warning':
                    profile.warning_count += 1
                    if profile.warning_count == 2:
                        StudentLog.objects.create(
                            profile=profile, recorder=request.user, log_type='warning_letter',
                            reason="[시스템 자동] 경고 2회 누적 -> 1차 경고장", action_taken="매니저 면담 필요", is_resolved=False
                        )
                        profile.status = 'counseling'
                        messages.warning(request, "⚠️ 경고 2회 누적! 1차 경고장이 자동 발부되었습니다.")
                    elif profile.warning_count == 3:
                        StudentLog.objects.create(
                            profile=profile, recorder=request.user, log_type='warning_letter',
                            reason="[시스템 자동] 경고 3회 누적 -> 2차 경고장", action_taken="PL 면담 필수", is_resolved=False
                        )
                        profile.status = 'counseling'
                        messages.error(request, "🚫 경고 3회 누적! 2차 경고장이 발부되었습니다.")
                    elif profile.warning_count >= 4:
                        profile.status = 'dropout'
                        profile.user.is_active = False
                        profile.user.save()
                        messages.error(request, "⛔ 경고 4회 누적! 퇴소 처리되었습니다.")
                    else:
                        profile.status = 'caution'
                        messages.info(request, "경고가 1회 적립되었습니다.")

                # (3) 경고장 수동 발부
                elif log_type == 'warning_letter':
                    if profile.warning_count < 2: profile.warning_count = 2
                    else: profile.warning_count += 1
                    
                    if profile.warning_count >= 4:
                        profile.status = 'dropout'
                        profile.user.is_active = False
                        profile.user.save()
                    else:
                        profile.status = 'counseling'
                    messages.warning(request, f"⛔ 경고장이 발부되었습니다.")

                # (4) 면담 및 잠금 해제
                elif log_type == 'counseling' or log_type == 'exam_fail':
                    if is_unlocked:
                        if related_quiz:
                            StudentLog.objects.filter(
                                profile=profile, related_quiz=related_quiz, log_type='exam_fail', is_resolved=False
                            ).update(is_resolved=True)
                            messages.success(request, f"시험 '{related_quiz.title}' 잠금 해제됨.")

                        if profile.warning_count == 3 and not pl_check:
                            messages.error(request, "🚫 3회 누적자는 'PL 면담 확인' 필수입니다.")
                            new_log.is_resolved = False
                            new_log.save()
                        elif profile.warning_count >= 4:
                            profile.status = 'dropout'
                            messages.error(request, "퇴소자는 잠금을 해제할 수 없습니다.")
                            new_log.is_resolved = False
                            new_log.save()
                        else:
                            if profile.status == 'counseling': profile.status = 'attending'
                            if not profile.user.is_active:
                                profile.user.is_active = True
                                profile.user.save()
                            messages.success(request, "계정이 정상화되었습니다.")
                    else:
                        messages.info(request, "면담 기록이 저장되었습니다.")

                profile.save()
        except Exception as e:
            messages.error(request, f"저장 중 오류 발생: {e}")

        return redirect('quiz:manage_student_logs', profile_id=profile.id)

    # ========================================================
    # [4] GET: 최종 데이터 렌더링
    # ========================================================
    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')

    # ★ 에러가 나던 badges 코드는 삭제했습니다.
    # ★ 템플릿도 manage_student_logs.html 로 정확히 지정했습니다.
    return render(request, 'quiz/manager/manage_student_logs.html', {
        'profile': profile,
        'exam_process_list': exam_process_list,
        'logs': logs,
        'locked_logs': locked_logs,
    })



# =========================================================
# 4. 특이사항/경고 관리 페이지 (블라인드 + 부분 권한 허용 + 파일 첨부)
# =========================================================
@login_required
def manage_student_logs(request, profile_id):
    # 1. 스태프 권한 체크
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    profile = get_object_or_404(Profile, pk=profile_id)
    student = profile.user

    # ★ [핵심 추가] 내 담당 공정 학생인지(또는 내가 슈퍼유저인지) 판별
    is_my_student = is_process_manager(request.user, profile)

    exam_process_list = []
    locked_logs = []

    # ========================================================
    # [1] View Data: 시험 진행 프로세스 및 잠긴 로그
    # ★ 내 학생일 때만 성적을 불러오고, 남의 학생이면 빈 리스트(블라인드) 처리!
    # ========================================================
    if is_my_student:
        target_quizzes = Quiz.objects.filter(
            Q(category='common') | Q(related_process=profile.process)
        ).distinct().order_by('category', 'title')

        for quiz in target_quizzes:
            history = TestResult.objects.filter(user=student, quiz=quiz).order_by('completed_at')
            attempts = list(history)
            count = len(attempts)
            last_result = attempts[-1] if count > 0 else None
            
            status = 'not_taken'
            score = '-'
            date = None
            
            if last_result:
                score = f"{last_result.score}점"
                date = last_result.completed_at
                status = 'pass' if last_result.is_pass else 'fail'

            is_locked = False
            if status == 'fail':
                is_locked = StudentLog.objects.filter(
                    profile=profile, related_quiz=quiz, log_type='exam_fail', is_resolved=False
                ).exists()
            
            quiz_logs = StudentLog.objects.filter(profile=profile, related_quiz=quiz).order_by('-created_at')

            exam_process_list.append({
                'quiz': quiz,
                'status': status,
                'score': score,
                'date': date,
                'try_1': attempts[0] if count >= 1 else None,
                'try_2': attempts[1] if count >= 2 else None,
                'try_3': attempts[2] if count >= 3 else None,
                'is_locked': is_locked,
                'logs': quiz_logs
            })

        locked_logs = StudentLog.objects.filter(
            profile=profile, log_type='exam_fail', is_resolved=False
        ).select_related('related_quiz').order_by('-created_at')


    # ========================================================
    # [2] POST 요청 처리: 로그 저장 및 로직 실행
    # ========================================================
    if request.method == 'POST':
        log_type = request.POST.get('log_type')
        reason = request.POST.get('reason')
        action_taken = request.POST.get('action_taken')
        
        is_unlocked = request.POST.get('resolve_lock') == 'on'
        pl_check = request.POST.get('pl_check') == 'on'
        related_quiz_id = request.POST.get('related_quiz_id')
        related_quiz = get_object_or_404(Quiz, pk=related_quiz_id) if related_quiz_id else None

        # ★ [수동 경고장 중복 방어막]
        # 매니저가 수동으로 '경고장'을 주면서 '지각'이라고 적었을 때
        if log_type == 'warning_letter' and '지각' in reason:
            today = timezone.now().date()
            already_auto_warned = StudentLog.objects.filter(
                profile=profile,
                log_type='warning_letter',
                created_at__date=today,
                reason__contains='[시스템 자동]'
            ).filter(reason__contains='지각').exists()

            # 오늘 이미 달력 출석부에서 자동으로 때린 지각 경고장이 있다면 튕겨냄!
            if already_auto_warned:
                messages.error(request, "⛔ 오늘 이미 달력 시스템에서 자동으로 발부된 지각 경고장이 있습니다. 중복으로 발부할 수 없습니다.")
                return redirect('quiz:manage_student_logs', profile_id=profile.id)

        
        # ★ [핵심 추가] 업로드된 첨부파일 가져오기
        attached_file = request.FILES.get('attached_file')
        related_quiz_id = request.POST.get('related_quiz_id') # 시험 선택 여부 확인
        
        # 파일이 꼭 필요한 경우: '잠금 해제'를 체크했는데 '시험 해제(related_quiz_id)'가 아닐 때!
        # 즉, 순수하게 '태도/인성 경고 누적'으로 잠긴 걸 풀어줄 때만 경고장 사본을 요구합니다.
        needs_file = (log_type == 'counseling' and is_unlocked and not related_quiz_id)
        
        if needs_file and not attached_file:
            messages.error(request, "⛔ 경고 누적으로 인한 잠금을 해제하려면 반드시 증빙 서류(경고장 사본 등)를 첨부해야 합니다. (시험 재응시 해제는 서류 불필요)")
            return redirect('quiz:manage_student_logs', profile_id=profile.id)

        # ★ [백엔드 철통 보안] 타 매니저가 해킹(소스 조작)으로 시험 잠금을 풀려 하면 튕겨냄!
        if not is_my_student:
            if is_unlocked or related_quiz_id or log_type in ['exam_fail', 'warning_letter']:
                messages.error(request, "⛔ 타 공정 학생의 성적 관련 조치나 경고장 발부는 불가능합니다. (태도 경고만 가능)")
                return redirect('quiz:manage_student_logs', profile_id=profile.id)

        try:
            with transaction.atomic():
                # (1) 로그 생성
                new_log = StudentLog.objects.create(
                    profile=profile,
                    recorder=request.user,
                    log_type=log_type,
                    reason=reason,
                    action_taken=action_taken,
                    related_quiz=related_quiz,
                    is_resolved=is_unlocked,
                    attached_file=attached_file, # ★ [핵심 적용] 파일 데이터를 DB에 쏙!
                    created_at=timezone.now()
                )

                # (2) 경고 누적 로직 (내 학생이든 남의 학생이든 태도 불량 경고 횟수는 올라감)
                if log_type == 'warning':
                    profile.warning_count += 1
                    if profile.warning_count == 2:
                        StudentLog.objects.create(
                            profile=profile, recorder=request.user, log_type='warning_letter',
                            reason="[시스템 자동] 경고 2회 누적 -> 1차 경고장", action_taken="매니저 면담 필요", is_resolved=False
                        )
                        profile.status = 'counseling'
                        messages.warning(request, "⚠️ 경고 2회 누적! 1차 경고장이 자동 발부되었습니다.")
                    elif profile.warning_count == 3:
                        StudentLog.objects.create(
                            profile=profile, recorder=request.user, log_type='warning_letter',
                            reason="[시스템 자동] 경고 3회 누적 -> 2차 경고장", action_taken="PL 면담 필수", is_resolved=False
                        )
                        profile.status = 'counseling'
                        messages.error(request, "🚫 경고 3회 누적! 2차 경고장이 발부되었습니다.")
                    elif profile.warning_count >= 4:
                        profile.status = 'dropout'
                        profile.user.is_active = False
                        profile.user.save()
                        messages.error(request, "⛔ 경고 4회 누적! 퇴소 처리되었습니다.")
                    else:
                        profile.status = 'caution'
                        messages.info(request, "경고가 1회 적립되었습니다.")

                # (3) 경고장 수동 발부 (내 담당 학생일 때만 가능)
                elif log_type == 'warning_letter':
                    if profile.warning_count < 2: profile.warning_count = 2
                    else: profile.warning_count += 1
                    
                    if profile.warning_count >= 4:
                        profile.status = 'dropout'
                        profile.user.is_active = False
                        profile.user.save()
                    else:
                        profile.status = 'counseling'
                    messages.warning(request, f"⛔ 경고장이 발부되었습니다.")

                    # ★ [핵심 추가] 교육생에게 '경고장 양식 다운로드' 알림 발송
                    from django.urls import reverse
                    download_url = reverse('quiz:print_warning_letter', args=[new_log.id])
                    Notification.objects.create(
                        recipient=profile.user,
                        sender=request.user,
                        notification_type='general', 
                        message=f"🚨 [경고장 발부] 규정 위반으로 경고장이 발부되었습니다. 클릭하여 양식을 인쇄 후 서명하여 제출 바랍니다.",
                        related_url=download_url
                    )
                    # 실시간 종소리 빵!
                    try:
                        broadcast_realtime_notification(profile.user.id)
                    except:
                        pass

                # (4) 면담 및 잠금 해제 (내 담당 학생일 때만 가능)
                elif log_type == 'counseling' or log_type == 'exam_fail':
                    if is_unlocked:
                        if related_quiz:
                            StudentLog.objects.filter(
                                profile=profile, related_quiz=related_quiz, log_type='exam_fail', is_resolved=False
                            ).update(is_resolved=True)
                            messages.success(request, f"시험 '{related_quiz.title}' 잠금 해제 및 파일 업로드 됨.")

                        if profile.warning_count == 3 and not pl_check:
                            messages.error(request, "🚫 3회 누적자는 'PL 면담 확인' 필수입니다.")
                            new_log.is_resolved = False
                            new_log.save()
                        elif profile.warning_count >= 4:
                            profile.status = 'dropout'
                            messages.error(request, "퇴소자는 잠금을 해제할 수 없습니다.")
                            new_log.is_resolved = False
                            new_log.save()
                        else:
                            if profile.status == 'counseling': profile.status = 'attending'
                            if not profile.user.is_active:
                                profile.user.is_active = True
                                profile.user.save()
                            messages.success(request, "계정이 정상화되었습니다.")
                    else:
                        messages.info(request, "일반 면담/경고 기록이 저장되었습니다.")

                profile.save()
        except Exception as e:
            messages.error(request, f"저장 중 오류 발생: {e}")

        return redirect('quiz:manage_student_logs', profile_id=profile.id)

    # ========================================================
    # [3] GET: 최종 데이터 렌더링
    # ========================================================
    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')

    return render(request, 'quiz/manager/manage_student_logs.html', {
        'profile': profile,
        'exam_process_list': exam_process_list,
        'logs': logs,
        'locked_logs': locked_logs,
        'is_my_student': is_my_student, # ★ [핵심] HTML에 권한 정보 전달!
    })

# =========================================================
# 3. 최종 평가서 작성 (데이터 채워넣기)
# =========================================================
@login_required
def manager_trainee_report(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    
    # 1. 통계 데이터 계산 (왼쪽 사이드바용)
    results = TestResult.objects.filter(user=profile.user)
    
    # 평균 점수
    avg_score = results.aggregate(Avg('score'))['score__avg']
    avg_score = round(avg_score, 1) if avg_score else 0
    
    # 재시험(불합격) 횟수
    fail_count = results.filter(is_pass=False).count()
    
    # 최근 특이사항 로그 (최신 5개)
    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')[:5]

    # 2. 저장(POST) 처리
    if request.method == 'POST':
        # (여기에 평가 저장 로직 구현 가능)
        messages.success(request, f"{profile.name}님의 최종 평가가 저장되었습니다.")
        return redirect('quiz:manager_trainee_detail', profile_id=profile.id)

    # 3. 화면 렌더링
    return render(request, 'quiz/manager/final_report.html', {
        'profile': profile,
        'avg_score': avg_score,
        'fail_count': fail_count,
        'logs': logs,
    })


# =========================================================
# 4. [AJAX] 상세페이지 모달용 로그 저장
# =========================================================
@login_required
@require_POST
def manager_create_log_ajax(request, profile_id):
    if not request.user.is_staff: 
        return JsonResponse({'status': 'error', 'message': '권한이 없습니다.'}, status=403)
    
    try:
        profile = get_object_or_404(Profile, pk=profile_id)
        
        # 데이터 수신
        content = request.POST.get('content')
        opinion = request.POST.get('opinion')
        is_passed = request.POST.get('is_passed') == 'on'
        quiz_id = request.POST.get('quiz_id')
        stage = request.POST.get('stage')

        # 시험 정보 가져오기
        related_quiz = None
        if quiz_id:
            related_quiz = get_object_or_404(Quiz, pk=quiz_id)

        # 로그 저장
        StudentLog.objects.create(
            profile=profile,
            recorder=request.user,
            log_type='counseling' if is_passed else 'warning',
            reason=content,
            action_taken=opinion,
            is_resolved=is_passed,
            related_quiz=related_quiz,
            stage=stage if stage else 1
        )

        # 상태 업데이트 (시험 불합격은 경고 카운트 증가 X)
        if is_passed:
            if profile.status == 'counseling': profile.status = 'attending'
        
        # [삭제됨] else: profile.warning_count += 1 
        # (사용자님 요청대로 시험 관련은 카운트 안 올림)
        
        profile.save()
        return JsonResponse({'status': 'success', 'message': '저장되었습니다.'})

    except Exception as e:
        print(f"❌ 로그 저장 에러: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)})


# --- (기타 액션 함수들: 가입승인, 비번초기화 등 기존 유지) ---
@login_required
@require_POST
def approve_signup_bulk(request):
    if not request.user.is_staff: 
        return JsonResponse({'status':'error', 'message': '권한이 없습니다.'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return JsonResponse({'status': 'error', 'message': '선택된 사용자가 없습니다.'})

        users = User.objects.filter(id__in=user_ids)
        
        if data.get('action') == 'approve':
            users.update(is_active=True)
            Profile.objects.filter(user__in=users).update(is_approved=True)
            return JsonResponse({'status': 'success', 'message': f'{users.count()}명의 가입이 승인되었습니다.'})
        else:
            users.delete()
            return JsonResponse({'status': 'success', 'message': '가입이 거절 및 삭제되었습니다.'})
            
    except Exception as e:
        # ★ 파이썬 에러가 나도 무조건 JSON으로 에러 이유를 알려줍니다!
        return JsonResponse({'status': 'error', 'message': f'서버 오류: {str(e)}'})
    users = User.objects.filter(id__in=data.get('user_ids', []))
    if data.get('action') == 'approve':
        users.update(is_active=True)
        Profile.objects.filter(user__in=users).update(is_approved=True)
        return JsonResponse({'status':'success', 'message': f'{users.count()}명 승인 완료'})
    else:
        users.delete()
        return JsonResponse({'status':'success', 'message': '거절 완료'})

@login_required
@require_POST
def reset_password_bulk(request):
    if not request.user.is_staff: return JsonResponse({'status':'error'}, status=403)
    data = json.loads(request.body)
    users = User.objects.filter(id__in=data.get('user_ids', []))
    for u in users:
        u.set_password('1234')
        if hasattr(u, 'profile'): u.profile.must_change_password = True; u.profile.save()
        u.save()
    return JsonResponse({'status':'success', 'message': '초기화 완료'})

@login_required
@require_POST
def unlock_account(request, profile_id):
    if not request.user.is_staff: return JsonResponse({'status':'error'}, status=403)
    p = get_object_or_404(Profile, pk=profile_id)
    if p.status in ['counseling', 'dropout']:
        p.status = 'attending'; p.save()
        return JsonResponse({'status':'success', 'message': '해제 완료'})
    return JsonResponse({'status':'info', 'message': '이미 정상입니다.'})




# 7. 응시 요청 관리 페이지
@login_required
def manager_exam_requests(request):
    """
    시험 응시 요청 및 공정 조회 권한 요청을 한 곳에서 관리하는 뷰
    """
    if not request.user.is_staff: return redirect('quiz:index')

    # 1. 시험 응시 요청 (QuizAttempt)
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.process:
        exam_reqs = QuizAttempt.objects.filter(
            status='대기중', 
            user__profile__process=request.user.profile.process
        ).order_by('requested_at')
    else:
        exam_reqs = QuizAttempt.objects.filter(status='대기중').order_by('requested_at')

    # 2. [신규 추가] 권한 조회 요청 (ProcessAccessRequest)
    access_reqs = []
    try:
        # 관리자: 모든 요청 확인
        if request.user.is_superuser:
            access_reqs = ProcessAccessRequest.objects.filter(status='pending').order_by('created_at')
        # 매니저: 내 공정에 대한 요청만 확인
        elif hasattr(request.user, 'profile') and request.user.profile.process:
            access_reqs = ProcessAccessRequest.objects.filter(
                target_process=request.user.profile.process,
                status='pending'
            ).order_by('created_at')
    except NameError:
        pass

    return render(request, 'quiz/manager/exam_requests.html', {
        'requests': exam_reqs,       # 시험 요청
        'access_requests': access_reqs # 권한 요청 (추가됨)
    })

# 1. PL 대시보드
@login_required
def pl_dashboard(request):
    # ==========================================
    # (1) 권한 체크
    # ==========================================
    # 스태프 권한이 있으면서, (PL 프로필이 있거나 슈퍼유저인 경우)만 통과
    if not (request.user.is_staff and (getattr(request.user, 'profile', None) and (request.user.profile.is_pl or request.user.is_superuser))):
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')
    
    # ==========================================
    # (2) 기본 대상 설정 (슈퍼유저 vs PL)
    # ==========================================
    if request.user.is_superuser:
        # 관리자는 전체 보기
        trainees = Profile.objects.select_related('user', 'cohort', 'process').all()
    else:
        try:
            # PL은 본인 파트의 교육생만 조회
            pl_obj = PartLeader.objects.get(email=request.user.email)
            trainees = Profile.objects.filter(pl=pl_obj).select_related('user', 'cohort', 'process')
        except PartLeader.DoesNotExist:
            trainees = Profile.objects.none()

    # ==========================================
    # (3) 검색 및 필터링 적용
    # ==========================================
    search_query = request.GET.get('q', '')
    filter_cohort = request.GET.get('cohort', '')
    filter_process = request.GET.get('process', '')

    if search_query:
        trainees = trainees.filter(name__icontains=search_query)
    if filter_cohort:
        trainees = trainees.filter(cohort_id=filter_cohort)
    if filter_process:
        trainees = trainees.filter(process_id=filter_process)

    # ==========================================
    # (4) 통계 데이터 계산
    # ==========================================
    total_count = trainees.count()
    no_data = total_count == 0

    # 상태별 카운트
    status_counts = {
        'attending': trainees.filter(status='attending').count(),
        'counseling': trainees.filter(status='counseling').count(),
        'dropout': trainees.filter(status='dropout').count(),
        'completed': trainees.filter(status='completed').count(),
    }

    # 평가 데이터 집계 (평가 데이터가 있는 인원만 대상)
    assessed = trainees.filter(final_assessment__isnull=False)
    
    if assessed.exists():
        # 전체 평균 점수
        avg_final = assessed.aggregate(Avg('final_assessment__final_score'))['final_assessment__final_score__avg']
        
        # 레이더 차트용 영역별 평균
        radar_data = assessed.aggregate(
            avg_exam=Avg('final_assessment__exam_avg_score'),
            avg_prac=Avg('final_assessment__practice_score'),
            avg_note=Avg('final_assessment__note_score'),
            avg_atti=Avg('final_assessment__attitude_score')
        )
        # 성적 우수자 Top 3
        top_trainees = assessed.order_by('-final_assessment__final_score')[:3]
    else:
        avg_final = 0
        radar_data = {'avg_exam': 0, 'avg_prac': 0, 'avg_note': 0, 'avg_atti': 0}
        top_trainees = []

    # 위험군 식별 (상담요망 상태이거나, 점수가 60점 미만인 경우)
    risk_trainees = trainees.filter(
        Q(status='counseling') | 
        Q(warning_count__gte=2) |
        (Q(final_assessment__final_score__lt=60) & Q(final_assessment__isnull=False))
    )

    # ==========================================
    # (5) 리스트 데이터 가공 (템플릿 출력용)
    # ==========================================
    trainee_list = []
    for t in trainees:
        # 1. 최종 성적 (FinalAssessment)
        fa = getattr(t, 'final_assessment', None)
        
        # 2. 매니저 평가 (ManagerEvaluation) 가져오기
        # [수정됨] profile -> t 로 변경 (여기가 에러 원인이었습니다)
        last_eval = t.managerevaluation_set.last() 
        
        # 3. 데이터 추출 (평가서가 없을 경우 대비)
        checklist = last_eval.selected_items.all() if last_eval else []
        manager_comment = last_eval.overall_comment if last_eval else ""

        # 4. 리스트에 담기
        trainee_list.append({
            'profile': t,
            'final_score': fa.final_score if fa else 0,
            'exam_avg': fa.exam_avg_score if fa else 0,
            'rank': getattr(fa, 'rank', '-'), 
            
            # 템플릿으로 보낼 데이터
            'manager_comment': manager_comment, 
            'checklist': checklist,
        })
    # ==========================================
    # (6) 컨텍스트 구성 및 렌더링
    # ==========================================
    context = {
        'no_data': no_data,
        'total_count': total_count,
        'status_counts': list(status_counts.values()), # 차트용 리스트 변환
        'avg_final': round(avg_final, 1) if avg_final else 0,
        
        # 레이더 차트 순서: [시험, 실습, 노트, 태도]
        'radar_data': [
            round(radar_data['avg_exam'] or 0, 1),
            round(radar_data['avg_prac'] or 0, 1), 
            round(radar_data['avg_note'] or 0, 1), 
            round(radar_data['avg_atti'] or 0, 1)
        ],
        
        'top_trainees': top_trainees,
        'risk_trainees': risk_trainees,
        'trainee_list': trainee_list,
        
        # 필터링 드롭다운 옵션
        'cohorts': Cohort.objects.all(),
        'processes': Process.objects.all(),
        
        # 현재 선택된 필터 유지
        'sel_q': search_query,
        'sel_cohort': int(filter_cohort) if filter_cohort else '',
        'sel_process': int(filter_process) if filter_process else '',
    }

    return render(request, 'quiz/pl_dashboard.html', context)


# 2. [핵심 수정] PL 교육생 상세 리포트 (중복 제거 및 HTML 렌더링 적용)
@login_required
def pl_trainee_detail(request, profile_id):
    """
    PL용 교육생 상세 리포트 (통합 버전)
    - 기본 정보 및 권한 체크
    - 시험 이력 (전체)
    - AI 취약점 분석 (태그별 정답률)
    - 매니저 평가 (체크리스트 & 종합의견)
    - 특이사항 (로그 & 재시험 횟수)
    """
    
    # 1. 프로필 가져오기
    profile = get_object_or_404(Profile, id=profile_id)

    # 2. 권한 체크 (슈퍼유저 OR 본인 파트 PL)
    if not request.user.is_superuser:
        # PL 프로필이 있는지 확인
        if not (hasattr(request.user, 'profile') and request.user.profile.is_pl):
            messages.error(request, "접근 권한이 없습니다.")
            return redirect('quiz:index')
        
        # 본인 파트인지 확인
        try:
            pl_obj = PartLeader.objects.get(email=request.user.email)
            if profile.pl != pl_obj:
                messages.error(request, "담당 교육생이 아닙니다.")
                return redirect('quiz:pl_dashboard')
        except PartLeader.DoesNotExist:
            messages.error(request, "PL 정보를 찾을 수 없습니다.")
            return redirect('quiz:pl_dashboard')

    # 3. 시험 결과 이력 (최신순 전체)
    results = TestResult.objects.filter(user=profile.user).select_related('quiz').order_by('-completed_at')
    

    # 4. AI 취약점 분석 (태그별 정답률 계산)
    taken_quiz_ids = results.values_list('quiz_id', flat=True).distinct()
    relevant_questions = Question.objects.filter(quizzes__in=taken_quiz_ids).distinct()
    relevant_tags = Tag.objects.filter(question__in=relevant_questions).distinct()
    
    tag_analysis = []
    for tag in relevant_tags:
        # 해당 태그가 달린 문제의 총 시도 횟수
        total = UserAnswer.objects.filter(
            test_result__user=profile.user, 
            question__tags=tag
        ).count()
        
        if total > 0:
            # 그 중에서 맞춘 횟수
            correct = UserAnswer.objects.filter(
                test_result__user=profile.user, 
                question__tags=tag, 
                is_correct=True
            ).count()
            
            accuracy = (correct / total) * 100
            tag_analysis.append({
                'name': tag.name, 
                'accuracy': accuracy, 
                'status': 'weak' if accuracy < 60 else 'strong' if accuracy >= 80 else 'normal'
            })
    
    # 정답률 낮은 순(취약점 우선) 또는 높은 순 정렬
    tag_analysis.sort(key=lambda x: x['accuracy'])

    # 5. 매니저 평가 데이터 (체크리스트 & 종합의견)
    # 가장 최근 평가 1건을 가져옴
    manager_eval = ManagerEvaluation.objects.filter(trainee_profile=profile).last()
    
    checklist = []
    manager_comment = None
    
    if manager_eval:
        # 체크리스트 항목 가져오기 (카테고리 순서대로 정렬 추천)
        checklist = manager_eval.selected_items.all().select_related('category').order_by('category__order')
        manager_comment = manager_eval.overall_comment

    # 6. 특이사항 및 통계
    # (A) 로그 기록 가져오기
    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')
    warning_count = logs.filter(log_type='warning').count()
    warning_letter_count = logs.filter(log_type='warning_letter').count() # 필드명 주의 (letter vs warning_letter)
    
    # (B) 재시험 횟수 계산
    # (같은 퀴즈를 몇 번 응시했는지 카운트)
    quiz_counts = TestResult.objects.filter(user=profile.user).values('quiz').annotate(attempt_cnt=Count('id'))
    
    retake_2_count = 0 
    retake_3_count = 0 
    
    for q in quiz_counts:
        if q['attempt_cnt'] == 2:   # 정확히 2번 응시 (재시험 1회)
            retake_2_count += 1
        elif q['attempt_cnt'] >= 3: # 3번 이상 응시 (재시험 2회 이상)
            retake_3_count += 1

    # 7. 템플릿 전달
    context = {
        'profile': profile,
        'results': results,
        'tag_analysis': tag_analysis,
        
        # 매니저 평가 관련
        'checklist': checklist,          # 템플릿에서 {% for item in checklist %} 사용
        'manager_comment': manager_comment, # 템플릿에서 {{ manager_comment }} 사용
        
        # 특이사항 관련
        'logs': logs,
        'warning_count': warning_count,
        'warning_letter_count': warning_letter_count,
        'retake_2_count': retake_2_count,
        'retake_3_count': retake_3_count,
    }
    
    return render(request, 'quiz/pl_trainee_detail.html', context)

# 3. PL 리포트 출력 뷰 (인쇄용)
@login_required
def pl_report_view(request):
    if not (request.user.is_staff and (request.user.profile.is_pl or request.user.is_superuser)):
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    if request.user.is_superuser:
        trainees = Profile.objects.select_related('user', 'cohort', 'process').all()
    else:
        try:
            pl_obj = PartLeader.objects.get(email=request.user.email)
            trainees = Profile.objects.filter(pl=pl_obj).select_related('user', 'cohort', 'process')
        except PartLeader.DoesNotExist:
            trainees = Profile.objects.none()

    # 데이터 구성
    all_quizzes = Quiz.objects.all().order_by('title')
    report_data = []

    for t in trainees:
        results = t.user.testresult_set.all().order_by('completed_at')
        scores_list = []
        for quiz in all_quizzes:
            attempts = results.filter(quiz=quiz)
            s1 = attempts[0].score if attempts.count() >= 1 else '-'
            s2 = attempts[1].score if attempts.count() >= 2 else '-'
            s3 = attempts[2].score if attempts.count() >= 3 else '-'
            scores_list.append({'title': quiz.title, 's1': s1, 's2': s2, 's3': s3})
        
        fa = getattr(t, 'final_assessment', None)
        report_data.append({
            'profile': t,
            'scores': scores_list,
            'assessment': {
                'final_score': fa.final_score if fa else '-',
                'rank': fa.rank if fa else '-',
                'comment': fa.manager_comment if fa else "의견 없음"
            }
        })

    return render(request, 'quiz/pl_report_print.html', {'report_data': report_data, 'today': timezone.now().date()})

# --- 1. 최종 점수 및 랭킹 계산 유틸리티 ---

def calculate_cohort_ranking(cohort_id):
    """특정 기수 내 최종 점수 기준으로 등수를 매기는 함수"""
    
    # 1. 기수 내 모든 FinalAssessment 가져오기 (최종 점수 기준으로 정렬)
    assessments = FinalAssessment.objects.filter(profile__cohort__id=cohort_id).order_by('-final_score')
    
    # 2. 랭킹 계산 (DenseRank 사용: 동점자에게 같은 등수를 부여합니다)
    ranked_assessments = assessments.annotate(
        rank=Window(
            expression=DenseRank(),
            order_by=[F('final_score').desc()]
        )
    )
    
    # 3. DB에 순위 반영
    for assessment in ranked_assessments:
        # 이미 랭킹이 계산된 값이 annotate 되어 있으므로 그대로 저장
        assessment.rank = assessment.rank
        assessment.save(update_fields=['rank'])


# --- 2. 랭킹 일괄 업데이트 (모든 기수) ---

def update_all_cohort_rankings():
    """DB에 있는 모든 기수의 랭킹을 일괄 계산하여 반영합니다."""
    cohort_ids = Cohort.objects.all().values_list('id', flat=True)
    for cohort_id in cohort_ids:
        calculate_cohort_ranking(cohort_id)

@login_required
def request_process_access(request):
    if request.method == 'POST':
        target_id = request.POST.get('target_process_id')
        
        # target_id가 'ALL'이면 전체 요청 (target_process=None)
        target_process = None
        target_name = "🌍 전체 공정"
        
        if target_id and target_id != 'ALL':
            target_process = get_object_or_404(Process, pk=target_id)
            target_name = target_process.name

        # 중복 요청 확인
        existing = ProcessAccessRequest.objects.filter(
            requester=request.user, 
            target_process=target_process, # None이면 전체 검색
            status__in=['pending', 'approved']
        ).first()
        
        if existing:
            msg_status = "승인되었습니다" if existing.status == 'approved' else "대기 중입니다"
            messages.warning(request, f"이미 '{target_name}' 권한이 {msg_status}.")
        else:
            ProcessAccessRequest.objects.create(
                requester=request.user,
                target_process=target_process # None이면 전체
            )
            from quiz.models import Notification
            from django.urls import reverse
            
            superusers = User.objects.filter(is_superuser=True)
            for admin in superusers:
                Notification.objects.create(
                    recipient=admin,
                    sender=request.user,
                    message=f"🔑 [권한 요청] {request.user.profile.name}님이 '{target_name}' 열람 권한을 요청했습니다.",
                    related_url=reverse('quiz:manage_access_requests'), # 클릭 시 권한 승인 창으로 이동
                    icon='bi-key-fill',
                    notification_type='general'
                )
            # =========================================================
            
            messages.success(request, f"'{target_name}' 열람 권한을 요청했습니다.")

    return redirect(request.META.get('HTTP_REFERER', 'quiz:dashboard'))

# 2. 요청 관리 페이지 (최고 관리자 전용)
@login_required
def manage_access_requests(request):
    if not request.user.is_superuser:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:dashboard')
        
    pending_requests = ProcessAccessRequest.objects.filter(status='pending').order_by('-created_at')
    
    return render(request, 'quiz/manage_access_requests.html', {'requests': pending_requests})

# 3. 승인/거절 처리 (최고 관리자 전용)
@login_required
def approve_access_request(request, request_id, action):
    if not request.user.is_superuser:
        return redirect('quiz:dashboard')
        
    from accounts.models import ProcessAccessRequest
    from django.utils import timezone
    from datetime import timedelta
        
    access_req = get_object_or_404(ProcessAccessRequest, pk=request_id)
    
    if action == 'approve':
        # HTML 폼에서 'hours' 값을 받아옵니다. (기본값 24)
        try:
            hours = int(request.POST.get('hours', 24))
        except ValueError:
            hours = 24
            
        access_req.status = 'approved'
        # ★ 지금 시간 + 선택한 시간(hours)을 계산해서 만료 시간으로 저장!
        access_req.expires_at = timezone.now() + timedelta(hours=hours)
        access_req.save()
        messages.success(request, f"{access_req.requester.profile.name}님의 요청을 {hours}시간 동안 승인했습니다.")
        
    elif action == 'reject':
        access_req.status = 'rejected'
        access_req.save()
        messages.warning(request, "요청을 거절했습니다.")
        
    return redirect('quiz:manage_access_requests')

@login_required
def manage_interviews(request, profile_id):
    """
    [구버전 호환용]
    예전 면담 페이지 URL로 접속 시, 새로운 '특이사항/경고 관리' 페이지로 이동시킵니다.
    """
    return redirect('quiz:manage_student_logs', profile_id=profile_id)

@login_required
def manager_quiz_list(request):
    """매니저용 시험 목록 관리"""
    if not request.user.is_staff: 
        return redirect('quiz:index')
    
    # 관리자는 전체 보기
    if request.user.is_superuser:
        quizzes = Quiz.objects.all().order_by('-id')

    # 매니저는 (공통 + 자기공정) 보기
    elif hasattr(request.user, 'profile') and request.user.profile.process:
        my_process = request.user.profile.process
        quizzes = Quiz.objects.filter(
            # [수정] associated_process -> related_process 로 변경
            Q(category=Quiz.Category.COMMON) | Q(related_process=my_process)
        ).distinct().order_by('-id')

    else:
        # 공정 정보가 없는 매니저는 공통만 보기
        quizzes = Quiz.objects.filter(category=Quiz.Category.COMMON).order_by('-id')

    return render(request, 'quiz/manager/quiz_list.html', {'quizzes': quizzes})

# ==================================================================
# 1. 시험 생성 함수 (Create) - 수동 처리 방식
# ==================================================================
# quiz/views.py

@login_required
def quiz_create(request):
    # 1. 관리자 권한 체크
    if not request.user.is_staff:
        messages.error(request, "관리자 권한이 필요합니다.")
        return redirect('quiz:index')

    # 2. [POST 요청] 데이터 저장
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            description = request.POST.get('description')
            category = request.POST.get('category')
            
            process_id = request.POST.get('related_process')
            process_instance = None
            if process_id and process_id.strip():
                process_instance = Process.objects.filter(id=process_id).first()

            q_count = request.POST.get('question_count') or 25
            p_score = request.POST.get('pass_score') or 80
            t_limit = request.POST.get('time_limit') or 30
            is_pub = request.POST.get('is_published') == 'on'

            # ★ [일단 작성자 없이 시도해봅니다]
            # 만약 작성자 필드가 필수라면 여기서 에러가 나겠지만,
            # 적어도 필드 이름 때문에 튕기지는 않게 일단 뺍니다.
            new_quiz = Quiz.objects.create(
                title=title,
                description=description,
                category=category,
                related_process=process_instance,
                question_count=int(q_count),
                pass_score=int(p_score),
                time_limit=int(t_limit),
                is_published=is_pub
                # recorder=request.user  <-- 삭제 (범인 후보 1)
                # created_by=request.user <-- 삭제 (범인 후보 2)
            )

            print(f"✅ 시험 생성 성공: {new_quiz.title}")
            messages.success(request, f"새 시험 '{title}'이(가) 생성되었습니다.")
            return redirect('quiz:question_list', quiz_id=new_quiz.id)

        except Exception as e:
            # 🚨 여기서 정답을 알아냅니다!
            print(f"\n❌ [생성 실패] 에러 내용: {e}")
            
            # Quiz 모델에 진짜 있는 필드 이름들을 출력해봅니다.
            print("\n🔍 [Quiz 모델의 실제 필드 목록]")
            try:
                field_names = [field.name for field in Quiz._meta.get_fields()]
                print(f"👉 {field_names}\n")
            except:
                print("👉 필드 목록을 가져오지 못했습니다.")

            messages.error(request, "시험 생성 중 오류가 발생했습니다. (터미널 로그를 확인해주세요)")
            return redirect('quiz:quiz_create')

    # 3. [GET 요청] 화면 표시
    processes = Process.objects.all()
    return render(request, 'quiz/manager/quiz_form.html', {
        'title': '새 시험 생성', 
        'processes': processes
    })

# ==================================================================
# 2. 시험 수정 함수 (Update) - 수동 처리 방식
# ==================================================================
@login_required
def quiz_update(request, quiz_id):
    # 1. 관리자 권한 체크
    if not request.user.is_staff:
        messages.error(request, "관리자 권한이 필요합니다.")
        return redirect('quiz:index')
    
    # 수정할 객체 가져오기
    quiz = get_object_or_404(Quiz, pk=quiz_id)

    # 2. [POST 요청] 데이터 수정 저장
    if request.method == 'POST':
        try:
            # (1) 텍스트 데이터 업데이트
            quiz.title = request.POST.get('title')
            quiz.description = request.POST.get('description')
            quiz.category = request.POST.get('category')
            
            # (2) 공정(Process) 연결 로직
            process_id = request.POST.get('related_process')
            if process_id and process_id.strip():
                quiz.related_process = Process.objects.filter(id=process_id).first()
            else:
                quiz.related_process = None

            # (3) ★ 숫자 데이터 업데이트 (핵심 로직)
            # 체크박스 체크됨 -> input disabled -> 값 안 넘어옴(None) -> 'or 25' (기본값 적용)
            # 체크박스 해제됨 -> input enabled -> 값 넘어옴(예: 70) -> 'or 25' 무시하고 70 사용
            q_count = request.POST.get('question_count') or 25
            p_score = request.POST.get('pass_score') or 80
            t_limit = request.POST.get('time_limit') or 30

            quiz.question_count = int(q_count)
            quiz.pass_score = int(p_score)
            quiz.time_limit = int(t_limit)

            quiz.is_published = (request.POST.get('is_published') == 'on')

            # (4) 저장
            quiz.save()
            
            messages.success(request, f"시험 '{quiz.title}' 정보가 수정되었습니다.")
            
            # 수정 후에는 보통 '문제 목록'이나 '시험 목록'으로 이동합니다.
            # (원하시는 곳으로 연결하세요. 여기선 시험 목록으로 보냅니다.)
            return redirect('quiz:manager_quiz_list')

        except Exception as e:
            print(f"❌ Quiz Update Error: {e}")
            messages.error(request, "수정 중 오류가 발생했습니다. 입력값을 확인해주세요.")
            # 에러 발생 시 수정 페이지에 머무름
            return redirect('quiz:quiz_update', quiz_id=quiz.id)

    # 3. [GET 요청] 수정 화면 표시
    processes = Process.objects.all()
    
    return render(request, 'quiz/manager/quiz_form.html', {
        'quiz': quiz,          # ★ 중요: 저장된 값(70점 등)을 HTML로 전달
        'title': '시험 설정 수정',
        'processes': processes
    })

@login_required
def quiz_delete(request, quiz_id):
    if not request.user.is_staff: return redirect('quiz:index')
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    if request.method == 'POST':
        quiz.delete()
        messages.success(request, "시험이 삭제되었습니다.")
    
    return redirect('quiz:manager_quiz_list')

# --- 문제(Question) 관리 뷰 ---

@login_required
def question_list(request, quiz_id):
    if not request.user.is_staff: return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # [수정됨] 1:N 방식(question_set) -> M:N 방식(questions)으로 변경
    # 이제 문제는 'quiz.questions'를 통해 가져와야 합니다.
    questions = quiz.questions.all().order_by('-created_at')
    
    return render(request, 'quiz/manager/question_list.html', {'quiz': quiz, 'questions': questions})

# ------------------------------------------------------------------
# 문제 등록 (Create)
# ------------------------------------------------------------------
@login_required
def question_create(request, quiz_id):
    if not request.user.is_staff:
        return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)

    if request.method == 'POST':
        try:
            # 1. 문제 생성
            question = Question.objects.create(
                question_text=request.POST.get('question_text'),
                question_type=request.POST.get('question_type'),
                difficulty=request.POST.get('difficulty')
            )
            question.quizzes.add(quiz)

            if request.FILES.get('question_image'):
                question.image = request.FILES['question_image']
                question.save()

            # 2. 태그 저장 (JSON 파싱 + 일반 콤마 지원)
            tags_input = request.POST.get('tags', '')
            if tags_input:
                try:
                    # Tagify가 보낸 JSON ([{"value":"태그1"}]) 처리
                    tag_list = json.loads(tags_input)
                    for item in tag_list:
                        t_name = item.get('value', '').strip()
                        if t_name:
                            tag_obj, _ = Tag.objects.get_or_create(name=t_name)
                            question.tags.add(tag_obj)
                except json.JSONDecodeError:
                    # JSON이 아닐 경우 콤마로 분리
                    for t in tags_input.split(','):
                        if t.strip():
                            tag_obj, _ = Tag.objects.get_or_create(name=t.strip())
                            question.tags.add(tag_obj)

            # 3. 정답/보기 처리
            q_type = question.question_type

            # (A) 주관식: 정답이 여러 개(예: 사과, 과자)인 경우만 콤마로 구분
            # Apple/apple 같은 대소문자는 채점할 때 처리하므로 하나만 입력해도 됨.
            if q_type == 'short_answer':
                # ★ HTML에서 name="short_answers[]" 로 보낸 여러 개의 칸 값을 배열로 받습니다!
                answers = request.POST.getlist('short_answers[]')
                
                # (혹시 기존의 단일 칸으로 들어올 경우를 대비한 안전장치)
                old_single_answer = request.POST.get('correct_answer_text', '')
                if old_single_answer and not answers:
                    answers = [old_single_answer]
                    
                for ans in answers:
                    if ans.strip():
                        Choice.objects.create(question=question, choice_text=ans.strip(), is_correct=True)

            # (B) 객관식 (단일/복수)
            elif q_type in ['multiple_choice', 'multiple_select']:
                for i in range(1, 5):
                    c_text = request.POST.get(f'choice_text_{i}', '').strip()
                    c_img = request.FILES.get(f'choice_image_{i}')
                    # 체크박스 값 확인 ('on'이면 True)
                    is_corr = request.POST.get(f'is_correct_{i}') == 'on'

                    if c_text or c_img:
                        Choice.objects.create(
                            question=question, choice_text=c_text, image=c_img, is_correct=is_corr
                        )

            # (C) OX
            elif q_type == 'true_false':
                ox_val = request.POST.get('ox_answer')
                Choice.objects.create(question=question, choice_text='O', is_correct=(ox_val == 'O'))
                Choice.objects.create(question=question, choice_text='X', is_correct=(ox_val == 'X'))

            messages.success(request, "새 문제가 등록되었습니다.")
            return redirect('quiz:question_list', quiz_id=quiz.id)

        except Exception as e:
            messages.error(request, f"오류 발생: {e}")
    
    # 태그 검색용 리스트
    all_tags_list = list(Tag.objects.values_list('name', flat=True))

    return render(request, 'quiz/manager/question_form.html', {'quiz': quiz})


# ------------------------------------------------------------------
# 문제 수정 (Update)
# ------------------------------------------------------------------
@login_required
def question_update(request, question_id):
    if not request.user.is_staff: 
        return redirect('quiz:index')
    
    question = get_object_or_404(Question, pk=question_id)
    related_quiz = question.quizzes.first()

    if request.method == 'POST':
        try:
            # 1. 정보 업데이트
            question.question_text = request.POST.get('question_text')
            question.question_type = request.POST.get('question_type')
            new_difficulty = request.POST.get('difficulty')
            if new_difficulty:
                question.difficulty = new_difficulty
            
            # [수정] 이미지 업데이트 (파일이 새로 업로드된 경우에만 교체)
            if 'question_image' in request.FILES:
                question.image = request.FILES['question_image']
            
            question.save()

            # 2. [변경] 태그 업데이트 (Tagify JSON 처리)
            # (기존 로직 유지)
            question.tags.clear()
            tags_json = request.POST.get('tags', '')
            if tags_json:
                try:
                    tag_list = json.loads(tags_json)
                    for tag_item in tag_list:
                        t_name = tag_item.get('value', '').strip()
                        if t_name:
                            tag_obj, _ = Tag.objects.get_or_create(name=t_name)
                            question.tags.add(tag_obj)
                except json.JSONDecodeError:
                    for t in tags_json.split(','):
                        if t.strip():
                            tag_obj, _ = Tag.objects.get_or_create(name=t.strip())
                            question.tags.add(tag_obj)

            # 3. 보기/정답 처리
            q_type = question.question_type

            # (A) 주관식 / OX -> 기존처럼 삭제 후 재생성 (간단하므로)
            if q_type in ['short_answer', 'true_false']:
                question.choice_set.all().delete()
                
                if q_type == 'short_answer':
                    # ★ 여기서부터 안쪽으로 쏙 들어가야 합니다!
                    answers = request.POST.getlist('short_answers[]')
                    
                    # (혹시 기존의 단일 칸으로 들어올 경우를 대비한 안전장치)
                    old_single_answer = request.POST.get('correct_answer_text', '')
                    if old_single_answer and not answers:
                        answers = [old_single_answer]
                        
                    for ans in answers:
                        if ans.strip():
                            Choice.objects.create(question=question, choice_text=ans.strip(), is_correct=True)
                
                elif q_type == 'true_false':
                    ox_val = request.POST.get('ox_answer')
                    Choice.objects.create(question=question, choice_text='O', is_correct=(ox_val == 'O'))
                    Choice.objects.create(question=question, choice_text='X', is_correct=(ox_val == 'X'))
            # (B) [핵심 수정] 객관식 (이미지 보존을 위해 Update 방식 사용)
            elif q_type in ['multiple_choice', 'multiple_select']:
                # 기존 보기들을 ID 순서대로 가져옴 (최대 4개라고 가정)
                old_choices = list(question.choice_set.all().order_by('id'))
                
                # 1번부터 4번까지 루프
                for i in range(1, 5):
                    c_text = request.POST.get(f'choice_text_{i}', '').strip()
                    c_file = request.FILES.get(f'choice_image_{i}') # 파일 객체 (HTML name="choice_image_1" 등)
                    is_corr = request.POST.get(f'is_correct_{i}') == 'on'

                    # 텍스트나 파일이 있는 경우에만 저장 (또는 기존 데이터가 있으면 수정)
                    if c_text or c_file or (i <= len(old_choices)):
                        
                        # 해당 순서에 기존 보기가 있으면 -> 수정 (Update)
                        if i <= len(old_choices):
                            choice_obj = old_choices[i-1]
                            choice_obj.choice_text = c_text
                            choice_obj.is_correct = is_corr
                            
                            # ★ 핵심: 새 파일이 올라왔을 때만 이미지 교체
                            if c_file:
                                choice_obj.image = c_file
                            
                            choice_obj.save()
                        
                        # 없으면 -> 새로 생성 (Create)
                        else:
                            if c_text or c_file: # 내용이 있을 때만
                                Choice.objects.create(
                                    question=question,
                                    choice_text=c_text,
                                    image=c_file, # 새 파일 저장
                                    is_correct=is_corr
                                )
                
                # (선택 사항) 4개 넘어가거나 불필요하게 남은 기존 보기 삭제
                if len(old_choices) > 4:
                    for del_choice in old_choices[4:]:
                        del_choice.delete()

            messages.success(request, "문제가 수정되었습니다.")
            
            if related_quiz:
                return redirect('quiz:question_list', quiz_id=related_quiz.id)
            else:
                return redirect('quiz:manager_quiz_list')

        except Exception as e:
            # 에러 발생 시 로그 출력 (디버깅용)
            print(f"Update Error: {e}")
            messages.error(request, f"수정 중 오류 발생: {e}")

    # GET 요청 처리 (화면 그리기) - 기존 코드 유지
    current_tags = ",".join(question.tags.values_list('name', flat=True))
    
    short_answer_val = ""
    if question.question_type == 'short_answer':
        correct_choices = question.choice_set.filter(is_correct=True).values_list('choice_text', flat=True)
        short_answer_val = ",".join(correct_choices)
            
    ox_answer_val = ""
    if question.question_type == 'true_false':
        correct_choice = question.choice_set.filter(is_correct=True).first()
        if correct_choice:
            ox_answer_val = correct_choice.choice_text

    choices_list = list(question.choice_set.all().order_by('id'))

    while len(choices_list) < 4:
        choices_list.append(None)
        
    all_tags_list = list(Tag.objects.values_list('name', flat=True))

    return render(request, 'quiz/manager/question_form.html', {
        'question': question,
        'quiz': related_quiz,
        'title': '문제 수정',
        'current_tags': current_tags,
        'short_answer_val': short_answer_val,
        'ox_answer_val': ox_answer_val,
        'choices': choices_list,
        'is_update': True,
        'all_tags_json': json.dumps(all_tags_list) 
    })


# ------------------------------------------------------------------
# 문제 삭제 (Delete) - 수정 사항 없음
# ------------------------------------------------------------------
@login_required
def question_delete(request, question_id):
    if not request.user.is_staff: return redirect('quiz:index')
    
    question = get_object_or_404(Question, pk=question_id)
    
    related_quiz = question.quizzes.first()
    quiz_id = related_quiz.id if related_quiz else None

    if request.method == 'POST':
        question.delete()
        messages.success(request, "문제가 삭제되었습니다.")
    
    if quiz_id:
        return redirect('quiz:question_list', quiz_id=quiz_id)
    return redirect('quiz:manager_quiz_list')

# ------------------------------------------------------------------
# 평가 (Evaluate Trainee) - [기존 유지]
# ------------------------------------------------------------------
@login_required
def evaluate_trainee(request, profile_id):
    if not request.user.is_staff:
        # from django.contrib import messages  <-- 삭제!
        messages.error(request, "접근 권한이 없습니다. (관리자 전용)")
        return redirect('quiz:index')

    # 1. 대상자 조회 및 권한 체크
    trainee = get_object_or_404(Profile, pk=profile_id)
    if not is_process_manager(request.user, trainee):
        # from django.contrib import messages  <-- 삭제!
        messages.error(request, "⛔ 권한 오류: 타 공정 교육생의 최종 평가서는 담당 공정 매니저 및 최고관리자만 조회/수정할 수 있습니다.")
        return redirect('quiz:manager_trainee_detail', profile_id=trainee.id)
    
    # [1-1] 경고 및 감점 계산 로직 (새로 추가됨)
    # 경고(warning)와 경고장(warning_letter) 횟수 조회
    warnings = StudentLog.objects.filter(profile=trainee, log_type='warning').count()
    letters = StudentLog.objects.filter(profile=trainee, log_type='warning_letter').count()
    
    # [감점 공식]
    # 경고: 1회는 봐줌(0점), 2회부터 10점씩 감점
    warning_penalty = (warnings - 1) * 10 if warnings > 1 else 0
    # 경고장: 횟수당 10점 감점
    letter_penalty = letters * 10
    total_penalty = warning_penalty + letter_penalty
    
    # 최대 감점 상한선 (40점)
    if total_penalty > 40:
        total_penalty = 40

    # 2. 기존 평가 데이터 가져오기 (수정 모드)
    existing_evaluation = ManagerEvaluation.objects.filter(trainee_profile=trainee).first()
    final_assessment, _ = FinalAssessment.objects.get_or_create(profile=trainee)

    today = timezone.now().date()
    # ✅ 수정됨: end_date가 값이 '있을 때만' 날짜 비교를 하도록 방어막 추가!
    is_cohort_ended = trainee.cohort and trainee.cohort.end_date and trainee.cohort.end_date < today
    is_finalized = (trainee.status == 'completed') or (trainee.status == 'dropout' and existing_evaluation is not None)

    # 3. POST 요청 처리 (저장 및 점수 계산)
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        
        if is_finalized and action == 'finalize':
            messages.error(request, "⛔ 이미 확정 처리된 평가서는 번복하거나 수정할 수 없습니다.")
            return redirect('quiz:manager_trainee_detail', profile_id=trainee.id)

        # 1. 폼 데이터 검증 및 저장 (빈칸 우회)
        form = EvaluationForm(request.POST, instance=existing_evaluation)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.manager = request.user
            evaluation.trainee_profile = trainee
            evaluation.save()
            form.save_m2m()
        else:
            # 빈칸이 있어서 폼 검증 실패해도 텍스트/체크박스 강제 저장
            if existing_evaluation is None:
                existing_evaluation = ManagerEvaluation(trainee_profile=trainee, manager=request.user)
            existing_evaluation.overall_comment = request.POST.get('overall_comment', '')
            existing_evaluation.save()
            existing_evaluation.selected_items.set(request.POST.getlist('selected_items'))

        # 2. 점수 안전하게 변환 (빈칸이 넘어와도 에러 없이 0점 처리!)

        try:
            practice_str = request.POST.get('practice_score', '0').strip()
            attitude_str = request.POST.get('attitude_score', '0').strip()
            
            practice = float(practice_str) if practice_str else 0.0
            raw_attitude = float(attitude_str) if attitude_str else 0.0

            real_attitude = max(0, raw_attitude - total_penalty)
            exam_avg = final_assessment.exam_avg_score
            final_score = round((exam_avg * 0.85) + (practice * 0.05) + (real_attitude * 0.10), 1)

            final_assessment.practice_score = practice
            final_assessment.note_score = 0
            final_assessment.attitude_score = raw_attitude
            final_assessment.final_score = final_score
            final_assessment.save()
        except ValueError:
            messages.error(request, "점수는 숫자만 입력 가능합니다.")
            final_score = 0
            
        # 3. ★ [액션 분기] '최종 확정'을 눌렀을 때만 수료/퇴소 처리 진행
        if action == 'finalize':
            final_status = request.POST.get('final_status')
            dropout_reason = request.POST.get('dropout_reason', '').strip()

            # 강제 퇴소 대상자 여부
            is_forced_dropout = (final_score < 80) or (trainee.status == 'dropout')

            if is_forced_dropout:
                if request.user.is_superuser:
                    # [관리자(Superuser) 예외 권한]
                    if final_status == 'completed':
                        if not dropout_reason:
                            final_status = 'dropout'
                            dropout_reason = "[시스템 방어] 예외 수료 사유 미작성으로 인한 강제 퇴소"
                        else:
                            StudentLog.objects.create(
                                profile=trainee, recorder=request.user, log_type='others',
                                reason=f"[관리자 예외 수료 승인] 사유: {dropout_reason}", is_resolved=True
                            )
                else:
                    # [일반 매니저] 예외 없이 무조건 퇴소
                    final_status = 'dropout'
                    if not dropout_reason:
                        if trainee.status == 'dropout':
                            dropout_reason = "[시스템 자동] 사전 퇴소 인원 최종 평가 마감"
                        else:
                            dropout_reason = f"[시스템 자동] 최종 평가 점수 미달 ({final_score}점)"

            # 최종 상태에 따른 처리
            if final_status == 'dropout':
                trainee.status = 'dropout'
                trainee.user.is_active = False
                trainee.user.save()
                    
                StudentLog.objects.create(
                    profile=trainee, recorder=request.user, log_type='others',
                    reason=f"최종 퇴소 처리 완료\n사유: {dropout_reason}", is_resolved=True
                )
                messages.error(request, f"⛔ 점수 미달 또는 수동 조치로 인해 '{trainee.name}' 교육생이 퇴소 처리되었습니다.")
                
            elif final_status == 'completed':
                trainee.status = 'completed'
                msg = f"🎓 축하합니다! '{trainee.name}' 교육생이 성공적으로 수료 처리되었습니다."
                if is_forced_dropout and request.user.is_superuser:
                    msg += " (관리자 예외 승인 적용)"
                messages.success(request, msg)
                
            trainee.save()


            from .models import Notification
            Notification.objects.filter(notification_type='pending_eval', related_url=f"/quiz/manager/evaluate/{trainee.id}/").delete()

            return redirect('quiz:manager_trainee_detail', profile_id=trainee.id)

        else:
            # ★ 중간 저장인 경우 (현재 페이지에 튕기지 않고 알림 띄우기)
            messages.success(request, f"✅ {trainee.name}님의 평가 내용이 성공적으로 임시 저장되었습니다.")
            return redirect('quiz:evaluate_trainee', profile_id=trainee.id)

    else:
        # GET 요청 시 폼 로드
        form = EvaluationForm(instance=existing_evaluation)
        
    # 4. [종합 데이터 로드] 평가를 위한 참고 자료 (기존 코드 유지)
    # (A) 성적 현황
    test_results = TestResult.objects.filter(user=trainee.user)
    avg_score = test_results.aggregate(Avg('score'))['score__avg'] or 0
    fail_count = test_results.filter(is_pass=False).count()
    
    # (B) 근태 현황
    attendance_stats = DailySchedule.objects.filter(profile=trainee).values('work_type__name').annotate(count=Count('id'))
    
    # (C) 로그 (상벌점 등)
    logs = StudentLog.objects.filter(profile=trainee).order_by('-created_at')

    # (D) 체크리스트 항목
    categories = EvaluationCategory.objects.prefetch_related('evaluationitem_set').order_by('order')

    context = {
        'trainee': trainee,
        'form': form,
        'categories': categories,
        'final_assessment': final_assessment,
        'is_finalized': is_finalized,
        # [추가됨] 감점 관련 정보
        'total_penalty': total_penalty,

        'is_cohort_ended': is_cohort_ended,
        'is_finalized': is_finalized,

        'warning_cnt': warnings,
        'letter_cnt': letters,
        
        # 기존 참고 데이터
        'avg_score': round(avg_score, 1),
        'fail_count': fail_count,
        'attendance_stats': attendance_stats,
        'logs': logs,
    }
    return render(request, 'quiz/evaluate_trainee.html', context)


@login_required
def certificate_view(request):
    # 수료 상태가 아니면 튕겨냄
    if request.user.profile.status != 'completed':
        messages.error(request, "수료한 교육생만 수료증을 발급받을 수 있습니다.")
        return redirect('quiz:my_page')
    
    return render(request, 'quiz/certificate.html', {'profile': request.user.profile})

@login_required
def pl_report_view(request):
    # 1. 권한 및 PL 정보 확인
    if not (request.user.is_staff and (request.user.profile.is_pl or request.user.is_superuser)):
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    # 2. 대상자 필터링 (대시보드와 동일한 로직 적용)
    if request.user.is_superuser:
        trainees = Profile.objects.select_related('user', 'cohort', 'process').all()
    else:
        try:
            pl_obj = PartLeader.objects.get(email=request.user.email)
            trainees = Profile.objects.filter(pl=pl_obj).select_related('user', 'cohort', 'process')
        except PartLeader.DoesNotExist:
            trainees = Profile.objects.none()

    # 3. 검색 조건 적용 (대시보드에서 선택한 조건 그대로 가져옴)
    search_query = request.GET.get('q', '')
    filter_cohort = request.GET.get('cohort', '')
    filter_process = request.GET.get('process', '')

    if search_query:
        trainees = trainees.filter(name__icontains=search_query)
    if filter_cohort:
        trainees = trainees.filter(cohort_id=filter_cohort)
    if filter_process:
        trainees = trainees.filter(process_id=filter_process)

    # 4. [핵심] 리포트용 상세 데이터 구성 (점수 + 의견)
    all_quizzes = Quiz.objects.all().order_by('title')
    report_data = []

    for t in trainees:
        # (1) 시험 점수 상세 내역
        results = t.user.testresult_set.all().order_by('completed_at')
        scores_list = []
        
        for quiz in all_quizzes:
            attempts = results.filter(quiz=quiz)
            # 1, 2, 3차 점수 추출
            s1 = attempts[0].score if attempts.count() >= 1 else '-'
            s2 = attempts[1].score if attempts.count() >= 2 else '-'
            s3 = attempts[2].score if attempts.count() >= 3 else '-'
            scores_list.append({'title': quiz.title, 's1': s1, 's2': s2, 's3': s3})

        # (2) 종합 평가 및 매니저 의견
        fa = getattr(t, 'final_assessment', None)
        final_info = {
            'final_score': fa.final_score if fa else '-',
            'rank': fa.rank if fa else '-',
            'comment': fa.manager_comment if fa and fa.manager_comment else "작성된 평가 의견이 없습니다."
        }

        report_data.append({
            'profile': t,
            'scores': scores_list,
            'assessment': final_info
        })

    context = {
        'report_data': report_data,
        'today': timezone.now().date(),
    }
    return render(request, 'quiz/pl_report_print.html', context)


def award_badges(user, test_result):
    try:
        user_profile = user.profile
        user_badges = user_profile.badges.all()
        user_badge_names = set(user_badges.values_list('name', flat=True))
    except Profile.DoesNotExist:
        return
    except Exception as e:
        print(f"뱃지 로직 오류 (프로필 로드 실패): {e}")
        return

    badges_to_add = []
    all_badges = {badge.name: badge for badge in Badge.objects.all()}

    # [1] 첫걸음
    badge_name = '첫걸음'
    if badge_name not in user_badge_names:
        if TestResult.objects.filter(user=user).count() == 1:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [2] 퍼펙트
    badge_name = '퍼펙트'
    if test_result.score == 100 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [3] 완벽한 시작
    badge_name = '완벽한 시작'
    if test_result.score == 100 and badge_name not in user_badge_names:
        previous_100s = TestResult.objects.filter(
            user=user, score=100
        ).exclude(pk=test_result.pk).exists()
        if not previous_100s:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [4] 지니어스
    badge_name = '지니어스'
    if test_result.score >= 90 and badge_name not in user_badge_names:
        quiz_has_hard_questions = test_result.quiz.question_set.filter(difficulty='상').exists()
        if quiz_has_hard_questions:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [5] 아차상
    badge_name = '아차상'
    if (test_result.score == 98 or test_result.score == 99) and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [6] 아슬아슬
    badge_name = '아슬아슬'
    if 60 <= test_result.score <= 65 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [7] 절반의 성공
    badge_name = '절반의 성공'
    if test_result.score == 50 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [8] 괜찮아, 다시 하면 돼
    badge_name = '괜찮아, 다시 하면 돼'
    if test_result.score <= 30 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [9] 빵점...?!
    badge_name = '빵점...?!'
    if test_result.score == 0 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            badges_to_add.append(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [10] 재도전자
    badge_name = '재도전자'
    if badge_name not in user_badge_names:
        attempts_count = TestResult.objects.filter(
            user=user, quiz=test_result.quiz
        ).count()
        if attempts_count >= 3:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [11] 성실한 응시자
    badge_name = '성실한 응시자'
    if badge_name not in user_badge_names:
        total_attempts = TestResult.objects.filter(user=user).count()
        if total_attempts >= 10:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [12] 연승가도
    badge_name = '연승가도'
    if test_result.is_pass and badge_name not in user_badge_names:
        last_three_results = TestResult.objects.filter(user=user).order_by('-completed_at')[:3]
        if len(last_three_results) == 3 and all(r.is_pass for r in last_three_results):
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [13] 불사조
    badge_name = '불사조'
    if test_result.is_pass and badge_name not in user_badge_names:
        had_failed_before = TestResult.objects.filter(
            user=user, quiz=test_result.quiz, is_pass=False
        ).exists()
        if had_failed_before:
             if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [14] 노력의 결실
    badge_name = '노력의 결실'
    if badge_name not in user_badge_names:
        first_attempt = TestResult.objects.filter(
            user=user, quiz=test_result.quiz
        ).order_by('completed_at').first()
        if first_attempt and first_attempt.pk != test_result.pk:
            if test_result.score >= first_attempt.score + 30:
                if all_badges.get(badge_name):
                    badges_to_add.append(all_badges[badge_name])
                    user_badge_names.add(badge_name)

    # [15] 정복자
    badge_name = '정복자'
    if badge_name not in user_badge_names:
        all_quiz_ids = set(Quiz.objects.values_list('id', flat=True))
        attempted_quiz_ids = set(TestResult.objects.filter(user=user).values_list('quiz_id', flat=True).distinct())
        
        if all_quiz_ids and all_quiz_ids.issubset(attempted_quiz_ids):
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [16] all 100
    badge_name = 'all 100'
    if test_result.score == 100 and badge_name not in user_badge_names:
        all_quiz_ids = set(Quiz.objects.values_list('id', flat=True))
        passed_100_quiz_ids = set(TestResult.objects.filter(user=user, score=100).values_list('quiz_id', flat=True).distinct())

        if all_quiz_ids and all_quiz_ids.issubset(passed_100_quiz_ids):
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [17] 공정 마스터
    badge_name = '공정 마스터'
    if test_result.is_pass and test_result.quiz.category == Quiz.Category.PROCESS and badge_name not in user_badge_names:
        passed_process_quizzes_count = TestResult.objects.filter(
            user=user, 
            quiz__category=Quiz.Category.PROCESS, 
            is_pass=True
        ).values('quiz_id').distinct().count()
        
        if passed_process_quizzes_count >= 3:
            if all_badges.get(badge_name):
                badges_to_add.append(all_badges[badge_name])
                user_badge_names.add(badge_name)

    # [18] 꾸준함
    badge_name = '꾸준함'
    if badge_name not in user_badge_names:
        recent_test_dates = list(TestResult.objects.filter(user=user).dates('completed_at', 'day', order='DESC')[:3])
        if len(recent_test_dates) == 3:
            is_consecutive = (
                recent_test_dates[0] - timedelta(days=1) == recent_test_dates[1] and
                recent_test_dates[1] - timedelta(days=1) == recent_test_dates[2]
            )
            if is_consecutive:
                if all_badges.get(badge_name):
                    badges_to_add.append(all_badges[badge_name])
                    user_badge_names.add(badge_name)

    if badges_to_add:
        user_profile.badges.add(*badges_to_add)

    final_badge_count = len(user_badge_names) 
    
    # [19] 수집가
    badge_name = '수집가'
    if final_badge_count >= 5 and badge_name not in user_badge_names:
        if all_badges.get(badge_name):
            user_profile.badges.add(all_badges[badge_name])
            user_badge_names.add(badge_name)

    # [20] 뱃지 콜렉터
    badge_name = '뱃지 콜렉터'
    if final_badge_count >= 10 and badge_name not in user_badge_names:
         if all_badges.get(badge_name):
            user_profile.badges.add(all_badges[badge_name])
            user_badge_names.add(badge_name)


@login_required
@require_POST
def manager_create_counseling_log(request, profile_id):
    """
    매니저가 시험 결과표에서 [면담] 버튼을 눌러 바로 기록을 남길 때 사용하는 함수
    """
    if not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '권한이 없습니다.'}, status=403)

    try:
        profile = get_object_or_404(Profile, pk=profile_id)
        
        # 폼 데이터 받기
        content = request.POST.get('content')
        opinion = request.POST.get('opinion')
        is_passed = request.POST.get('is_passed') == 'on' # 체크박스 (잠금 해제용)

        if not content:
            return JsonResponse({'status': 'error', 'message': '면담 내용을 입력해주세요.'}, status=400)

        # 로그 저장 (StudentLog 사용)
        log = StudentLog.objects.create(
            profile=profile,
            recorder=request.user,  # <--- 여기 수정됨
            log_type='counseling',
            reason=content, 
            action_taken=opinion, 
            is_resolved=is_passed 
        )

        # 잠금 해제 로직 (체크 시)
        if is_passed and profile.status == 'counseling':
            profile.status = 'attending'
            profile.save()
            msg = "면담 기록 저장 및 잠금 해제 완료"
        else:
            msg = "면담 기록이 저장되었습니다."

        return JsonResponse({'status': 'success', 'message': msg})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@login_required
def student_log_detail(request, log_id):
    """
    교육생이 자신의 특이사항/경고/평가 로그의 상세 내용을 확인하는 뷰
    """
    # 본인의 로그인지 확인 (보안)
    log = get_object_or_404(StudentLog, pk=log_id, profile=request.user.profile)
    
    return render(request, 'quiz/student_log_detail.html', {'log': log})

@login_required
def quiz_question_manager(request, quiz_id):
    """
    [좌측: 내 시험지] vs [우측: 전체 문제 은행] (필터링 기능 강화 + 시험 제목 필터)
    """
    if not request.user.is_staff: return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # 1. 현재 시험에 담긴 문제들
    added_questions = quiz.questions.all().order_by('-created_at')
    
    # 2. 문제 은행 (전체 문제 - 이미 담긴 문제 제외)
    bank_questions = Question.objects.exclude(id__in=added_questions.values_list('id', flat=True)).order_by('-created_at')

    # --- [검색 및 필터링 적용] ---
    search_query = request.GET.get('q', '')
    filter_tag = request.GET.get('tag', '')
    filter_difficulty = request.GET.get('difficulty', '')
    filter_quiz = request.GET.get('quiz_filter', '') # [신규] 시험 제목 필터

    # (A) 검색어 필터 (내용)
    if search_query:
        bank_questions = bank_questions.filter(question_text__icontains=search_query)
    
    # (B) 태그 필터 (공정 등)
    if filter_tag:
        bank_questions = bank_questions.filter(tags__id=filter_tag)
        
    # (C) 난이도 필터
    if filter_difficulty:
        bank_questions = bank_questions.filter(difficulty=filter_difficulty)
        
    # (D) [신규] 특정 시험에 포함된 문제만 보기
    if filter_quiz:
        bank_questions = bank_questions.filter(quizzes__id=filter_quiz)

    bank_questions = bank_questions.distinct()

    # 페이지네이션 (문제 은행만)
    paginator = Paginator(bank_questions, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    # 필터용 데이터
    all_tags = Tag.objects.all().order_by('name')
    difficulty_choices = Question.Difficulty.choices
    
    # [신규] 필터링용 시험 목록 (현재 시험 제외)
    all_quizzes_for_filter = Quiz.objects.exclude(id=quiz_id).order_by('title')

    return render(request, 'quiz/manager/quiz_question_manager.html', {
        'quiz': quiz,
        'added_questions': added_questions,
        'bank_questions': page_obj,
        
        # 필터링 상태 유지
        'search_query': search_query,
        'filter_tag': int(filter_tag) if filter_tag else '',
        'filter_difficulty': filter_difficulty,
        'filter_quiz': int(filter_quiz) if filter_quiz else '',
        
        # 드롭다운 메뉴용 데이터
        'all_tags': all_tags,
        'difficulty_choices': difficulty_choices,
        'all_quizzes_for_filter': all_quizzes_for_filter, # 추가됨
    })

@login_required
@require_POST
def add_question_to_quiz(request):
    """AJAX: 문제 은행에서 -> 내 시험지로 담기"""
    if not request.user.is_staff: return JsonResponse({'status':'error'}, status=403)
    try:
        data = json.loads(request.body)
        quiz = get_object_or_404(Quiz, pk=data.get('quiz_id'))
        questions = Question.objects.filter(id__in=data.get('question_ids', []))
        quiz.questions.add(*questions) # M2M 추가
        return JsonResponse({'status': 'success', 'count': questions.count()})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required
@require_POST
def remove_question_from_quiz(request):
    """AJAX: 내 시험지에서 -> 문제 빼기 (삭제 아님, 관계만 끊기)"""
    if not request.user.is_staff: return JsonResponse({'status':'error'}, status=403)
    try:
        data = json.loads(request.body)
        quiz = get_object_or_404(Quiz, pk=data.get('quiz_id'))
        questions = Question.objects.filter(id__in=data.get('question_ids', []))
        quiz.questions.remove(*questions) # M2M 제거
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
@login_required
def my_notifications(request):
    """
    교육생 전용 알림/피드백 전체 목록 페이지
    """
    profile = request.user.profile
    
    # 필터링
    filter_type = request.GET.get('type', '')
    
    logs = StudentLog.objects.filter(profile=profile).order_by('-created_at')
    
    if filter_type:
        logs = logs.filter(log_type=filter_type)
        
    # 페이지네이션 (10개씩)
    paginator = Paginator(logs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    # 읽지 않은 알림 개수 (예시 로직)
    # unread_count = logs.filter(is_read=False).count() 

    return render(request, 'quiz/my_notifications.html', {
        'page_obj': page_obj,
        'filter_type': filter_type,
        'log_types': StudentLog.LOG_TYPES,
    })

@login_required
def admin_full_data_view(request):
    """
    [관리자 전용] 엑셀 스타일의 마스터 그리드 뷰 (안전한 조회 방식 적용)
    """
    if not request.user.is_superuser:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:dashboard')

    # 1. 파라미터 수신
    filter_cohort = request.GET.get('cohort', '')
    filter_process = request.GET.get('process', '')
    filter_company = request.GET.get('company', '')
    search_query = request.GET.get('q', '')
    start_date_param = request.GET.get('start_date', '')
    end_date_param = request.GET.get('end_date', '')

    # 2. [석차 계산]
    all_assessments = FinalAssessment.objects.filter(
        final_score__isnull=False
    ).select_related('profile').values(
        'profile__id', 'final_score', 
        'profile__cohort_id', 'profile__process_id', 'profile__company_id'
    )

    data_pool = list(all_assessments)
    data_pool.sort(key=lambda x: x['final_score'], reverse=True)

    rank_map = defaultdict(dict)

    # (A) 전체 석차
    curr_rank = 1
    for i, item in enumerate(data_pool):
        if i > 0 and item['final_score'] < data_pool[i-1]['final_score']:
            curr_rank = i + 1
        rank_map[item['profile__id']]['overall'] = curr_rank

    # (B) 그룹별 석차
    def calculate_group_rank(group_key, rank_name):
        grouped = defaultdict(list)
        for item in data_pool:
            grouped[item[group_key]].append(item)
        
        for g_id, items in grouped.items():
            g_rank = 1
            for i, item in enumerate(items):
                if i > 0 and item['final_score'] < items[i-1]['final_score']:
                    g_rank = i + 1
                rank_map[item['profile__id']][rank_name] = g_rank

    calculate_group_rank('profile__cohort_id', 'cohort')
    calculate_group_rank('profile__process_id', 'process')
    calculate_group_rank('profile__company_id', 'company')


    # 3. 화면 표시용 프로필 조회
    # (A) 기본 쿼리셋 생성
    profiles = Profile.objects.select_related(
        'user', 'cohort', 'company', 'process', 'pl', 'final_assessment'
    ).prefetch_related(
        'user__testresult_set', 
        'user__testresult_set__quiz',
        'dailyschedule_set__work_type',
        'managerevaluation_set__selected_items'
    )

    # (B) 정렬 로직 적용: [1순위] 최신 기수, [2순위] 내 공정, [3순위] 이름
    try:
        # 현재 로그인한 관리자의 공정 확인
        my_process = request.user.profile.process
    except AttributeError:
        my_process = None

    if my_process:
        # 내 공정이면 0, 아니면 1로 점수를 매겨서 정렬
        profiles = profiles.annotate(
            is_my_process=Case(
                When(process=my_process, then=Value(0)), 
                default=Value(1),
                output_field=IntegerField(),
            )
        ).order_by(
            '-cohort__start_date',  # 기수 내림차순 (최신 기수 위로)
            'is_my_process',        # 내 공정 우선
            'name'                  # 이름순
        )
    else:
        # 공정이 없는 관리자(슈퍼유저 등)는 그냥 최신 기수 -> 이름순
        profiles = profiles.order_by('-cohort__start_date', 'process__name', 'name')

    # 필터 적용
    if filter_cohort: profiles = profiles.filter(cohort_id=filter_cohort)
    if filter_process: profiles = profiles.filter(process_id=filter_process)
    if filter_company: profiles = profiles.filter(company_id=filter_company)
    
    if start_date_param: profiles = profiles.filter(joined_at__gte=start_date_param)
    if end_date_param: profiles = profiles.filter(joined_at__lte=end_date_param)
    
    if search_query:
        profiles = profiles.filter(
            Q(name__icontains=search_query) | 
            Q(user__username__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        )

    # 4. 데이터 가공
    all_quizzes = Quiz.objects.all().order_by('title')
    table_rows = []

    for p in profiles:
        # (1) 퀴즈 점수
        user_results = p.user.testresult_set.all()
        result_map = defaultdict(list)
        for r in user_results:
            result_map[r.quiz.id].append(r)
        
        ordered_scores = [] 
        for quiz in all_quizzes:
            attempts = sorted(result_map[quiz.id], key=lambda x: x.completed_at)
            scores_pkg = []
            for i in range(3):
                if i < len(attempts):
                    scores_pkg.append({'val': attempts[i].score, 'is_pass': attempts[i].is_pass})
                else:
                    scores_pkg.append({'val': '-', 'is_pass': False})
            ordered_scores.append(scores_pkg)

        # (2) 근태 (기수 기간 한정)
        schedules = p.dailyschedule_set.all()
        
        if p.cohort:
            if p.cohort.start_date:
                schedules = schedules.filter(date__gte=p.cohort.start_date)
            if p.cohort.end_date:
                schedules = schedules.filter(date__lte=p.cohort.end_date)

        w_cnt = schedules.filter(work_type__deduction=0).count()
        l_cnt = schedules.filter(work_type__deduction=1.0).count()
        h_cnt = schedules.filter(work_type__deduction=0.5).count()
        
        # (3) 로그 및 평가
        # ★★★ [핵심 수정] 역참조 이름 몰라도 되는 '직접 조회' 방식 사용 ★★★
        # p.studentlog_set.all() 대신 StudentLog 테이블에서 직접 찾습니다.
        logs_list = StudentLog.objects.filter(profile=p).order_by('-created_at')
        
        fa = getattr(p, 'final_assessment', None)
        last_eval = p.managerevaluation_set.last()
        manager_comment = last_eval.overall_comment if last_eval else ""

        checklist_items = last_eval.selected_items.all() if last_eval else []

        # (4) 석차
        my_ranks = rank_map.get(p.id, {})

        # ========================================================
        # ★ 모달 상세 표시를 위한 요약 데이터 가공
        # ========================================================
        warning_cnt = sum(1 for log in logs_list if log.log_type == 'warning')
        letter_cnt = sum(1 for log in logs_list if log.log_type == 'warning_letter')
        
        exam_details = []
        first_scores, second_scores, third_scores = [], [], []

        for quiz in all_quizzes:
            attempts = sorted(result_map[quiz.id], key=lambda x: x.completed_at)
            if attempts:
                s1 = attempts[0].score if len(attempts) >= 1 else None
                s2 = attempts[1].score if len(attempts) >= 2 else None
                s3 = attempts[2].score if len(attempts) >= 3 else None
                
                if s1 is not None: first_scores.append(s1)
                if s2 is not None: second_scores.append(s2)
                if s3 is not None: third_scores.append(s3)

                exam_details.append({
                    'title': quiz.title,
                    's1': s1, 's2': s2, 's3': s3,
                    'is_pass': attempts[-1].is_pass
                })
        
        # 건수와 평균을 모두 계산해서 넘김
        table_rows.append({
            'profile': p,
            'ordered_scores': ordered_scores,
            'attendance': {'work': w_cnt, 'leave': l_cnt, 'half': h_cnt},
            'final': fa,
            'ranks': my_ranks,
            'logs': logs_list,
            'manager_comment': manager_comment,
            'checklist': checklist_items,
            'log_count': logs_list.count(),
            # ★ 과목 수(Count)와 평균(Avg) 전달
            'first_cnt': len(first_scores),
            'first_avg': round(sum(first_scores)/len(first_scores), 1) if first_scores else 0,
            'second_cnt': len(second_scores),
            'second_avg': round(sum(second_scores)/len(second_scores), 1) if second_scores else 0,
            'third_cnt': len(third_scores),
            'third_avg': round(sum(third_scores)/len(third_scores), 1) if third_scores else 0,
            'warning_cnt': warning_cnt,
            'letter_cnt': letter_cnt,
            'exam_details': exam_details,
        })

    context = {
        'table_rows': table_rows,
        'quizzes': all_quizzes,
        'total_count': profiles.count(),
        'cohorts': Cohort.objects.all(),
        'processes': Process.objects.all(),
        'companies': Company.objects.all(),
        'sel_cohort': int(filter_cohort) if filter_cohort else '',
        'sel_process': int(filter_process) if filter_process else '',
        'sel_company': int(filter_company) if filter_company else '',
        'sel_start': start_date_param,
        'sel_end': end_date_param,
        'sel_q': search_query,
    }

    return render(request, 'quiz/manager/admin_full_data.html', context)

@login_required
def student_log_create(request, student_id):
    """
    [통합 로그 작성/수정 뷰]
    1. 간편 면담 팝업에서 호출됨
    2. 특이사항 관리 페이지에서 호출됨
    3. 핵심기능: '잠금 해제' 체크 시, 해당 학생의 'exam_fail' 로그를 찾아 해결 처리함.
    """
    if request.method == 'POST':
        student = get_object_or_404(User, pk=student_id) # 또는 Profile 모델
        profile = getattr(student, 'profile', None)

        # 폼 데이터 가져오기
        log_type = request.POST.get('log_type', 'counseling') # 기본값 면담
        content = request.POST.get('content', '')
        action_taken = request.POST.get('action_taken', '')
        
        # [핵심 1] 어떤 시험에 대한 면담인가? (HTML에서 quiz_id를 넘겨줘야 함)
        related_quiz_id = request.POST.get('related_quiz_id')
        related_quiz = None
        if related_quiz_id:
            related_quiz = get_object_or_404(Quiz, pk=related_quiz_id)

        # [핵심 2] '조치 완료 및 잠금 해제' 체크박스 확인
        # HTML에서 <input type="checkbox" name="resolve_lock"> 라고 되어 있어야 함
        should_resolve = request.POST.get('resolve_lock') == 'on'

        with transaction.atomic():
            # 1. 매니저가 작성한 '면담 기록'을 새로 생성 (이건 이력용)
            new_log = StudentLog.objects.create(
                profile=profile,
                recorder=request.user, # 작성자(매니저)
                log_type=log_type,     # 'counseling' 등
                reason=content,
                action_taken=action_taken,
                related_quiz=related_quiz,
                is_resolved=True # 면담 기록 자체는 작성 즉시 완료됨
            )

            # 2. ★ 만약 '잠금 해제'를 체크했다면? -> 기존의 'exam_fail' 로그를 찾아 해결 처리
            if should_resolve and related_quiz:
                # 이 학생의, 이 시험에 대한, 해결안된(False), 시험불합격(exam_fail) 로그를 모두 찾음
                blocking_logs = StudentLog.objects.filter(
                    profile=profile,
                    related_quiz=related_quiz,
                    log_type='exam_fail',
                    is_resolved=False
                )
                
                if blocking_logs.exists():
                    # 찾아서 전부 True로 바꿔버림 (잠금 해제)
                    # update()를 쓰면 여러 개(중복)가 있어도 한방에 해결됨
                    count = blocking_logs.update(is_resolved=True)
                    messages.success(request, f"면담 기록이 저장되었으며, {count}건의 불합격 잠금이 해제되었습니다.")
                else:
                    messages.warning(request, "면담은 저장되었으나, 해제할 잠금(불합격 기록)을 찾지 못했습니다.")
            else:
                messages.success(request, "면담 기록이 저장되었습니다.")

        # 저장 후 원래 페이지로 리다이렉트 (HTTP_REFERER 사용)
        return redirect(request.META.get('HTTP_REFERER', 'quiz:index'))

    return redirect('quiz:index')

@login_required
def notification_list(request):
    """
    내게 온 알림 목록을 보여주는 페이지
    """
    # 내 알림만 가져오기 (최신순)
    notis = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    
    # 읽지 않은 알림 개수
    unread_count = notis.filter(is_read=False).count()
    
    return render(request, 'quiz/notification_list.html', {
        'notifications': notis,
        'unread_count': unread_count
    })

# -----------------------------------------------------------
# [누락된 함수 2] 알림 읽음 처리 및 이동
# -----------------------------------------------------------
@login_required
def notification_read(request, noti_id):
    """
    알림을 클릭하면 '읽음' 처리하고 해당 링크로 이동
    """
    noti = get_object_or_404(Notification, pk=noti_id)
    
    # 본인 알림인지 확인 (남의 알림을 볼 수 없도록)
    if noti.recipient != request.user:
        messages.error(request, "권한이 없습니다.")
        return redirect('quiz:index')
        
    # 읽음 처리
    noti.is_read = True
    noti.save()
    
    # 연결된 주소(related_url)가 있으면 이동, 없으면 알림 목록으로 리다이렉트
    return redirect(noti.related_url if noti.related_url else 'quiz:notification_list')

def bulk_upload_file(request):
    return render(request, 'base.html', {'message': '기능 복구 중'})

@login_required
def exam_result(request, result_id):
    """
    시험 결과 상세 조회 뷰 (최종 수정: earned 변수 누락 해결)
    """
    # 1. TestResult 찾기
    try:
        if request.user.is_staff:
            result = get_object_or_404(TestResult, pk=result_id)
        else:
            result = get_object_or_404(TestResult, pk=result_id, user=request.user)
            
        answers = result.useranswer_set.select_related('question').all()
        quiz = result.quiz
        
    except TestResult.DoesNotExist:
        return redirect('quiz:index')

    # 2. '결과 확인했음' 처리
    if hasattr(result, 'is_viewed') and not result.is_viewed:
        result.is_viewed = True
        result.save()
        try:
            update_student_stats_force(result.user.profile)
        except: pass

    # 3. 데이터 가공
    detail_results = []
    
    # 기본 배점 계산
    total_count = answers.count()
    default_score = 100 / total_count if total_count > 0 else 0
    
    for ans in answers:
        # (1) 사용자 답안
        user_val = ""
        if hasattr(ans, 'selected_choice') and ans.selected_choice:
            user_val = ans.selected_choice.choice_text
        elif hasattr(ans, 'answer_text'):
            user_val = ans.answer_text
        elif hasattr(ans, 'short_answer_text'):
            user_val = ans.short_answer_text
            
        # (2) 진짜 정답
        real_val = getattr(ans.question, 'answer', '')
        if not real_val:
            correct_choices = ans.question.choice_set.filter(is_correct=True)
            if correct_choices.exists():
                real_val = ", ".join([c.choice_text for c in correct_choices])
            else:
                real_val = "정답 비공개"

        # (3) 문제 지문 (필드명 호환성 체크)
        q_content = getattr(ans.question, 'text', None)
        if not q_content:
            q_content = getattr(ans.question, 'question_text', None)
        if not q_content:
            q_content = getattr(ans.question, 'content', "문제 내용을 불러올 수 없습니다.")

        # (4) ★ [누락되었던 부분] 배점 및 획득 점수 계산 ★
        # 모델에 score가 있으면 쓰고, 없으면 1/N 점수(default_score) 사용
        q_score = getattr(ans.question, 'score', default_score)
        
        # 정답이면 배점만큼 획득, 아니면 0점 (이게 없어서 에러 났음)
        earned = q_score if ans.is_correct else 0

        detail_results.append({
            'question': q_content, 
            'user_answer': user_val,
            'real_answer': real_val,
            'is_correct': ans.is_correct,
            'score_earned': round(earned, 1) # 이제 earned가 정의되었으므로 에러 안 남
        })

    return render(request, 'quiz/exam_result.html', {
        'result': result, 
        'quiz': quiz,
        'detail_results': detail_results 
    })


# [Helper] 통계 강제 업데이트 함수 (지우지 마세요!)
def update_student_stats_force(profile):
    """
    프로필의 FinalAssessment(종합평가) 데이터를 강제로 최신화합니다.
    """
    try:
        # 1. 평균 계산 (Avg import 필요)
        avg_data = TestResult.objects.filter(user=profile.user).aggregate(avg=Avg('score'))
        new_avg = avg_data['avg'] if avg_data['avg'] is not None else 0
        
        # 2. 모델 연결 (Lazy Import로 순환참조 방지)
        from accounts.models import FinalAssessment
        
        # 3. 데이터 갱신
        assessment, created = FinalAssessment.objects.get_or_create(profile=profile)
        assessment.exam_avg_score = round(new_avg, 1)
        assessment.save()
        
    except Exception as e:
        print(f"⚠️ 통계 강제 업데이트 실패: {e}")

@login_required
def bulk_add_sheet_view(request):
    """
    엑셀 대량 등록 페이지 뷰
    """
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    # 등록된 모든 시험 목록을 가져옴 (드롭다운 선택용)
    quizzes = Quiz.objects.all().order_by('-created_at')

    # ★★★ 여기를 수정했습니다! (기존 파일명으로 연결) ★★★
    return render(request, 'quiz/bulk_add_sheet.html', {
        'quizzes': quizzes
    })

# =============================================================
# [7번 기능] 상태별 접속 제한 페이지 뷰
# =============================================================
@login_required
def counseling_required_view(request):
    """퇴소 안내 (기수 진행 중) or 면담 필요"""
    return render(request, 'quiz/status/counseling_required.html')

@login_required
def dropout_alert_view(request):
    """최종 퇴소 확정 (기수 종료 후)"""
    return render(request, 'quiz/status/dropout_alert.html')

@login_required
def completed_alert_view(request):
    """수료 축하 페이지"""
    return render(request, 'quiz/status/completed_alert.html')


@login_required
@require_POST
def bulk_approve_attempts(request):
    """
    시험 응시 요청 일괄 승인 처리
    - HTML의 체크박스(name='attempt_ids') 값을 받아 한 번에 처리
    - 보안: 각 요청별로 담당 공정 매니저인지 확인 후 승인
    """
    # 1. 권한 체크
    if not request.user.is_staff:
        messages.error(request, "권한이 없습니다.")
        return redirect('quiz:index')

    # 2. 체크된 ID 목록 가져오기
    attempt_ids = request.POST.getlist('attempt_ids')
    
    if not attempt_ids:
        messages.warning(request, "선택된 항목이 없습니다.")
        return redirect('quiz:manager_exam_requests')

    # 3. 승인 처리 로직
    success_count = 0
    fail_count = 0
    
    # 선택된 요청들을 가져옴 (대기중인 것만)
    attempts = QuizAttempt.objects.filter(id__in=attempt_ids, status='대기중')

    for attempt in attempts:
        target_profile = attempt.user.profile
        
        # [보안 체크] 슈퍼유저이거나, 해당 공정 담당자인 경우에만 승인
        if is_process_manager(request.user, target_profile):
            attempt.status = '승인됨'
            attempt.save()
            success_count += 1
        else:
            fail_count += 1

    # 4. 결과 메시지
    if success_count > 0:
        messages.success(request, f"✅ 총 {success_count}건의 요청을 승인했습니다.")
    
    if fail_count > 0:
        messages.error(request, f"🚫 권한이 없는 {fail_count}건은 승인되지 않았습니다.")

    # ★★★ [수정 완료] 처리 후 다시 요청 관리함으로 복귀
    return redirect('quiz:manager_exam_requests')

@login_required
def exam_analytics_dashboard(request):
    """
    [Final] PMTC 시험 분석 센터
    - 핵심 변경: 매니저는 '본인 공정' 및 '공통' 시험만 조회 (타 공정 노출 차단)
    - 관리자(Superuser)는 모든 시험 조회
    """
    if not request.user.is_staff:
        return redirect('quiz:index')

    user = request.user
    
    # 1. 퀴즈 조회 범위 설정 (필터링 핵심 로직)
    if user.is_superuser:
        # 관리자: 모든 시험 조회
        quizzes = Quiz.objects.all()
    else:
        # 매니저: '내 공정' + '공통(Process 없음)' 시험만 조회
        target_process = user.profile.process if hasattr(user, 'profile') else None
        
        quizzes = Quiz.objects.filter(
            Q(related_process=target_process) | Q(related_process__isnull=True)
        )

    # 정렬 (카테고리 -> 공정 -> 이름)
    quizzes = quizzes.select_related('related_process').order_by('category', 'related_process', 'title')
    
    # 2. 데이터를 담을 구조 (기존 유지)
    grouped_data = {
        'process': {'label': '🏭 내 공정 직무 평가', 'exams': [], 'icon': 'bi-gear-wide-connected', 'color': 'primary'},
        'common':  {'label': '📘 공통 필수 평가',   'exams': [], 'icon': 'bi-book-half', 'color': 'success'},
        'safety':  {'label': '🦺 환경/안전 교육',   'exams': [], 'icon': 'bi-shield-check', 'color': 'danger'},
        'etc':     {'label': '🧩 기타 평가',       'exams': [], 'icon': 'bi-puzzle-fill', 'color': 'secondary'},
    }

    total_stats = {'exam_count': 0, 'total_attempts': 0}

    for quiz in quizzes:
        results = TestResult.objects.filter(quiz=quiz)
        total_count = results.count()
        
        avg_score = 0
        pass_rate = 0
        
        if total_count > 0:
            avg_score = results.aggregate(Avg('score'))['score__avg'] or 0
            pass_count = results.filter(is_pass=True).count()
            pass_rate = (pass_count / total_count) * 100
            
        exam_data = {
            'id': quiz.id,
            'title': quiz.title,
            'process_name': quiz.related_process.name if quiz.related_process else "공통",
            'question_count': quiz.questions.count(),
            'total_attempts': total_count,
            'avg_score': round(avg_score, 1),
            'pass_rate': round(pass_rate, 1),
        }

        cat = quiz.category
        if cat in grouped_data:
            grouped_data[cat]['exams'].append(exam_data)
        else:
            grouped_data['etc']['exams'].append(exam_data)
            
        total_stats['exam_count'] += 1
        total_stats['total_attempts'] += total_count

    # 1. 태그별 통계 (오답률이 높은 = 정답률이 낮은 순으로 상위 5개만)
    tag_stats = Tag.objects.annotate(
        total_tries=Count('question__useranswer'),
        correct_tries=Count('question__useranswer', filter=Q(question__useranswer__is_correct=True))
    ).annotate(
        accuracy=Case(
            When(total_tries=0, then=Value(0.0)),
            default=Cast(F('correct_tries'), FloatField()) / Cast(F('total_tries'), FloatField()) * 100.0,
            output_field=FloatField()
        )
    ).filter(total_tries__gt=0).order_by('accuracy')[:5]

    # 2. 전체 문제별 통계 (마의 문항 상위 5개)
    question_stats = Question.objects.prefetch_related('quizzes').annotate(
        total_tries=Count('useranswer'),
        correct_tries=Count('useranswer', filter=Q(useranswer__is_correct=True))
    ).annotate(
        accuracy=Case(
            When(total_tries=0, then=Value(0.0)),
            default=Cast(F('correct_tries'), FloatField()) / Cast(F('total_tries'), FloatField()) * 100.0,
            output_field=FloatField()
        )
    ).filter(total_tries__gt=0).order_by('accuracy')[:5]

    context = {
        'grouped_data': grouped_data,
        'total_stats': total_stats,
        'tag_stats': tag_stats,           # <-- 이거 추가
        'question_stats': question_stats, # <-- 이거 추가
    }
    return render(request, 'quiz/manager/exam_analytics.html', context)


@login_required
def question_analytics_detail(request, quiz_id):
    """
    [NEW] 문제별 정답률 및 변별력 분석 상세 (AI 분석 화면용)
    """
    if not request.user.is_staff:
        return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    questions = quiz.questions.all().order_by('id')
    
    # 1. 변별력 계산을 위한 그룹 나누기
    # ★ [수정 1] user_id가 아니라 'id(시험지 고유번호)'를 써야 재시험 통계가 정확함
    all_results = TestResult.objects.filter(quiz=quiz).order_by('-score')
    total_count = all_results.count()
    
    cut = int(total_count * 0.3) # 상/하위 30% 커트라인
    
    top_result_ids = []
    bottom_result_ids = []

    if cut > 0:
        # 상위 30% 시험지 ID 목록
        top_result_ids = list(all_results[:cut].values_list('id', flat=True))
        # 하위 30% 시험지 ID 목록 (뒤에서부터)
        bottom_result_ids = list(all_results[total_count - cut:].values_list('id', flat=True))

    q_stats = []
    
    for q in questions:
        # 이 문제에 대한 모든 답변 가져오기
        answers = UserAnswer.objects.filter(question=q, test_result__quiz=quiz)
        total_try = answers.count()
        
        accuracy = 0.0
        discriminator = 0.0
        
        # ★ [수정 2] HTML과 맞추기 위해 기본값을 'mid'로 설정
        suggested_diff = 'mid'
        
        if total_try > 0:
            correct_cnt = answers.filter(is_correct=True).count()
            accuracy = (correct_cnt / total_try) * 100
            
            # 변별력 지수 계산 (상위권 정답률 - 하위권 정답률)
            if top_result_ids and bottom_result_ids:
                # ★ [수정 3] test_result__user__id__in 대신 test_result_id__in 사용
                top_correct = answers.filter(test_result_id__in=top_result_ids, is_correct=True).count()
                bot_correct = answers.filter(test_result_id__in=bottom_result_ids, is_correct=True).count()
                
                top_rate = top_correct / len(top_result_ids)
                bot_rate = bot_correct / len(bottom_result_ids)
                
                discriminator = round(top_rate - bot_rate, 2)
            
            # 난이도 자동 추천 로직
            if accuracy >= 80:
                suggested_diff = 'high' # 너무 쉬우니 난이도를 '상'으로 올려라
            elif accuracy <= 30:
                suggested_diff = 'low'  # 너무 어려우니 난이도를 '하'로 낮춰라
            # 그 외는 'mid' (유지)
            
        else:
            # 응시 데이터가 없을 때
            accuracy = 0
            discriminator = 0
            # DB에 저장된 현재 난이도 가져오기 (없으면 mid)
            suggested_diff = getattr(q, 'difficulty', 'mid')

        q_stats.append({
            'question': q,
            'accuracy': round(accuracy, 1),
            'discriminator': discriminator,
            'current_diff': getattr(q, 'difficulty', 'mid'),
            'suggested_diff': suggested_diff
        })

    context = {
        'quiz': quiz,
        'q_stats': q_stats
    }
    
    # ★ 확인: 파일명이 'question_analytics_detail.html' 맞는지 꼭 체크!
    return render(request, 'quiz/manager/question_analytics_detail.html', context)


@login_required
@require_POST
def auto_adjust_difficulty(request, quiz_id):
    """
    [기능] 난이도 자동 보정 시스템 (Auto-Tuning)
    - 버튼 클릭 시 정답률에 따라 상/중/하 난이도를 DB에 업데이트
    """
    if not request.user.is_staff: return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    questions = quiz.questions.all()
    updated_count = 0
    
    for q in questions:
        answers = UserAnswer.objects.filter(question=q) # 전체 데이터 기준
        total = answers.count()
        if total < 5: continue # 표본 부족 시 스킵
        
        correct = answers.filter(is_correct=True).count()
        accuracy = (correct / total) * 100
        
        new_diff = 'medium'
        if accuracy >= 80: new_diff = 'low'
        elif accuracy <= 30: new_diff = 'high'
        
        # 실제 값이랑 다를 때만 업데이트 (DB 부하 절감)
        # (모델의 choices 값이 'low', 'medium', 'high' 또는 '하', '중', '상' 인지 확인 필요)
        # 여기선 영어 코드값 기준으로 작성함. models.py 확인 후 한글이면 수정하세요.
        if q.difficulty != new_diff:
            q.difficulty = new_diff
            q.save()
            updated_count += 1
            
    messages.success(request, f"총 {updated_count}개 문제의 난이도가 AI 로직에 의해 자동 조절되었습니다.")
    return redirect('quiz:question_analytics_detail', quiz_id=quiz.id)

@login_required
def student_dashboard_excel(request):
    """
    [Excel] 교육생 성적 현황 엑셀 다운로드
    - 현재 대시보드에 적용된 필터링(회사/공정/기수/검색)을 그대로 반영하여 다운로드
    """
    # 1. 권한 체크
    user = request.user
    if not (user.is_staff or (hasattr(user, 'profile') and (user.profile.is_manager or user.profile.is_pl))):
        return HttpResponse("권한이 없습니다.", status=403)

    # 2. 필터링 데이터 가져오기 (dashboard 뷰와 동일한 로직)
    sel_company = request.GET.get('company')
    sel_process = request.GET.get('process')
    sel_cohort = request.GET.get('cohort')
    search_query = request.GET.get('q', '')

    # 3. 대상 시험(Header) 선정
    target_process = None
    if not user.is_superuser and hasattr(user, 'profile'):
        target_process = user.profile.process

    header_quizzes = Quiz.objects.filter(category='common')
    if target_process:
        header_quizzes = header_quizzes | Quiz.objects.filter(related_process=target_process)
    elif sel_process:
        header_quizzes = header_quizzes | Quiz.objects.filter(related_process_id=sel_process)
    
    header_quizzes = header_quizzes.exclude(category__in=['safety', 'etc']).order_by('category', 'title')

    # 4. 대상 교육생 조회
    profiles = Profile.objects.filter(user__is_staff=False).select_related('company', 'process', 'cohort')

    if target_process:
        profiles = profiles.filter(process=target_process)
    elif sel_process:
        profiles = profiles.filter(process_id=sel_process)
    
    if sel_company:
        profiles = profiles.filter(company_id=sel_company)
    
    if sel_cohort:
        profiles = profiles.filter(cohort_id=sel_cohort)

    if search_query:
        profiles = profiles.filter(Q(name__icontains=search_query) | Q(employee_id__icontains=search_query))
    
    # 정렬: 기수 -> 이름 순
    profiles = profiles.order_by('-cohort__start_date', 'name')

    # ---------------------------------------------------------
    # 5. 엑셀 워크북 생성 및 디자인 설정
    # ---------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "교육생 성적 현황"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # 파란색 배경
    center_align = Alignment(horizontal="center", vertical="center")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # (1) 헤더 작성
    # 기본 정보 컬럼
    headers = ["기수", "회사", "공정", "이름", "사번", "상태", "1차 평균"]
    
    # 시험별 컬럼 (시험 하나당 1차, 2차, 3차 3칸씩 차지)
    for quiz in header_quizzes:
        headers.extend([f"{quiz.title}(1차)", f"{quiz.title}(2차)", f"{quiz.title}(3차)"])

    ws.append(headers)

    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style

    # (2) 데이터 행 작성
    for p in profiles:
        row_data = [
            p.cohort.name if p.cohort else "-",
            p.company.name if p.company else "-",
            p.process.name if p.process else "-",
            p.name,
            p.employee_id,
            p.get_status_display(),
            "" # 평균 점수 자리 (계산 후 넣음)
        ]

        total_1st_score = 0
        count_1st = 0

        # 시험별 점수 채우기
        for quiz in header_quizzes:
            # 해당 퀴즈 기록 가져오기 (최대 3개)
            attempts = TestResult.objects.filter(user=p.user, quiz=quiz).order_by('completed_at')[:3]
            
            # 1차 점수 평균 계산용
            if attempts.exists():
                s1 = attempts[0].score
                total_1st_score += s1
                count_1st += 1
            
            # 1, 2, 3차 점수 리스트 만들기
            scores = ["-", "-", "-"]
            for i, att in enumerate(attempts):
                # 점수 표시 (합격이면 '80(P)', 불합격이면 '60(F)' 등으로 표시 가능, 여기선 점수만)
                scores[i] = att.score
            
            row_data.extend(scores)

        # 1차 평균 계산하여 리스트의 6번 인덱스(7번째 칸)에 삽입
        avg_1st = round(total_1st_score / count_1st, 1) if count_1st > 0 else 0
        row_data[6] = avg_1st

        # 엑셀에 행 추가
        ws.append(row_data)

    # (3) 컬럼 너비 자동 조절 (약간의 여유)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # A, B, C...
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 10 else 10

    # 6. HTTP 응답 생성
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=student_scores.xlsx'
    wb.save(response)
    
    return response



@login_required
def exam_analytics_detail(request, quiz_id):
    """
    [통합 수정본] 시험 상세 분석 (통계 + AI 난이도 추천 + 최근 기록)
    - 기존의 '관리' 기능을 이곳으로 통합
    """
    if not request.user.is_staff:
        return redirect('quiz:index')

    quiz = get_object_or_404(Quiz, pk=quiz_id)
    
    # 1. 문항 데이터 로드 (M2M 역참조 문제 해결: quiz.questions.all())
    questions = quiz.questions.all().order_by('id')
    
    # 2. 전체 응시 기록 (변별력 계산용: 점수 높은 순)
    all_results = TestResult.objects.filter(quiz=quiz).order_by('-score')
    total_attempts = all_results.count()
    
    # 3. 최근 응시 기록 (화면 우측 표시용)
    recent_results = TestResult.objects.filter(quiz=quiz).order_by('-completed_at')[:10]
    
    q_stats = []
    
    # 텍스트 변환용 맵 (화면 표시용)
    diff_map = {'high': '상', 'mid': '중', 'low': '하'}

    for q in questions:
        # 이 문항의 답변 기록 (현재 시험지 내에서만)
        answers = UserAnswer.objects.filter(question=q, test_result__quiz=quiz)
        total_ans = answers.count()
        
        accuracy = 0.0
        if total_ans > 0:
            correct_ans = answers.filter(is_correct=True).count()
            accuracy = round((correct_ans / total_ans) * 100, 1)

        # 4. 변별력 계산 (상위 50% vs 하위 50%)
        discriminator = 0
        if total_attempts >= 4:
            mid_idx = total_attempts // 2
            top_group = all_results[:mid_idx]
            low_group = all_results[mid_idx:]
            
            top_correct = answers.filter(test_result__in=top_group, is_correct=True).count()
            low_correct = answers.filter(test_result__in=low_group, is_correct=True).count()
            
            top_rate = top_correct / len(top_group) if len(top_group) > 0 else 0
            low_rate = low_correct / len(low_group) if len(low_group) > 0 else 0
            
            discriminator = round(top_rate - low_rate, 2)

        # 5. AI 난이도 추천 로직 (정답률 기반)
        # - 정답률 80% 이상 -> 쉬움 -> 난이도 '하(low)' 추천
        # - 정답률 30% 이하 -> 어려움 -> 난이도 '상(high)' 추천
        suggested_diff = 'mid'
        if accuracy >= 80: suggested_diff = 'low'
        elif accuracy <= 30: suggested_diff = 'high'
        
        # 현재 난이도 정보
        current_diff_val = getattr(q, 'difficulty', 'mid')
        
        q_stats.append({
            'question': q,
            'accuracy': accuracy,
            'discriminator': discriminator,
            
            # [중요] 템플릿 변수명과 일치시킴
            'difficulty': diff_map.get(current_diff_val, '중'), # 화면용(상/중/하)
            'current_val': current_diff_val,                   # 로직용(high/mid/low)
            'suggested_diff': suggested_diff,                  # 추천 로직용(high/mid/low)
            'sug_text': diff_map.get(suggested_diff, '중'),    # 추천 화면용(상/중/하)
            'total_try': total_ans
        })

    # 6. 정답률 낮은 순(어려운 순) 정렬
    q_stats.sort(key=lambda x: x['accuracy'])

    context = {
        'quiz': quiz,
        'q_stats': q_stats,
        'recent_results': recent_results,
    }
    # ★ 통합된 페이지로 렌더링
    return render(request, 'quiz/manager/exam_analytics_detail.html', context)


@login_required
def auto_adjust_difficulty(request, quiz_id):
    """
    [통합 수정본] 난이도 일괄 자동 보정 실행
    - 실행 후 '통계 페이지(exam_analytics_detail)'로 돌아가야 함 (관리 페이지 삭제됨)
    """
    if not request.user.is_staff:
        return redirect('quiz:index')
    
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    questions = quiz.questions.all()
    
    updated_count = 0
    
    for q in questions:
        answers = UserAnswer.objects.filter(question=q, test_result__quiz=quiz)
        total = answers.count()
        
        if total > 0:
            correct = answers.filter(is_correct=True).count()
            accuracy = (correct / total) * 100
            
            # 현재 값
            old_diff = q.difficulty
            
            # 목표 값 계산 (View 로직과 동일)
            new_diff = 'mid'
            if accuracy >= 80:
                new_diff = 'low'
            elif accuracy <= 30:
                new_diff = 'high'
            
            # [최적화] 값이 다를 때만 저장 (불필요한 DB 쓰기 방지)
            if old_diff != new_diff:
                q.difficulty = new_diff
                q.save()
                updated_count += 1
    
    # ★ [핵심 수정] 작업 완료 후 '통계 페이지(exam_analytics_detail)'로 돌아갑니다.
    # (더 이상 관리 페이지는 쓰지 않으므로 여기로 와야 합니다.)
    return redirect('quiz:exam_analytics_detail', quiz_id=quiz.id)

@login_required
@require_POST
def request_dropout(request):
    try:
        from accounts.models import Profile, DropOutRequest
        
        data = json.loads(request.body)
        trainee_id = data.get('trainee_id')
        drop_date = data.get('drop_date')
        reason = data.get('reason')

        # 1. 대상 학생 프로필 찾기
        trainee = Profile.objects.get(id=trainee_id)

        # 2. 퇴사 요청서 생성 (pending 상태)
        DropOutRequest.objects.create(
            trainee=trainee,
            requester=request.user,
            drop_date=drop_date,
            reason=reason
        )

        from django.urls import reverse # 함수 맨 위에 없으면 추가
        
        superusers = User.objects.filter(is_superuser=True)
        for admin in superusers:
            Notification.objects.create(
                recipient=admin,
                sender=request.user,
                notification_type='general', 
                message=f"🚨 [퇴사 결재 대기] {trainee.name} 교육생의 중도 퇴사 요청이 접수되었습니다.",
                related_url="/quiz/admin/dropout/requests/" # ✅ 새로 만든 전용 결재함 주소로 변경!
            )

        return JsonResponse({"status": "success", "message": "성공적으로 최고 관리자에게 퇴소 결재가 요청되었습니다."})
        
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)

@login_required
def dropout_request_list(request):
    """최고 관리자용 중도 퇴사 결재함 페이지"""
    if not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, "최고 관리자만 접근할 수 있는 페이지입니다.")
        return redirect('quiz:index')

    from accounts.models import DropOutRequest
    # 대기 중인 요청과 처리 완료된 요청을 분리해서 가져옵니다
    pending_requests = DropOutRequest.objects.filter(status='pending').order_by('-created_at')
    resolved_requests = DropOutRequest.objects.exclude(status='pending').order_by('-resolved_at')[:50] # 최근 50건만

    context = {
        'pending_requests': pending_requests,
        'resolved_requests': resolved_requests,
    }
    return render(request, 'quiz/manager/dropout_approval_list.html', context)

@login_required
@require_POST
def approve_dropout(request, req_id):
    """중도 퇴사 승인 처리"""
    if not request.user.is_superuser:
        return JsonResponse({"status": "error", "message": "권한이 없습니다."})
    
    from accounts.models import DropOutRequest
    from quiz.models import Notification

    try:
        req = DropOutRequest.objects.get(id=req_id)
        req.status = 'approved'
        req.resolved_at = timezone.now()
        req.save() # 저장되는 순간 Signal이 발동하여 학생 상태가 'dropout'으로 변합니다.

        Notification.objects.create(
            recipient=req.requester, # 퇴사를 요청했던 매니저
            sender=request.user,     # 최고 관리자
            notification_type='general',
            message=f"📢 [{req.trainee.name}] 교육생의 중도 퇴사가 승인되었습니다. 최종 평가(퇴소 사유 등)를 작성해 마감해 주세요.",
            related_url=f"/quiz/manager/evaluate/{req.trainee.id}/" # 평가서 작성 페이지로 바로 이동!
        )
        
        return JsonResponse({"status": "success", "message": f"{req.trainee.name} 교육생의 퇴소가 승인되었습니다."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required
@require_POST
def reject_dropout(request, req_id):
    """중도 퇴사 반려 처리"""
    if not request.user.is_superuser:
        return JsonResponse({"status": "error", "message": "권한이 없습니다."})
    
    from accounts.models import DropOutRequest
    try:
        req = DropOutRequest.objects.get(id=req_id)
        req.status = 'rejected'
        req.resolved_at = timezone.now()
        req.save()
        
        return JsonResponse({"status": "success", "message": "퇴사 요청이 반려되었습니다."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required
def cohort_final_report(request, cohort_id):
    """기수별 최종 마감 종합 리포트 (전체 기수 조회 & 상세 점수 포함)"""
    
    user = request.user
    user_profile = getattr(user, 'profile', None)
    if not (user.is_staff or (user_profile and (user_profile.is_manager or user_profile.is_pl))):
        from django.contrib import messages
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    from accounts.models import Cohort, Profile, DropOutRequest, Process, Company, FinalAssessment
    from quiz.models import TestResult, UserAnswer
    
    all_cohorts = Cohort.objects.all().order_by('-start_date')
    f_process = request.GET.get('process', '')
    f_company = request.GET.get('company', '')

    # =========================================================
    # ★ [보안] 타 공정 세부 조회 권한 체크
    # =========================================================
    access_denied = False
    target_process_id = ''
    f_process_clean = f_process.strip() # 공백 제거 방어막

    # 1. 특정 공정(f_process)을 선택했고, 최고관리자가 아닐 때
    if f_process_clean and not request.user.is_superuser:
        my_process_name = user_profile.process.name if user_profile and user_profile.process else ''
        
        # 2. 내 공정이 아니면 티켓 검사
        if f_process_clean != my_process_name:
            target_proc = Process.objects.filter(name=f_process_clean).first()
            if target_proc:
                target_process_id = target_proc.id
                from accounts.models import ProcessAccessRequest
                
                has_ticket = ProcessAccessRequest.objects.filter(
                    requester=request.user, 
                    status='approved',
                    expires_at__gte=timezone.now()
                ).filter(Q(target_process=target_proc) | Q(target_process__isnull=True)).exists()
                
                if not has_ticket:
                    access_denied = True
            else:
                access_denied = True # 없는 공정을 입력하면 얄짤없이 차단

    # [핵심] cohort_id가 0이면 "모든 기수(전체)"로 간주합니다.
    if cohort_id == 0:
        cohort = None
        base_students = Profile.objects.filter(
            is_manager=False, is_pl=False, is_approved=True
        ).exclude(cohort__isnull=True).exclude(user__is_staff=True)
    else:
        cohort = get_object_or_404(Cohort, id=cohort_id)
        base_students = Profile.objects.filter(
            cohort=cohort, is_manager=False, is_pl=False, is_approved=True
        ).exclude(user__is_staff=True)

    available_processes = Process.objects.filter(profile__in=base_students).distinct()
    available_companies = Company.objects.filter(profile__in=base_students).distinct()

    students = base_students
    if f_process: students = students.filter(process__name=f_process)
    if f_company: students = students.filter(company__name=f_company)

    total_students = students.count()
    completed_count = students.filter(status='completed').count()
    completion_rate = round((completed_count / total_students) * 100, 1) if total_students > 0 else 0

    # 중도 퇴사 vs 미수료 분리
    approved_dropout_ids = DropOutRequest.objects.filter(
        trainee__in=students, status='approved'
    ).values_list('trainee_id', flat=True)

    all_dropouts = students.filter(status='dropout')
    mid_drop_count = all_dropouts.filter(id__in=approved_dropout_ids).count()
    fail_drop_count = all_dropouts.exclude(id__in=approved_dropout_ids).count()

    # ★ [수정됨] 공통 평가 vs 공정 평가 상세 평균 점수 분리
    test_results = TestResult.objects.filter(user__profile__in=students)
    
    # [수정] 전체 평균 점수
    common_avg = test_results.filter(quiz__related_process__isnull=True).aggregate(Avg('score'))['score__avg'] or 0
    process_avg = test_results.filter(quiz__related_process__isnull=False).aggregate(Avg('score'))['score__avg'] or 0

    # ★ [추가] 세부 과목(시험)별 평균 점수 리스트
    common_quizzes = test_results.filter(quiz__related_process__isnull=True).values('quiz__title').annotate(avg_score=Avg('score')).order_by('quiz__title')
    process_quizzes = test_results.filter(quiz__related_process__isnull=False).values('quiz__title').annotate(avg_score=Avg('score')).order_by('quiz__title')

    # ★ [수정] 마의 오답률을 공통과 공정으로 분리!
    common_top_wrong = UserAnswer.objects.filter(
        test_result__in=test_results, test_result__quiz__related_process__isnull=True, is_correct=False
    ).values('question__id', 'question__question_text', 'question__question_type', 'question__difficulty').annotate(wrong_count=Count('id')).order_by('-wrong_count')[:5]

    process_top_wrong = UserAnswer.objects.filter(
        test_result__in=test_results, test_result__quiz__related_process__isnull=False, is_correct=False
    ).values('question__id', 'question__question_text', 'question__question_type', 'question__difficulty').annotate(wrong_count=Count('id')).order_by('-wrong_count')[:5]

    # 우수자 Top 3
    top_trainees = FinalAssessment.objects.filter(
        profile__in=students, profile__status='completed'
    ).select_related('profile', 'profile__process').order_by('-final_score')[:3]

    # 트렌드 비교 (전체 기수 보기일 때는 트렌드 비교를 숨김)
    prev_cohort = None
    trend_diff = 0
    if cohort:
        prev_cohort = Cohort.objects.filter(start_date__lt=cohort.start_date).order_by('-start_date').first()
        if prev_cohort:
            prev_students = Profile.objects.filter(cohort=prev_cohort, is_manager=False, is_pl=False, is_approved=True)
            if f_process: prev_students = prev_students.filter(process__name=f_process)
            if f_company: prev_students = prev_students.filter(company__name=f_company)
            
            prev_total = prev_students.count()
            if prev_total > 0:
                prev_completed = prev_students.filter(status='completed').count()
                prev_completion_rate = round((prev_completed / prev_total) * 100, 1)
                trend_diff = round(completion_rate - prev_completion_rate, 1)

    context = {
        'cohort': cohort,
        'is_all_cohorts': cohort_id == 0, 
        'all_cohorts': all_cohorts,
        'available_processes': available_processes,
        'available_companies': available_companies,
        'f_process': f_process,
        'f_company': f_company,
        'total_students': total_students,
        'completed_count': completed_count,
        'completion_rate': completion_rate,
        'mid_drop_count': mid_drop_count,
        'fail_drop_count': fail_drop_count,
        
        'common_avg': round(common_avg, 1),
        'process_avg': round(process_avg, 1),
        
        # ★ 새로 쪼개진 데이터들 (옛날 top_wrong_questions 대신 들어감)
        'common_quizzes': common_quizzes,
        'process_quizzes': process_quizzes,
        'common_top_wrong': common_top_wrong,
        'process_top_wrong': process_top_wrong,
        
        'top_trainees': top_trainees,
        'prev_cohort': prev_cohort,
        'trend_diff': trend_diff,
        'access_denied': access_denied,
        'target_process_id': target_process_id,
    }
    return render(request, 'quiz/manager/cohort_final_report.html', context)

@login_required
def latest_cohort_report(request):
    """
    네비게이션 바에서 접근 시, 상황에 맞는 기수 리포트로 자동 연결해주는 뷰
    - 현재 진행 중인 기수(예: 34기)가 있다면 -> 바로 이전 기수(예: 33기) 마감 리포트로 이동
    - 현재 진행 중인 기수가 없다면 -> 가장 최근에 끝난 기수 리포트로 이동
    """
    from accounts.models import Cohort
    from django.utils import timezone
    
    today = timezone.now().date()
    
    # 1. 오늘 날짜 기준으로 '진행 중인 기수'가 있는지 찾습니다.
    running_cohort = Cohort.objects.filter(start_date__lte=today, end_date__gte=today).order_by('-start_date').first()
    
    target_cohort = None
    
    if running_cohort:
        # [진행 중인 기수가 있는 경우] -> 진행 중인 기수보다 '시작일'이 과거인 이전 기수를 찾음
        target_cohort = Cohort.objects.filter(start_date__lt=running_cohort.start_date).order_by('-start_date').first()
    else:
        # [진행 중인 기수가 없는 경우] -> '종료일'이 오늘보다 과거인 것 중 가장 최근 기수를 찾음 (가장 최근에 마감된 기수)
        target_cohort = Cohort.objects.filter(end_date__lt=today).order_by('-end_date').first()
        
        # 만약 마감된 기수조차 없다면 최후의 수단으로 그냥 가장 최근에 만들어진 기수를 띄움
        if not target_cohort:
            target_cohort = Cohort.objects.order_by('-start_date').first()

    # 결과에 따라 이동
    if target_cohort:
        return redirect('quiz:cohort_final_report', cohort_id=target_cohort.id)
    else:
        from django.contrib import messages
        messages.warning(request, "등록된 기수 정보가 없어 리포트를 표시할 수 없습니다.")
        return redirect('quiz:manager_dashboard')

@login_required
def export_cohort_report_excel(request, cohort_id):
    """기수 리포트 엑셀 다운로드 (필터 조건 반영)"""
    user = request.user
    user_profile = getattr(user, 'profile', None)
    if not (user.is_staff or (user_profile and (user_profile.is_manager or user_profile.is_pl))):
        from django.contrib import messages
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    from accounts.models import Cohort, Profile, FinalAssessment

    # 1. 필터 조건 가져오기
    f_process = request.GET.get('process', '')
    f_company = request.GET.get('company', '')

    if cohort_id == 0:
        cohort_name = "전체_기수"
        students = Profile.objects.filter(is_manager=False, is_pl=False, is_approved=True).exclude(cohort__isnull=True)
    else:
        from django.shortcuts import get_object_or_404
        cohort = get_object_or_404(Cohort, id=cohort_id)
        cohort_name = cohort.name
        students = Profile.objects.filter(cohort=cohort, is_manager=False, is_pl=False, is_approved=True)

    # 필터 적용
    if f_process: students = students.filter(process__name=f_process)
    if f_company: students = students.filter(company__name=f_company)

    # 2. 엑셀 워크북(파일) 만들기
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cohort_name}_운영결과"

    # [디자인 설정]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # 3. 엑셀 헤더(첫 줄) 작성
    headers = ["기수", "소속회사", "공정", "사번", "이름", "최종 상태", "시험 평균", "실습 점수", "태도 점수", "최종 환산점수", "등수"]
    ws.append(headers)

    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 4. 학생 데이터 채워넣기
    # 학생 명단을 가져오면서 최종 평가서 데이터도 같이 가져옵니다.
    assessments = FinalAssessment.objects.filter(profile__in=students).select_related('profile', 'profile__cohort', 'profile__company', 'profile__process')
    
    # 딕셔너리로 만들어서 매칭 속도 높이기
    assessment_dict = {a.profile_id: a for a in assessments}

    for student in students.order_by('company__name', 'process__name', 'name'):
        assess = assessment_dict.get(student.id)
        
        row_data = [
            student.cohort.name if student.cohort else "-",
            student.company.name if student.company else "-",
            student.process.name if student.process else "공통",
            student.employee_id or "-",
            student.name or "-",
            student.get_status_display(),
            assess.exam_avg_score if assess else 0,
            assess.practice_score if assess else 0,
            assess.attitude_score if assess else 0,
            assess.final_score if assess else 0,
            assess.rank if assess else "-",
        ]
        ws.append(row_data)

    # 5. 열 너비 자동 맞춤
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # 예: 'A', 'B'
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 6. 파일 응답으로 반환 (다운로드 창 띄우기)
    
    import urllib.parse # 한글 파일명 인코딩용 라이브러리 추가

    current_time = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f"교육운영리포트_{cohort_name}_{current_time}.xlsx"

    # [핵심 수정] 클라우드(GitHub) 환경에서 502 에러가 나지 않도록 브라우저 표준(UTF-8)으로 파일명 포장
    encoded_filename = urllib.parse.quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    
    wb.save(response)
    return response

@login_required
@require_POST
def quick_add_warning(request):
    """대시보드 퀵 경고/특이사항 등록 (다중 선택 지원) 및 자동 알림"""
    user = request.user
    user_profile = getattr(user, 'profile', None)
    if not (user.is_staff or (user_profile and (user_profile.is_manager or user_profile.is_pl))):
        return JsonResponse({'status': 'error', 'message': '권한이 없습니다.'}, status=403)

    from accounts.models import Profile
    from quiz.models import StudentLog, Notification

    # 프론트엔드에서 "1,3,5" 처럼 쉼표로 연결되어 날아옵니다.
    profile_ids_str = request.POST.get('profile_ids')
    log_type = request.POST.get('log_type')
    reason = request.POST.get('reason')
    action_taken = request.POST.get('action_taken', '')

    if not profile_ids_str:
        return JsonResponse({'status': 'error', 'message': '선택된 교육생이 없습니다.'})

    profile_ids = profile_ids_str.split(',') # 쉼표를 기준으로 리스트로 분리
    success_names = [] # 처리 완료된 학생 이름 모음

    try:
        # 선택된 여러 명의 학생을 한 명씩 돌면서 동일하게 처리!
        for pid in profile_ids:
            profile = Profile.objects.get(id=pid)
            
            # 1. 히스토리 기록 생성
            StudentLog.objects.create(
                profile=profile,
                recorder=user,
                log_type=log_type,
                reason=reason,
                action_taken=action_taken,
                is_resolved=False
            )

            # 2. ★ [핵심 수정] 상혁님의 징계 공식 적용 (경고장 = 최소 2회 취급)
            if log_type == 'warning':
                profile.warning_count += 1
            elif log_type == 'warning_letter':
                # 경고장을 바로 때리면 무조건 2회(1차 경고장)로 점프! 이미 2회 이상이면 +1 (2차 경고장)
                if profile.warning_count < 2:
                    profile.warning_count = 2
                else:
                    profile.warning_count += 1
                
            # 3. 누적 점수에 따른 상태 자동 변경 (대시보드 기준과 100% 일치)
            if profile.warning_count == 1:
                profile.status = 'caution' # 1회: 주의 (노란색)
                
            elif profile.warning_count == 2:
                profile.status = 'counseling' # 2회(1차경고장): 주의 (노란색, 계정 잠금)
                
            elif profile.warning_count == 3:
                profile.status = 'counseling' # 3회(2차경고장): 위험 (빨간색, PL 면담 필수)
                
                # ★ 위험 단계 도달 시 매니저에게 즉시 알림!
                Notification.objects.create(
                    recipient=user,
                    notification_type='general',
                    message=f"🚨 [{profile.name}] 누적 3회(2차 경고장) 도달! 즉시 PL 면담이 필요합니다.",
                    related_url=f"/quiz/manager/trainees/{profile.id}/logs/"
                )
                broadcast_realtime_notification(user.id)
                
            elif profile.warning_count >= 4:
                profile.status = 'dropout' # 4회 이상: 강제 퇴소
                profile.user.is_active = False
                profile.user.save()
            
            profile.save()
            success_names.append(profile.name)

        # 최종 성공 메시지
        msg = f"총 {len(success_names)}명 ({', '.join(success_names)}) 특이사항 일괄 등록 및 반영 완료!"
        return JsonResponse({'status': 'success', 'message': msg})
        
    except Profile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '일부 교육생 데이터를 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# [1] 미승인 대기 화면 (기존에 다른 곳에 있다면 생략 가능)
@login_required
def status_pending(request):
    if hasattr(request.user, 'profile') and request.user.profile.is_approved:
        return redirect('quiz:index')
    # 승인 대기 페이지 템플릿 경로를 적어주세요. (없다면 임의로 생성)
    return render(request, 'accounts/pending.html') 

# [2] 퇴소자 접속 불가 화면 (기존 파일명 사용)
@login_required
def status_dropout(request):
    if hasattr(request.user, 'profile') and request.user.profile.status != 'dropout':
        return redirect('quiz:index')
    return render(request, 'accounts/dropout_alert.html')

# [3] 수료자 축하 마이페이지 (기존 파일명 사용)
@login_required
def status_graduated(request):
    if hasattr(request.user, 'profile') and request.user.profile.status != 'graduated':
        return redirect('quiz:index')
    return render(request, 'accounts/completed_alert.html', {'profile': request.user.profile})

# [4] 수료증 팝업/인쇄 화면 (새로 만든 파일명 사용)
@login_required
def print_certificate(request):
    profile = request.user.profile
    if profile.status != 'graduated':
        return redirect('quiz:index')
    
    context = {
        'profile': profile,
        'cohort_name': profile.cohort.name if profile.cohort else "미지정",
        'process_name': profile.process.name if profile.process else "미지정",
        'start_date': profile.cohort.start_date if profile.cohort else None,
        'end_date': profile.cohort.end_date if profile.cohort else None,
    }
    return render(request, 'accounts/certificate_print.html', context)

@login_required
def chat_room(request):
    return render(request, 'quiz/chat.html')

from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt  # 채팅창 비동기 통신을 위해 임시로 CSRF 면제 (실무에선 토큰 방식 권장)
@login_required
def chat_file_upload(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        # 서버의 media/chat_files/ 폴더에 안전하게 저장
        fs = FileSystemStorage()
        filename = fs.save(f"chat_files/{uploaded_file.name}", uploaded_file)
        file_url = fs.url(filename) # 다운로드 주소 생성
        
        return JsonResponse({
            'status': 'success', 
            'file_url': file_url, 
            'file_name': uploaded_file.name
        })
    return JsonResponse({'status': 'error', 'message': '파일이 없습니다.'})

from .models import ChatRoom, ChatMessage

@login_required
@xframe_options_exempt # ★ 누락 방지: 서랍장(iframe)에서 열리게 하는 필수 옵션
def chat_home(request):
    """메신저 홈 화면 (연락처 및 내 채팅방 목록 + 단톡방 생성)"""
    
    # ========================================================
    # ★ 누락 방지: 단톡방 만들기 버튼을 눌렀을 때 처리 (POST)
    # ========================================================
    if request.method == 'POST':
        room_name = request.POST.get('room_name', '새 단톡방')
        user_ids = request.POST.getlist('user_ids') # 선택된 유저들
        if user_ids:
            new_room = ChatRoom.objects.create(name=room_name, is_group_chat=True)
            new_room.participants.add(request.user) # 나 추가
            for uid in user_ids:
                new_room.participants.add(uid) # 선택한 사람들 추가
            return redirect('quiz:chat_room_detail', room_id=new_room.id)

    # ========================================================
    # 1. 내 채팅방 목록 가져오기 (★ 고정된 방 먼저, 그다음 최신순 정렬)
    # ========================================================
    my_rooms = ChatRoom.objects.filter(
        participants=request.user
    ).exclude(
        hidden_by=request.user # 기존에 있던 숨김 처리 유지
    ).annotate(
        # 내가 고정한 방이면 1점, 아니면 0점을 줘서 정렬 기준으로 씁니다.
        is_pinned=Case(
            When(pinned_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        )
    ).order_by('-is_pinned', '-last_activity')

    # ========================================================
    # 2. 화면에 뿌려줄 데이터 가공 (기존 로직 100% 유지 + 핀 색깔 추가)
    # ========================================================
    room_data = []
    for room in my_rooms:
        if room.is_group_chat:
            display_name = f"[단톡방] {room.name}"
        else:
            other_user = room.participants.exclude(id=request.user.id).first()
            display_name = other_user.profile.name if other_user and hasattr(other_user, 'profile') else (other_user.username if other_user else "알 수 없음")
            
        last_msg = room.messages.order_by('-created_at').first()
        unread_count = room.messages.exclude(sender=request.user).exclude(read_by=request.user).count()
        
        last_time = last_msg.created_at if last_msg else room.last_activity

        room_data.append({
            'room': room,
            'display_name': display_name,
            'participant_count': room.participants.count(),
            'last_msg': last_msg,
            'unread_count': unread_count,
            'last_time': last_time, # ★ [추가 2] 이 줄이 있어야 정렬이 정상 작동합니다!
            'is_pinned': request.user in room.pinned_by.all() 
        })

    room_data.sort(key=lambda x: (
        x['is_pinned'],         # 1. 핀 꽂힌 방을 가장 위로!
        x['unread_count'] > 0,  # 2. 안 읽은 알림 뱃지가 있는 방을 그 다음으로!
        x['last_time']          # 3. 나머지는 최신 시간순으로!
    ), reverse=True)
    
    # ========================================================
    # 3. 연락처 목록 (수료/퇴소자 제외)
    # ========================================================
    from accounts.models import Profile, Company, Process
    
    profiles = Profile.objects.exclude(
        user=request.user
    ).exclude(
        status__in=['completed', 'dropout'] # ★ 핵심: 수료, 퇴소 상태 제외!
    ).filter(
        user__is_active=True,
        is_profile_complete=True, # 2차 정보 기입 완료자만
        name__isnull=False        # 이름 있는 사람만
    ).select_related('user', 'company', 'process')
    
    managers = profiles.filter(is_manager=True)
    trainees = profiles.filter(is_manager=False)
    
    # 단톡방 만들 때 쓸 전체 유저 목록
    all_users = profiles.order_by('name')
    
    return render(request, 'quiz/chat_home.html', {
        'room_data': room_data,
        'managers': managers,
        'trainees': trainees,
        'all_users': all_users,
        'companies': Company.objects.all(),
        'processes': Process.objects.all() 
    })
    
    # 2. 연락처 목록
    from accounts.models import Profile
    
    profiles = Profile.objects.exclude(
        user=request.user
    ).exclude(
        status__in=['completed', 'dropout'] # ★ 핵심: 수료, 퇴소 상태 제외!
    ).filter(
        user__is_active=True,
        is_profile_complete=True, # 2차 정보 기입 완료자만
        name__isnull=False        # 이름 있는 사람만
    ).select_related('user', 'company', 'process')
    
    managers = profiles.filter(is_manager=True)
    trainees = profiles.filter(is_manager=False)
    
    # 단톡방 만들 때 쓸 전체 유저 목록 (여기서도 수료/퇴소자는 빠집니다)
    all_users = profiles.order_by('name')

    from accounts.models import Company, Process # 위쪽에 임포트 안 되어 있다면 추가
    
    return render(request, 'quiz/chat_home.html', {
        'room_data': room_data,
        'managers': managers,
        'trainees': trainees,
        'all_users': all_users,
        'companies': Company.objects.all(), # ★ 추가됨
        'processes': Process.objects.all()  # ★ 추가됨
    })

@login_required
def chat_start_1on1(request, target_user_id):
    target_user = get_object_or_404(User, id=target_user_id)
    common_rooms = ChatRoom.objects.filter(is_group_chat=False, participants=request.user).filter(participants=target_user)
    
    if common_rooms.exists():
        room = common_rooms.first()
        # ★ 예전에 나갔던 방이라면 다시 숨김 해제!
        room.hidden_by.remove(request.user)
    else:
        room = ChatRoom.objects.create(is_group_chat=False)
        room.participants.add(request.user, target_user)
        
    return redirect('quiz:chat_room_detail', room_id=room.id)

# (참고) 진짜 1:1 채팅창 화면 뷰는 다음 스텝에서 추가할 예정입니다!
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden

@login_required
@xframe_options_exempt
def chat_room_detail(request, room_id):
    room = get_object_or_404(ChatRoom, id=room_id)
    if request.user not in room.participants.all() and not request.user.is_superuser:
        return HttpResponseForbidden("권한이 없습니다.")

    # (읽음 처리 로직 삭제됨 - WebSockets에서 실시간으로 처리하여 1이 뿅 사라지게 만듦!)

    messages = room.messages.exclude(deleted_by=request.user).select_related('sender__profile').order_by('created_at')[:50]
    total_participants = room.participants.count()

    # ★ 2. 각 메시지마다 '안 읽은 사람 수' 계산
    for msg in messages:
        msg.unread_count = total_participants - msg.read_by.count()

    other_user_name = "단톡방"
    if not room.is_group_chat:
        other_user = room.participants.exclude(id=request.user.id).first()
        if other_user:
            other_user_name = other_user.profile.name if hasattr(other_user, 'profile') else other_user.username

    return render(request, 'quiz/chat_room.html', {
        'room': room,
        'messages': messages,
        'other_user_name': other_user_name
    })

@login_required
def chat_leave_room(request, room_id):
    room = get_object_or_404(ChatRoom, id=room_id)
    
    if room.is_group_chat:
        room.participants.remove(request.user)
        if room.participants.count() == 0:
            room.delete()
    else:
        # ★ 1:1 방은 방을 숨기고, 기존 메시지를 내 화면에서 지움
        room.hidden_by.add(request.user)
        for msg in room.messages.all():
            msg.deleted_by.add(request.user)
            
    return redirect('quiz:chat_home')

@login_required
def chat_unread_count(request):
    """대시보드에서 안 읽은 총 메시지 개수 가져오기"""
    # 내가 속한 방들의 메시지 중, 내가 읽지 않은(read_by에 내가 없는) 메시지 개수
    count = ChatMessage.objects.filter(room__participants=request.user).exclude(sender=request.user).exclude(read_by=request.user).count()
    return JsonResponse({'unread_chat_count': count})

def broadcast_realtime_notification(user_id):
    """
    특정 유저(user_id)의 브라우저에게 "새 종모양 알림이 왔으니 숫자를 당장 올려라!" 
    라고 실시간 무전(WebSocket)을 쏘는 헬퍼 함수입니다.
    """
    channel_layer = get_channel_layer()
    if channel_layer:
        # NotificationConsumer(알림 교환원)의 'chat_notification'과 동일한 신호를 보내어
        # 브라우저가 loadNotifications(true)를 즉시 실행하게 만듭니다.
        async_to_sync(channel_layer.group_send)(
            f'user_{user_id}',
            {
                'type': 'chat_notification', 
                'message': '시스템 새 알림 도착',
                'sender_name': '시스템',
                'room_id': 'system' # 채팅방 알림과 구분하기 위한 가짜 ID
            }
        )

@login_required
def chat_read_status(request, msg_id):
    """특정 메시지를 읽은 사람 / 안 읽은 사람 명단 반환"""
    msg = get_object_or_404(ChatMessage, id=msg_id)
    room = msg.room
    
    read_users = msg.read_by.all()
    # 방 참가자 중 읽은 사람을 제외하면 안 읽은 사람!
    unread_users = room.participants.exclude(id__in=read_users.values_list('id', flat=True))
    
    # 예쁘게 이름 리스트로 포장
    read_list = [{'name': u.profile.name if hasattr(u, 'profile') else u.username} for u in read_users]
    unread_list = [{'name': u.profile.name if hasattr(u, 'profile') else u.username} for u in unread_users]
    
    return JsonResponse({'read': read_list, 'unread': unread_list})

@login_required
@require_POST
def chat_pin_message(request, room_id):
    """특정 메시지를 방 상단에 공지로 고정(Pin)"""
    room = get_object_or_404(ChatRoom, id=room_id)
    msg_id = request.POST.get('msg_id')
    
    if msg_id:
        msg = get_object_or_404(ChatMessage, id=msg_id, room=room)
        room.pinned_message = msg
    else:
        room.pinned_message = None # 공지 내리기
        
    room.save()
    
    # 웹소켓으로 방에 있는 사람들에게 "공지 바뀌었음!" 새로고침 신호 쏘기
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    channel_layer = get_channel_layer()
    if channel_layer:
        async_to_sync(channel_layer.group_send)(
            f'chat_{room.id}',
            {'type': 'system_message', 'action': 'reload_pin'}
        )
        
    return JsonResponse({'status': 'success'})

@login_required
def chat_search_messages(request, room_id):
    """채팅방 내부 키워드 검색 API"""
    room = get_object_or_404(ChatRoom, id=room_id)
    q = request.GET.get('q', '')
    
    if not q:
        return JsonResponse({'results': []})
    
    # 내가 숨긴/지운 메시지는 제외하고, 내용이나 파일명에 검색어가 포함된 것 찾기 (최신 30개)
    msgs = room.messages.exclude(deleted_by=request.user).filter(
        Q(content__icontains=q) | Q(file_name__icontains=q)
    ).select_related('sender__profile').order_by('-created_at')[:30]
    
    data = []
    for m in msgs:
        sender_name = m.sender.profile.name if hasattr(m.sender, 'profile') else m.sender.username
        text = m.content if m.content else f"📁 [{m.file_name}]"
        data.append({
            'id': m.id,
            'sender': sender_name,
            'text': text,
            'date': m.created_at.strftime('%y.%m.%d %H:%M')
        })
        
    return JsonResponse({'results': data})

@login_required
@require_POST
def chat_invite_users(request, room_id):
    """카톡 스타일 초대 로직"""
    current_room = get_object_or_404(ChatRoom, id=room_id)
    # 초대할 유저 ID 리스트 (JSON으로 받음)
    import json
    data = json.loads(request.body)
    invited_user_ids = data.get('user_ids', [])
    
    if not invited_user_ids:
        return JsonResponse({'status': 'error', 'message': '초대할 대상을 선택해주세요.'})

    # 1. 기존 방이 1:1 방일 경우 -> '새로운 단톡방' 생성 (카톡과 동일)
    if not current_room.is_group_chat:
        # 새로운 단톡방 생성
        new_room = ChatRoom.objects.create(
            is_group_chat=True,
            name=f"{request.user.profile.name}님의 단톡방" # 기본 방 이름
        )
        # 기존 1:1 방 멤버들 + 새로 초대된 유저들 합치기
        existing_members = list(current_room.participants.all())
        new_room.participants.add(*existing_members)
        new_room.participants.add(*invited_user_ids)
        
        return JsonResponse({
            'status': 'success', 
            'action': 'new_room', 
            'new_room_id': new_room.id
        })

    # 2. 기존 방이 이미 단톡방일 경우 -> '현재 방'에 인원만 추가
    else:
        current_room.participants.add(*invited_user_ids)
        
        # (참고) 이전 대화 숨김: 
        # 우리 시스템은 메시지마다 '읽은 사람(read_by)'을 기록하므로, 
        # 새로 들어온 사람은 과거 메시지의 read_by에 포함되지 않아 자연스럽게 '안 읽음' 상태가 되거나, 
        # 로직상 조인 시점 이후 메시지만 뿌려주는 식으로 처리됩니다.
        
        # 시스템 메시지 발송 (선택사항)
        for uid in invited_user_ids:
            u = User.objects.get(id=uid)
            
            # ★ 이름이 없으면 아이디라도 뜨도록 방어 로직 추가!
            name = getattr(u.profile, 'name', None) if hasattr(u, 'profile') else None
            display_name = name if name else u.username
            
            ChatMessage.objects.create(
                room=current_room,
                sender=request.user, # 초대한 사람 이름으로 시스템 메시지 처리
                content=f"📢 {display_name}님이 초대되었습니다."
            )
            
        return JsonResponse({'status': 'success', 'action': 'added'})

@login_required
def chat_invite_targets(request, room_id):
    """현재 채팅방 멤버를 제외한 초대 가능 인원 목록을 반환"""
    room = get_object_or_404(ChatRoom, id=room_id)
    
    # 1. 현재 방에 이미 들어와 있는 인원들의 ID 목록
    existing_participant_ids = room.participants.values_list('id', flat=True)
    
    # 2. 초대 가능 대상 필터링:
    #    - 나 제외 / 이미 방에 있는 사람 제외 / 활성 계정 / 수료·퇴소자 제외
    targets = Profile.objects.exclude(
        user__id__in=existing_participant_ids
    ).filter(
        user__is_active=True,
        is_profile_complete=True
    ).exclude(
        status__in=['completed', 'dropout']
    ).select_related('process')

    # 3. JSON 데이터로 포장
    user_list = []
    for t in targets:
        # ★ 이름이 비어있으면(null) 아이디를 띄우도록 방어막 추가
        display_name = t.name if t.name else t.user.username 
        user_list.append({
            'id': t.user.id,
            'name': display_name,
            'process': t.process.name if t.process else '소속 없음'
        })
        
    return JsonResponse({'users': user_list})

@login_required
def chat_load_more_messages(request, room_id):
    """[무한 스크롤] 과거 메시지 50개씩 추가 로딩"""
    room = get_object_or_404(ChatRoom, id=room_id)
    first_msg_id = request.GET.get('first_msg_id')
    
    if not first_msg_id:
        return JsonResponse({'messages': []})
        
    # 현재 화면에 있는 가장 오래된 메시지보다 '더 과거'의 메시지 50개 호출
    msgs = room.messages.exclude(deleted_by=request.user).filter(
        id__lt=first_msg_id
    ).select_related('sender__profile', 'parent__sender__profile').order_by('-created_at')[:50]
    
    data = []
    # 시간순으로 정렬하기 위해 다시 뒤집음(reversed)
    for m in reversed(msgs):
        sender_name = m.sender.profile.name if hasattr(m.sender, 'profile') else m.sender.username
        is_system = '📢' in m.content if m.content else False
        
        # 파일 URL 안전하게 가져오기
        f_url = getattr(m, 'file_url', None)
        if not f_url and hasattr(m, 'file') and m.file:
            f_url = m.file.url
            
        data.append({
            'msg_id': m.id,
            'username': sender_name,
            'is_me': m.sender == request.user,
            'is_system': is_system,
            'message': m.content,
            'file_url': f_url,
            'file_name': getattr(m, 'file_name', None),
            'time_str': m.created_at.strftime('%H:%M'),
            'unread_count': room.participants.count() - m.read_by.count(),
            'parent_text': m.parent.content if m.parent else None,
            'parent_sender': m.parent.sender.profile.name if m.parent and hasattr(m.parent.sender, 'profile') else None,
        })
        
    return JsonResponse({'messages': data})

@login_required
def chat_toggle_pin(request, room_id):
    """채팅방 상단 고정/해제 토글 API"""
    room = get_object_or_404(ChatRoom, id=room_id)
    
    # 이미 고정되어 있으면 빼고, 없으면 넣습니다.
    if request.user in room.pinned_by.all():
        room.pinned_by.remove(request.user)
        pinned = False
    else:
        room.pinned_by.add(request.user)
        pinned = True
        
    return JsonResponse({'status': 'success', 'is_pinned': pinned})

@login_required
@require_POST
def quick_update_practice_score(request):
    """대시보드 빠른 실습 점수 개별 입력 및 최종 점수 재계산"""
    user = request.user
    user_profile = getattr(user, 'profile', None)
    
    if not (user.is_staff or (user_profile and (user_profile.is_manager or user_profile.is_pl))):
        return JsonResponse({'status': 'error', 'message': '권한이 없습니다.'}, status=403)

    from accounts.models import Profile, FinalAssessment
    from quiz.models import StudentLog
    import json
    from django.db import transaction

    try:
        data = json.loads(request.body)
        scores_data = data.get('scores', []) # [{'id': '1', 'score': '95'}, {'id': '2', 'score': '80'}] 형태

        if not scores_data:
            return JsonResponse({'status': 'error', 'message': '선택된 교육생이나 입력된 점수가 없습니다.'})

        success_names = []

        with transaction.atomic():
            for item in scores_data:
                pid = item.get('id')
                score_str = item.get('score')

                if not pid or not score_str:
                    continue

                score = float(score_str)
                profile = Profile.objects.get(id=pid)

                # [보안] 최고관리자가 아니면, '내 공정' 학생인지 확인
                if not user.is_superuser and user_profile.process != profile.process:
                    continue 

                # 최종 평가서 가져오기 또는 생성
                assessment, created = FinalAssessment.objects.get_or_create(profile=profile)

                # 1. 개별 실습 점수 업데이트
                assessment.practice_score = score

                # 2. 태도 점수 감점 로직 계산
                warnings = StudentLog.objects.filter(profile=profile, log_type='warning').count()
                letters = StudentLog.objects.filter(profile=profile, log_type='warning_letter').count()
                
                warning_penalty = (warnings - 1) * 10 if warnings > 1 else 0
                letter_penalty = letters * 10
                total_penalty = min(warning_penalty + letter_penalty, 40)

                raw_attitude = assessment.attitude_score or 0.0
                real_attitude = max(0, raw_attitude - total_penalty)
                exam_avg = assessment.exam_avg_score or 0.0

                # 3. 최종 환산 점수 재계산 및 저장
                final_score = round((exam_avg * 0.85) + (score * 0.05) + (real_attitude * 0.10), 1)
                assessment.final_score = final_score
                assessment.save()

                success_names.append(profile.name)

        if not success_names:
            return JsonResponse({'status': 'error', 'message': '점수를 부여할 수 있는 학생이 없습니다.'})

        msg = f"총 {len(success_names)}명 실습 점수 개별 반영 및 최종 성적 갱신 완료!"
        return JsonResponse({'status': 'success', 'message': msg})

    except Profile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '일부 교육생 데이터를 찾을 수 없습니다.'})
    except ValueError:
        return JsonResponse({'status': 'error', 'message': '점수는 숫자로 입력해주세요.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
@login_required
@require_GET
def get_manual_exam_targets(request):
    """지정된 차수에 맞춰 응시 대상자(과락자 등)만 필터링해서 내려줌 (면담 잠금 여부 포함)"""
    cohort_id = request.GET.get('cohort_id')
    quiz_id = request.GET.get('quiz_id')
    attempt = int(request.GET.get('attempt', 1))

    base_profiles = Profile.objects.filter(
        cohort_id=cohort_id, is_manager=False, is_pl=False, status__in=['attending', 'caution', 'counseling']
    ).select_related('user', 'company', 'process')

    data = []
    
    if attempt == 1:
        targets = base_profiles
        for p in targets:
            data.append({
                'id': p.id,
                'user_id': p.user.id,
                'name': p.name,
                'company': p.company.name if p.company else '-',
                'is_locked': False # 1차는 무조건 안 잠김
            })
    else:
        # 2, 3차는 직전 차수 불합격자만
        prev_attempt = attempt - 1
        failed_results = TestResult.objects.filter(
            quiz_id=quiz_id, 
            attempt_number=prev_attempt, 
            is_pass=False
        ).select_related('user__profile')
        
        for res in failed_results:
            p = res.user.profile
            # ★ 핵심: 이전 불합격으로 인해 생성된 로그(잠금)가 해결(면담완료) 안 되었는지 확인
            is_locked = StudentLog.objects.filter(
                profile=p, related_quiz_id=quiz_id, log_type='exam_fail', is_resolved=False
            ).exists()
            
            data.append({
                'id': p.id,
                'user_id': p.user.id,
                'name': p.name,
                'company': p.company.name if p.company else '-',
                'is_locked': is_locked # 잠겼으면 True 전달
            })

    return JsonResponse({'status': 'success', 'data': data})


@login_required
@require_POST
def submit_manual_exam_scores(request):
    """점수만 쏙 저장하고, 불합격 시 기존 CBT처럼 자동 알림+잠금 처리"""
    data = json.loads(request.body)
    quiz_id = data.get('quiz_id')
    attempt = int(request.GET.get('attempt', 1))
    results = data.get('results', []) 

    quiz = Quiz.objects.get(id=quiz_id)
    pass_score = quiz.pass_score

    with transaction.atomic():
        for res in results:
            profile = Profile.objects.get(id=res['profile_id'])
            score = float(res['score'])
            is_pass = score >= pass_score

            TestResult.objects.update_or_create(
                user_id=res['user_id'],
                quiz=quiz,
                attempt_number=attempt,
                defaults={'score': score, 'is_pass': is_pass}
            )

            # 불합격 시 기존 시스템과 완벽하게 똑같이 동작!
            if not is_pass:
                if attempt < 3:
                    # 1. 잠금 로그 생성
                    reason_msg = f"[{quiz.title}] {attempt}차 수기 평가 불합격 - 재응시 잠금"
                    if attempt == 2: reason_msg += " (PL 면담 필요)"

                    StudentLog.objects.create(
                        profile=profile,
                        recorder=request.user,
                        log_type='exam_fail',
                        reason=reason_msg,
                        related_quiz=quiz,
                        stage=attempt,
                        is_resolved=False
                    )
                    
                    # 2. 알림 센터로 종소리 빵! (기존 로직)
                    target_url = f"/quiz/manager/trainees/{profile.id}/logs/"
                    receivers = set(User.objects.filter(is_superuser=True))
                    if profile.process:
                        managers = User.objects.filter(is_staff=True, profile__is_manager=True, profile__process=profile.process)
                        receivers.update(managers)
                        
                    for recv in receivers:
                        Notification.objects.create(
                            recipient=recv,
                            sender=request.user,
                            message=f"🚨 {profile.name}님 '{quiz.title}' 시험 불합격! 면담(잠금 해제) 기록 작성이 필요합니다.",
                            notification_type='counseling',
                            related_url=target_url
                        )

                elif attempt == 3:
                    # 3차 최종 탈락 퇴소 처리
                    profile.status = 'dropout'
                    profile.save()
                    StudentLog.objects.create(
                        profile=profile, recorder=request.user, log_type='exam_fail',
                        reason=f"{quiz.title} 3차 수기 시험 과락으로 인한 퇴소 처리",
                        related_quiz=quiz, stage=3, is_resolved=False
                    )

    return JsonResponse({'status': 'success', 'message': '채점 점수가 저장되었습니다. 불합격자는 자동 잠금(면담 필요) 처리되었습니다.'})

@login_required
@require_POST
def submit_manual_exam_scores(request):
    """입력된 점수와 면담 내용을 일괄 저장하고 3차 탈락 시 퇴소 처리"""
    data = json.loads(request.body)
    quiz_id = data.get('quiz_id')
    attempt = int(data.get('attempt', 1))
    results = data.get('results', []) # [{profile_id, user_id, score, note}, ...]

    quiz = Quiz.objects.get(id=quiz_id)
    pass_score = quiz.pass_score

    with transaction.atomic():
        for res in results:
            profile = Profile.objects.get(id=res['profile_id'])
            score = float(res['score'])
            note = res.get('note', '').strip()
            is_pass = score >= pass_score

            # 1. 시험 결과(TestResult) 저장 (CBT와 완벽 호환)
            TestResult.objects.update_or_create(
                user_id=res['user_id'],
                quiz=quiz,
                attempt_number=attempt,
                defaults={
                    'score': score,
                    'is_pass': is_pass,
                }
            )

            # 2. 80점 미만이고 면담을 썼다면 면담 로그(StudentLog) 저장
            if not is_pass and note:
                log_type = 'counseling' if attempt == 1 else ('warning' if attempt == 2 else 'exam_fail')
                StudentLog.objects.create(
                    profile=profile,
                    recorder=request.user,
                    log_type=log_type,
                    related_quiz=quiz,
                    stage=attempt,
                    reason=f"{quiz.title} {attempt}차 수기 평가 과락",
                    action_taken=note
                )

            # 3. [초강력 룰] 3차 마저 떨어졌다면 강제 퇴소 처리!
            if attempt == 3 and not is_pass:
                profile.status = 'dropout'
                profile.save()
                
                # 추가 로그 남기기
                StudentLog.objects.create(
                    profile=profile,
                    recorder=request.user,
                    log_type='exam_fail',
                    reason=f"{quiz.title} 3차(최종) 시험 과락으로 인한 자동 퇴소 처리",
                )

    return JsonResponse({'status': 'success', 'message': '채점 결과 및 조치가 완벽하게 저장되었습니다.'})

@login_required
def global_analytics_full(request, analysis_type):
    """
    [관리자 전용] 문제 은행 전체 누적 정답률 & 태그 통계 (전체 스크롤 보기)
    - analysis_type: 'tag' 또는 'question'
    """
    if not request.user.is_staff:
        messages.error(request, "접근 권한이 없습니다.")
        return redirect('quiz:index')

    if analysis_type == 'tag':
        # 태그 전체 데이터 (정답률 낮은 순)
        stats = Tag.objects.annotate(
            total_tries=Count('question__useranswer'),
            correct_tries=Count('question__useranswer', filter=Q(question__useranswer__is_correct=True))
        ).annotate(
            accuracy=Case(
                When(total_tries=0, then=Value(0.0)),
                default=Cast(F('correct_tries'), FloatField()) / Cast(F('total_tries'), FloatField()) * 100.0,
                output_field=FloatField()
            )
        ).filter(total_tries__gt=0).order_by('accuracy')
        title = "🏷️ 전체 취약 개념 (태그) 랭킹"

    else:
        # 문제 전체 데이터 (정답률 낮은 순)
        # ★ 여기도 Question.objects 바로 뒤에 .prefetch_related('quizzes') 를 추가했습니다!
        stats = Question.objects.prefetch_related('quizzes').annotate(
            total_tries=Count('useranswer'),
            correct_tries=Count('useranswer', filter=Q(useranswer__is_correct=True))
        ).annotate(
            accuracy=Case(
                When(total_tries=0, then=Value(0.0)),
                default=Cast(F('correct_tries'), FloatField()) / Cast(F('total_tries'), FloatField()) * 100.0,
                output_field=FloatField()
            )
        ).filter(total_tries__gt=0).order_by('accuracy')
        title = "🚨 전체 마의 문항 (문제) 랭킹"

    context = {
        'stats': stats,
        'type': analysis_type,
        'title': title
    }
    
    return render(request, 'quiz/manager/global_analytics_full.html', context)

@login_required
def print_warning_letter(request, log_id):
    from .models import StudentLog
    log = get_object_or_404(StudentLog, id=log_id, log_type='warning_letter')
    
    if not request.user.is_staff and log.profile.user != request.user:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("본인의 경고장만 열람할 수 있습니다.")

    # ★ [추가] 이 학생이 받은 경고장 날짜 이전까지의 모든 위반/경고 내역을 가져옴
    warning_history = StudentLog.objects.filter(
        profile=log.profile,
        log_type__in=['warning', 'warning_letter', 'exam_fail'],
        created_at__lte=log.created_at
    ).order_by('created_at')
        
    return render(request, 'quiz/manager/warning_letter_print.html', {
        'log': log,
        'warning_history': warning_history  # ★ 템플릿으로 내역 전달
    })