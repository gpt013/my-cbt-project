# quiz/views_facility.py

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_datetime
from datetime import timedelta
from django.utils import timezone
from django.db.models import Count
from .models import Room, Reservation, Notification # 모델 임포트 필수!
from quiz.views import broadcast_realtime_notification
from django.contrib.auth.models import User
# accounts 앱의 모델 가져오기 (없으면 에러나니 꼭 확인)
try:
    from accounts.models import Profile, Process, Cohort,Company
except ImportError:
    # 혹시 모를 에러 방지용 더미 클래스
    Profile = None
    Process = None
    Cohort = None

# [헬퍼] 알림 발송
def send_notification(user, message, notification_type='facility', related_url=None):
    if user:
        Notification.objects.create(
            recipient=user, 
            message=message, 
            notification_type=notification_type,
            related_url=related_url # ★ 드디어 주소가 DB에 저장됩니다!
        )
        broadcast_realtime_notification(user.id)

# [헬퍼] 안전한 날짜 변환 (Timezone 에러 방지)
def safe_parse_datetime(date_str):
    if not date_str: return None
    dt = parse_datetime(date_str)
    if dt and timezone.is_naive(dt):
        return timezone.make_aware(dt)
    return dt

# [1] 대시보드
@login_required
def facility_dashboard(request):
    if not request.user.is_staff:
        return redirect('quiz:index')
    
    rooms = Room.objects.filter(is_active=True)
    companies = Company.objects.all()  # ★ 이 줄을 반드시 추가해야 합니다!
    today = timezone.localtime(timezone.now())
    
    # 1. 이번 달 통계 데이터
    start_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_res = Reservation.objects.filter(start_time__gte=start_month)
    
    total_confirmed = month_res.filter(status='confirmed').count()
    total_pending = month_res.filter(status='pending').count()
    total_rejected = month_res.filter(status='rejected').count()
    
    company_stats = month_res.filter(status='confirmed').values('company_name').annotate(count=Count('id')).order_by('-count')
    process_stats = month_res.filter(status='confirmed').values('process_name').annotate(count=Count('id')).order_by('-count')
    
    # 2. ★ [수정됨] "오늘의 예약" 리스트 (강의실별 그룹화)
    today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # (1) 데이터 가져오기: 반드시 '강의실 이름' 순으로 정렬해야 그룹화가 가능합니다.
    raw_reservations = Reservation.objects.filter(
        start_time__lte=today_end,
        end_time__gte=today_start
    ).exclude(status='rejected').select_related('room', 'user__profile').order_by('room__name', 'start_time')

    # (2) 파이썬 로직으로 그룹화: [{'room': RoomObj, 'events': [Res1, Res2]}, ...] 형태
    today_reservations = []
    if raw_reservations:
        current_room = None
        current_group = None
        
        for res in raw_reservations:
            # 방이 바뀌면 새로운 그룹 시작
            if res.room != current_room:
                current_room = res.room
                current_group = {
                    'room': current_room, 
                    'events': []
                }
                today_reservations.append(current_group)
            
            # 현재 그룹에 예약 추가
            current_group['events'].append(res)

    # 3. 내 알림
    my_notifications = Notification.objects.filter(
        recipient=request.user, 
        is_read=False, 
        notification_type='facility'  # <--- 이 부분이 추가되었습니다.
    ).order_by('-created_at')

    my_reservations = Reservation.objects.filter(user=request.user).order_by('-start_time')[:10]

    is_manager = request.user.is_superuser or request.user.groups.filter(name='FacilityManager').exists()

    context = {
        'rooms': rooms,
        'company_stats': company_stats,
        'process_stats': process_stats,
        'today_reservations': today_reservations, 
        'is_manager': is_manager,
        'notifications': my_notifications,
        'companies': companies,
        'my_reservations': my_reservations,
        'total_confirmed': total_confirmed,
        'total_pending': total_pending,
        'total_rejected': total_rejected,
    }
    return render(request, 'quiz/manager/facility_dashboard.html', context)

@login_required
def facility_events(request):
    action = request.GET.get('action')
    if action == 'monthly_stats':
        date_str = request.GET.get('date')
        try:
            view_date = parse_datetime(date_str + "T00:00:00")
            if timezone.is_naive(view_date): view_date = timezone.make_aware(view_date)
        except:
            view_date = timezone.localtime(timezone.now())

        start_month = view_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        import calendar
        last_day = calendar.monthrange(start_month.year, start_month.month)[1]
        end_month = start_month.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

        # ★ 해당 월에 '걸쳐있는' 모든 내 예약 가져오기
        my_res_qs = Reservation.objects.filter(
            user=request.user, start_time__lte=end_month, end_time__gte=start_month
        ).order_by('start_time')

        def get_my_list(status):
            return [{
                'title': r.title, 'room_name': r.room.name, 
                'company': r.company_name or '소속 미상',
                # ★ [수정] 다중일이면 시작~종료 날짜 모두 표시, 당일이면 시간만 표시
                'start': f"{timezone.localtime(r.start_time).strftime('%m/%d %H:%M')} ~ {timezone.localtime(r.end_time).strftime('%m/%d %H:%M')}" if timezone.localtime(r.start_time).date() != timezone.localtime(r.end_time).date() else f"{timezone.localtime(r.start_time).strftime('%m/%d %H:%M')} ~ {timezone.localtime(r.end_time).strftime('%H:%M')}",
                'user_name': r.user.profile.name if hasattr(r.user, 'profile') else r.user.username
            } for r in my_res_qs.filter(status=status).order_by('start_time')]

        my_stats = {
            'confirmed': get_my_list('confirmed'),
            'pending': get_my_list('pending'),
            'rejected': get_my_list('rejected'),
        }
    

        # 2. 내 예약 리스트 (다중일 표기 로직 추가)
        my_res_data = []
        for r in my_res_qs[:10]: # 목록은 최근 10개만
            s_dt = timezone.localtime(r.start_time)
            e_dt = timezone.localtime(r.end_time)
            
            # 당일 예약이면 시간만, 다중일이면 날짜~날짜 표기!
            if s_dt.date() == e_dt.date():
                time_str = f"{s_dt.strftime('%m/%d')} {s_dt.strftime('%H:%M')} ~ {e_dt.strftime('%H:%M')}"
            else:
                time_str = f"{s_dt.strftime('%m/%d %H:%M')} ~ {e_dt.strftime('%m/%d %H:%M')}"

            my_res_data.append({
                'title': r.title, 'status': r.status, 'room_name': r.room.name,
                'time_str': time_str
            })

        # 3. 관리자 통계 (해당 월 기준)
        is_manager = request.user.is_superuser or request.user.groups.filter(name='FacilityManager').exists()
        stats_data = {}
        if is_manager:
            all_res = Reservation.objects.filter(start_time__lte=end_month, end_time__gte=start_month)
            
            def get_list(status):
                return [{
                    'title': r.title, 'room_name': r.room.name, 
                    'company': r.company_name or '소속 미상',
                    # ★ [수정] 다중일이면 시작~종료 날짜 모두 표시, 당일이면 시간만 표시
                    'start': f"{timezone.localtime(r.start_time).strftime('%m/%d %H:%M')} ~ {timezone.localtime(r.end_time).strftime('%m/%d %H:%M')}" if timezone.localtime(r.start_time).date() != timezone.localtime(r.end_time).date() else f"{timezone.localtime(r.start_time).strftime('%m/%d %H:%M')} ~ {timezone.localtime(r.end_time).strftime('%H:%M')}",
                    'user_name': r.user.profile.name if hasattr(r.user, 'profile') else r.user.username
                } for r in all_res.filter(status=status).order_by('start_time')]

            stats_data = {
                'confirmed': get_list('confirmed'),
                'pending': get_list('pending'),
                'rejected': get_list('rejected'),
            }
            companies = all_res.filter(status='confirmed').values('company_name').annotate(count=Count('id')).order_by('-count')
            stats_data['companies'] = [{'name': c['company_name'] or '소속 미상', 'count': c['count']} for c in companies]

        return JsonResponse({
            'is_manager': is_manager,
            'my_stats': my_stats,        # ★ 내 통계 추가
            'my_reservations': my_res_data,
            'stats': stats_data,
            'month_str': f"{start_month.year}년 {start_month.month}월"
        })
    
    # 1. 파라미터 가져오기
    start = safe_parse_datetime(request.GET.get('start'))
    end = safe_parse_datetime(request.GET.get('end'))
    room_id = request.GET.get('room_id')

    events = []
    if start and end:
        # 2. 날짜 범위 필터링 (월이 넘어가는 다중일 일정도 잡히도록 교집합 조건으로 변경!)
        query = Reservation.objects.filter(start_time__lt=end, end_time__gt=start)
        
        # 3. 특정 강의실 필터링 (All이 아닐 경우)
        if room_id and room_id != 'all':
             query = query.filter(room_id=room_id)
        
        # 쿼리 최적화 (매니저 정보 미리 가져오기)
        events = query.select_related('room', 'user__profile').prefetch_related('room__managers')
    
    data = []
    for e in events:
        # 색상 로직
        if e.status == 'pending': color = '#adb5bd'   # 대기 (회색)
        elif e.status == 'rejected': color = '#dc3545' # 반려 (빨강)
        else: color = e.room.color                    # 확정 (방 고유색)
        
        # 이름/소속 표시 로직
        try:
            profile = e.user.profile
            user_name = profile.name
            company = e.company_name or (profile.company.name if profile.company else "")
            process = e.process_name or (profile.process.name if profile.process else "")
            prefix = f"[{company}/{process}] {user_name}"
        except:
            user_name = e.user.username
            prefix = f"[{user_name}]"

        # ★★★ [권한 로직 강화] ★★★
        is_superuser = request.user.is_superuser
        is_designated_manager = request.user in e.room.managers.all()
        is_owner = (e.user == request.user)
        
        can_approve = (is_superuser or is_designated_manager)
        can_edit = (can_approve or is_owner)

        data.append({
            'id': e.id,
            'room_id': e.room.id,
            'title': f"{prefix} - {e.title}",
            'start': e.start_time.isoformat(),
            'end': e.end_time.isoformat(),
            'color': color,
            'editable': can_edit, 
            'extendedProps': {
                'status_code': e.status, 
                'status_label': e.get_status_display(),
                'company': e.company_name, 
                'process': e.process_name,
                'username': user_name,
                'attendees': e.attendees,   # ★ 이 줄 추가
                'full_desc': e.title,
                'is_admin': can_approve,
                'can_approve': can_approve,
                'can_manage': can_edit,
                'can_edit': can_edit
            }
        })
        
    # =========================================================
    # ★ [추가] PMTC 기수(교육기간) 전체 일정 표시 (캘린더 맨 위 띠)
    # =========================================================
    if Cohort and start and end:
        # 달력 현재 화면(시작~끝 날짜)과 겹치는 기수 찾기
        overlapping_cohorts = Cohort.objects.filter(
            start_date__lte=end.date(),
            end_date__gte=start.date()
        )
        
        for c in overlapping_cohorts:
            current_start = c.start_date
            week_num = 1
            
            while current_start <= c.end_date:
                # ★ [핵심] 무조건 7일이 아니라 '토요일'을 기준으로 자릅니다!
                # 파이썬 weekday(): 월=0, 화=1, 수=2, 목=3, 금=4, 토=5, 일=6
                # 현재 날짜에서 이번 주 토요일까지 남은 일수 계산
                days_to_saturday = (5 - current_start.weekday()) % 7
                current_end = current_start + timedelta(days=days_to_saturday)
                
                # 기수 종료일을 넘지 않도록 방어
                if current_end > c.end_date:
                    current_end = c.end_date
                
                # 현재 달력 화면에 보이는 날짜인지 확인
                if current_start <= end.date() and current_end >= start.date():
                    # FullCalendar 종일(allDay) 이벤트는 종료일에 하루를 더해줘야 예쁘게 꽉 찹니다.
                    fc_end_date = current_end + timedelta(days=1)
                    
                    data.append({
                        'id': f'cohort_{c.id}_w{week_num}',
                        'title': f'🏫[ PMTC {c.name} 교육기간 {week_num}W ] ',
                        'start': current_start.isoformat(),
                        'end': fc_end_date.isoformat(),
                        'color': '#ffecb5',      # 부드러운 노란색 배경
                        'textColor': '#664d03',  # 진한 갈색 텍스트
                        'allDay': True,          # 달력 맨 위에 '종일' 띠로 고정
                        'editable': False,       # 마우스로 드래그 금지
                        'extendedProps': {
                            'is_cohort': True
                        }
                    })
                
                # 다음 주차 시작일 (다음 주 일요일)
                current_start = current_end + timedelta(days=1)
                week_num += 1

    return JsonResponse(data, safe=False)

# [3] 예약 신청 (알림 로직 수정: 지정 관리자에게 알림 발송)
@login_required
def facility_reserve(request):
    if request.method == 'POST':
        try:
            # 1. 수정인지 생성인지 확인
            event_id = request.POST.get('id')
            
            # 2. 다중 선택된 강의실 ID 리스트 가져오기 (room_ids)
            room_ids = request.POST.getlist('room_ids') 
            if not room_ids:
                return JsonResponse({'status': 'error', 'message': '강의실을 선택해주세요.'})

            start_dt = safe_parse_datetime(request.POST.get('start_time'))
            end_dt = safe_parse_datetime(request.POST.get('end_time'))
            title = request.POST.get('title')
            attendees = request.POST.get('attendees', 0)
            company_name = request.POST.get('company_name', '')

            if not start_dt or not end_dt: 
                return JsonResponse({'status': 'error', 'message': '날짜 오류'})
            if start_dt >= end_dt: 
                return JsonResponse({'status': 'error', 'message': '종료 시간이 시작 시간보다 빠를 수 없습니다.'})

            # 3. 예약 처리 (루프 돌면서 선택한 방 개수만큼 예약 생성)
            # 수정 모드일 때는 선택된 첫 번째 방 하나만 처리하는 것이 안전합니다.
            if event_id:
                res = get_object_or_404(Reservation, pk=event_id)
                room = get_object_or_404(Room, pk=room_ids[0]) # 수정 시엔 첫 번째 선택된 방으로
                
                # 권한 체크
                is_approver = request.user.is_superuser or (request.user in room.managers.all())
                if not (is_approver or res.user == request.user):
                    return JsonResponse({'status': 'error', 'message': '수정 권한이 없습니다.'})
                
                res.room = room
                res.title = title
                res.start_time = start_dt
                res.end_time = end_dt
                res.attendees = attendees
                res.company_name = company_name
                if not is_approver: res.status = 'pending'
                
                # 중복 체크 후 저장
                if res.check_overlap():
                    return JsonResponse({'status': 'error', 'message': f'[{room.name}] 해당 시간에 이미 예약이 있습니다.'})
                res.save()
            
            else:
                # [신규 생성 모드] 선택한 모든 강의실에 대해 각각 예약 생성
                for r_id in room_ids:
                    room = get_object_or_404(Room, pk=r_id)
                    is_approver = request.user.is_superuser or (request.user in room.managers.all())
                    
                    new_res = Reservation(
                        room=room, 
                        user=request.user, 
                        title=title,
                        start_time=start_dt, 
                        end_time=end_dt,
                        attendees=attendees,
                        company_name=company_name,
                        status='confirmed' if is_approver else 'pending'
                    )
                    
                    # 중복 체크
                    if new_res.check_overlap():
                        return JsonResponse({'status': 'error', 'message': f'[{room.name}] 이미 예약된 시간입니다.'})
                    new_res.save()

                    # 관리자 알림 발송 (생략 가능)
                    if not is_approver:
                        msg = f"📢 [예약신청] {request.user.profile.name if hasattr(request.user, 'profile') else request.user.username}님이 {room.name} 예약을 신청했습니다."
                        
                        target_users = set() # 중복 수신 방지 바구니
                        
                        # 1. 해당 강의실의 우선 배정 공정(target_process) 매니저들 싹 다 담기
                        if room.target_process:
                            for p in Profile.objects.filter(process=room.target_process, is_manager=True):
                                target_users.add(p.user)
                                
                        # 2. 지정 관리자들 담기
                        for m in room.managers.all():
                            target_users.add(m)
                            
                        # 3. 바구니에 담긴 모두에게 알림 발송!
                        for user_to_notify in target_users:
                            send_notification(user_to_notify, msg)

            return JsonResponse({'status': 'success', 'message': '예약이 정상적으로 처리되었습니다.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'서버 오류: {str(e)}'})
            
    return JsonResponse({'status': 'error', 'message': '잘못된 요청입니다.'})

# [4] 예약 변경 (권한 체크)
@login_required
def facility_update(request):
    if request.method == 'POST':
        try:
            event = get_object_or_404(Reservation, pk=request.POST.get('id'))
            
            # 권한: 슈퍼유저 OR 지정 관리자 OR 본인
            is_admin = request.user.is_superuser or (request.user in event.room.managers.all())
            if not (is_admin or event.user == request.user): 
                return JsonResponse({'status': 'error', 'message': '권한 없음'})
            
            start_dt = safe_parse_datetime(request.POST.get('start'))
            end_dt = safe_parse_datetime(request.POST.get('end'))
            
            event.start_time, event.end_time = start_dt, end_dt
            if event.check_overlap(): 
                return JsonResponse({'status': 'error', 'message': '중복된 시간입니다.'})
            
            # 일반인은 수정 시 승인 대기로 전환
            if not is_admin:
                reason = request.POST.get('reason', '사유 미기재')
                event.status = 'pending'
                msg = '시간이 변경되었습니다. (관리자 재승인 필요)'
                
                # ★ 알림 발송: 지정 관리자 + 공정 매니저
                target_users = set()
                if event.room.target_process:
                    for p in Profile.objects.filter(process=event.room.target_process, is_manager=True):
                        target_users.add(p.user)
                for m in event.room.managers.all():
                    target_users.add(m)
                    
                for user_to_notify in target_users:
                    send_notification(user_to_notify, f"✏️ [수정 요청] {event.title} 시간 변경\n사유: {reason}")
            else:
                msg = '시간 변경 완료'

            event.save()
            return JsonResponse({'status': 'success', 'message': msg})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error'})

# [5] 예약 관리 (승인/반려) - 지정 관리자도 가능하게
@login_required
def facility_action(request, event_id):
    event = get_object_or_404(Reservation, pk=event_id)
    
    # ★ [핵심] 슈퍼유저 이거나 이 방의 지정 관리자여야 함
    is_admin = request.user.is_superuser or (request.user in event.room.managers.all())
    action = request.POST.get('action')

    if action in ['approve', 'reject'] and not is_admin: 
        return JsonResponse({'status': 'error', 'message': '관리 권한이 없습니다.'})
    
    if action == 'delete' and not (is_admin or event.user == request.user): 
        return JsonResponse({'status': 'error', 'message': '권한 없음'})

    if action == 'approve':
        if event.check_overlap(): return JsonResponse({'status': 'error', 'message': '승인 실패(이미 확정된 예약 존재)'})
        
        # ★ 중복된 대기열(pending) 찾아서 자동 반려 처리!
        reject_reason = request.POST.get('reject_reason', f"중복 중 선택한 일정({event.title})으로 인한 변경")
        overlapping_pending = Reservation.objects.filter(
            room=event.room, status='pending',
            start_time__lt=event.end_time, end_time__gt=event.start_time
        ).exclude(pk=event.pk)
        
        from django.urls import reverse # 상단에 없으면 에러나니 여기서 안전하게 임포트
        facility_url = reverse('quiz:facility_dashboard') # 대시보드 주소

        for p_res in overlapping_pending:
            p_res.status = 'rejected'
            p_res.save()
            # ★ related_url 추가
            send_notification(p_res.user, f"❌ 중복 예약이 반려되었습니다: {p_res.title}\n사유: {reject_reason}", notification_type='facility', related_url=facility_url)

        event.status = 'confirmed'
        event.save()
        # ★ related_url 추가
        send_notification(event.user, f"✅ 예약이 승인되었습니다: {event.title}", notification_type='facility', related_url=facility_url)
        
        # ★ [알림 발송] 확정 소식도 양쪽 모두 공유
        target_users = set()
        if event.room.target_process:
            for p in Profile.objects.filter(process=event.room.target_process, is_manager=True):
                target_users.add(p.user)
        for m in event.room.managers.all():
            target_users.add(m)
            
        for user_to_notify in target_users:
            if user_to_notify != request.user:
                # ★ related_url 추가
                send_notification(user_to_notify, f"🆗 [확정] {event.title} 예약이 승인처리 되었습니다.", notification_type='facility', related_url=facility_url)

    elif action == 'reject':
        reject_reason = request.POST.get('reject_reason', '사유 없음')
        event.status = 'rejected'
        event.save()
        from django.urls import reverse
        # ★ related_url 추가
        send_notification(event.user, f"❌ 예약이 반려되었습니다: {event.title}\n사유: {reject_reason}", notification_type='facility', related_url=reverse('quiz:facility_dashboard'))

    elif action == 'delete':
        event.delete()
        if event.user != request.user: 
            send_notification(event.user, f"🗑 예약이 취소되었습니다: {event.title}")
        
    return JsonResponse({'status': 'success', 'message': '처리 완료'})

# =========================================================
# ★ [추가] 알림 관련 기능 (삭제, 전체삭제, API)
# =========================================================

@login_required
def notification_read(request, noti_id):
    if int(noti_id) == 0: return JsonResponse({'status': 'success'}) # 시스템 가짜 알림은 무시
    noti = get_object_or_404(Notification, id=noti_id, recipient=request.user)
    noti.is_read = True
    noti.save()
    return JsonResponse({'status': 'success'})

@login_required
def notification_delete(request, noti_id):
    if int(noti_id) == 0: return JsonResponse({'status': 'success'}) # 시스템 가짜 알림은 무시
    noti = get_object_or_404(Notification, id=noti_id, recipient=request.user)
    noti.delete()
    return JsonResponse({'status': 'success'})

@login_required
def notification_clear_all(request):
    Notification.objects.filter(recipient=request.user).delete()
    return JsonResponse({'status': 'success'})

@login_required
def notification_api_list(request):
    """상단 벨 아이콘용 알림 목록 API (+ 5일 지난 알림 자동 청소 & 평가 대기 알림 생성)"""
    
    # 1. 5일이 지난 일반 알림은 자동 청소
    expiration_date = timezone.now() - timedelta(days=5)
    Notification.objects.filter(recipient=request.user, created_at__lt=expiration_date).delete()

    # =========================================================
    # ★ [핵심] 평가 대기 인원 '진짜 개별 알림' 자동 생성 로직
    # =========================================================
    if request.user.is_staff:
        today = timezone.now().date()
        
        # 평가를 받아야 하는(기수 종료되었으나 재직중인) 대상자 모두 찾기
        pending_profiles = Profile.objects.filter(
            cohort__end_date__lt=today,
            status__in=['attending', 'caution', 'counseling']
        ).exclude(user__is_superuser=True).exclude(is_manager=True).select_related('cohort', 'process')

        if not request.user.is_superuser:
            if hasattr(request.user, 'profile') and request.user.profile.process:
                pending_profiles = pending_profiles.filter(process=request.user.profile.process)
            else:
                # 공정이 할당되지 않은 일반 스태프는 알림을 받지 않음
                pending_profiles = pending_profiles.none()
        # =====================================================

        from django.urls import reverse

        for p in pending_profiles:
            # (이하 기존 코드 동일)
            cohort_name = p.cohort.name if p.cohort else "미지정"
            process_name = p.process.name if p.process else "미지정"
            days_passed = (today - p.cohort.end_date).days

            # 대상자별 최종 평가서 주소
            target_url = reverse('quiz:evaluate_trainee', args=[p.id])

            # 알림 메시지 세팅 (지연 일수에 따라 독촉 메시지로 변경)
            if days_passed <= 1:
                msg = f"[{cohort_name}/{process_name}] {p.name}님 기수가 종료되었습니다. 상세 페이지에서 최종 평가 및 수료 처리를 진행해주세요."
            else:
                msg = f"🚨 [D+{days_passed}일 지연] [{cohort_name}/{process_name}] {p.name}님 수료 처리가 안되었습니다! 즉시 작성 부탁드립니다."

            # 이 매니저에게 이 학생에 대한 알림이 이미 있는지 확인
            existing_noti = Notification.objects.filter(
                recipient=request.user,
                notification_type='pending_eval',
                related_url=target_url
            ).first()

            if not existing_noti:
                # 알림이 없으면 새로 생성 (DB에 진짜로 저장됨)
                Notification.objects.create(
                    recipient=request.user,
                    message=msg,
                    related_url=target_url,
                    notification_type='pending_eval'
                )
            else:
                # 이미 알림이 있는데 날짜가 어제 것이라면? -> 오늘 날짜로 갱신하고 다시 띄움! (독촉)
                if existing_noti.created_at.date() < today:
                    existing_noti.message = msg
                    existing_noti.is_read = False # 다시 안 읽음 상태로 (빨간 점 켜짐)
                    existing_noti.created_at = timezone.now() # 최상단으로 끌어올림
                    existing_noti.save()

    # =========================================================
    # ★ [추가] 시설/장비 예약 당일 아침 브리핑 (Daily Reminder)
    # =========================================================
    # 시설 관리자 권한이 있는 사람에게만 작동합니다.
    if request.user.is_superuser or request.user.groups.filter(name='FacilityManager').exists() or hasattr(request.user, 'profile'):
        
        today = timezone.localtime().date()
        today_start = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
        today_end = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.max.time()))

        # 1. 오늘 이 유저에게 이미 '당일 브리핑(facility_daily)' 알림을 보냈는지 확인 (중복 알림 방지)
        already_sent_today = Notification.objects.filter(
            recipient=request.user,
            notification_type='facility_daily',
            created_at__gte=today_start
        ).exists()

        if not already_sent_today:
            # 2. 오늘 날짜와 1초라도 겹치는 '확정(confirmed)'된 모든 예약 가져오기
            today_res = Reservation.objects.filter(
                status='confirmed',
                start_time__lte=today_end,
                end_time__gte=today_start
            ).select_related('room', 'user__profile')

            # 3. 방(Room)별로 예약 묶기
            from collections import defaultdict
            room_map = defaultdict(list)
            
            for res in today_res:
                # 내가 관리자인 방인지 확인 (슈퍼유저거나, 담당공정이거나, 지정관리자거나)
                is_my_room = False
                if request.user.is_superuser:
                    is_my_room = True
                elif res.room.target_process and hasattr(request.user, 'profile') and request.user.profile.process == res.room.target_process and request.user.profile.is_manager:
                    is_my_room = True
                elif request.user in res.room.managers.all():
                    is_my_room = True

                if is_my_room:
                    room_map[res.room.name].append(res)

            # 4. 방별로 묶인 예약을 리스트 형태의 텍스트로 만들기
            from django.urls import reverse
            facility_url = reverse('quiz:facility_dashboard')

            for room_name, res_list in room_map.items():
                # 시간 순 정렬
                res_list.sort(key=lambda x: x.start_time)
                
                # 헤더 메시지
                msg_lines = [f"🔔 [오늘의 일정] {room_name}에 금일 확정된 예약이 {len(res_list)}건 있습니다."]
                
                for r in res_list:
                    s_dt = timezone.localtime(r.start_time)
                    e_dt = timezone.localtime(r.end_time)
                    user_name = r.user.profile.name if hasattr(r.user, 'profile') else r.user.username
                    
                    # 다중일 예약 시간 표기 로직 (오늘 시작인지, 어제부터 계속되는지 등)
                    if s_dt.date() < today < e_dt.date():
                        time_str = "종일 (연속 예약)"
                    elif s_dt.date() == today < e_dt.date():
                        time_str = f"{s_dt.strftime('%H:%M')} ~ (내일로 이어짐)"
                    elif s_dt.date() < today == e_dt.date():
                        time_str = f"(이전부터) ~ {e_dt.strftime('%H:%M')}"
                    else:
                        time_str = f"{s_dt.strftime('%H:%M')} ~ {e_dt.strftime('%H:%M')}"

                    msg_lines.append(f" - {time_str} : {r.title} ({user_name})")

                final_msg = "\n".join(msg_lines)

                # 5. 최종 알림 DB에 저장
                Notification.objects.create(
                    recipient=request.user,
                    message=final_msg,
                    notification_type='facility_daily', # ★ 중복 방지용 고유 타입
                    related_url=facility_url
                )

    # =========================================================
    # 2. 안 읽은 알림 목록 화면에 전송
    # =========================================================
    notis = Notification.objects.filter(recipient=request.user, is_read=False).order_by('-created_at')
    count = notis.count()
    
    data = []
    for n in notis:
        # ★ iicon 오타를 icon으로 고치고, 기본 아이콘을 확실하게 지정합니다.
        icon = 'bi-info-circle text-primary' 
        
        if n.notification_type == 'facility': icon = 'bi-building text-success'
        elif n.notification_type == 'facility_daily': icon = 'bi-calendar-check-fill text-primary'
        elif n.notification_type == 'signup': icon = 'bi-person-plus-fill text-info'
        elif n.notification_type == 'exam': icon = 'bi-pencil-square text-warning'
        elif n.notification_type == 'pending_eval': icon = 'bi-exclamation-square-fill text-danger'
        # ★ 우리가 추가했던 멘션 알림 아이콘
        elif n.notification_type == 'chat_mention': icon = 'bi-chat-dots-fill text-primary' 
        # 혹시 모를 일반 알림 아이콘 방어막
        elif n.notification_type == 'general': icon = 'bi-bell-fill text-secondary'

        data.append({
            'id': n.id,
            'message': n.message,
            'link': n.related_url if n.related_url else '#',
            'icon': icon,
            'time': n.created_at.strftime('%m/%d %H:%M')
        })
        
    return JsonResponse({'count': count, 'notifications': data})


def read_notification(request, id):
    # 1. 데이터베이스에서 해당 id의 알림 객체를 가져옵니다.
    notification = get_object_or_404(Notification, id=id)
    
    # 2. 읽음 처리를 합니다. (모델의 필드명에 맞게 수정하세요)
    notification.is_read = True 
    notification.save()
    
    # 3. 프론트엔드의 .then(res => res.json())이 잘 작동하도록 JSON 객체를 반환합니다.
    return JsonResponse({"status": "success", "message": "알림 읽음 처리 완료"})
        

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from collections import defaultdict
import urllib.parse
from .models import Reservation
from datetime import datetime, date, timedelta

@login_required
def export_facility_schedule_excel(request):
    if not request.user.is_staff:
        return HttpResponse("권한이 없습니다. (관리자 전용)", status=403)

    # 1. 화면에서 넘겨받은 시작일과 종료일
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if not start_date_str or not end_date_str:
        return HttpResponse("시작일과 종료일을 선택해주세요.", status=400)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    if start_date > end_date:
        return HttpResponse("종료일이 시작일보다 빠를 수 없습니다.", status=400)

    # 2. 해당 기간과 '겹치는' 예약들 가져오기
    reservations = Reservation.objects.filter(
        start_time__date__lte=end_date,
        end_time__date__gte=start_date,
        status='confirmed'
    ).select_related('room')

    # 3. 혹시 PMTC가 예약(Reservation) 테이블이 아닌 기수(Cohort) 테이블에 있을 경우를 대비한 가드코드
    cohort_list = []
    try:
        from accounts.models import Cohort
        cohorts = Cohort.objects.all()
        for c in cohorts:
            if hasattr(c, 'start_date') and hasattr(c, 'end_date') and c.start_date and c.end_date:
                if c.start_date <= end_date and c.end_date >= start_date:
                    cohort_list.append(c)
    except Exception:
        pass

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 엑셀 디자인 세팅
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    weekend_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    pmtc_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
    pmtc_font = Font(color="9C0006", bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["일자", "요일", "교육장 주요 일정 (PMTC 등)", "기타 확정된 예약 내역"]

    # 시작일부터 종료일까지 하루씩 넘어가면서 엑셀에 쓰기
    current_date = start_date
    while current_date <= end_date:
        current_year = current_date.year
        current_month = current_date.month
        sheet_name = f"{current_year}년 {current_month}월"
        
        # 월이 바뀔 때마다 시트 새로 만들기
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 35 
            ws.column_dimensions['D'].width = 50 

        row_idx = ws.max_row + 1
        weekday = current_date.weekday()
        weekday_str = ["월", "화", "수", "목", "금", "토", "일"][weekday]

        # ★ 핵심 버그 수정: 예약의 시작일~종료일 사이에 '오늘(current_date)'이 껴있으면 몽땅 가져옴!
        day_res = [r for r in reservations if r.start_time.date() <= current_date <= r.end_time.date()]
        
        is_pmtc = False
        pmtc_text = ""
        res_texts = []

        # 4-1. 예약(Reservation) 목록에서 PMTC 찾기
        for r in day_res:
            if 'PMTC' in r.title.upper():
                is_pmtc = True
                pmtc_text = f"🚨 {r.title}\n(전 랩실 사용 불가)"
            else:
                room_name = r.room.name if r.room else "미지정"
                res_texts.append(f"[{room_name}] {r.title} ({r.start_time.strftime('%H:%M')}~{r.end_time.strftime('%H:%M')})")
        
        # 4-2. 기수(Cohort) 모델에서 PMTC 찾기 (예약에 안 적혀 있을 경우)
        for c in cohort_list:
            if c.start_date <= current_date <= c.end_date:
                is_pmtc = True
                name = getattr(c, 'name', '')
                pmtc_text = f"🚨 {name} PMTC 진행\n(전 랩실 사용 불가)"

        res_combined = "\n".join(res_texts) if res_texts else ""

        # 5. 엑셀에 한 줄 쓰기
        row_data = [
            current_date.strftime('%Y-%m-%d'),
            weekday_str,
            pmtc_text,
            res_combined
        ]

        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.border = thin_border
            if col_num in [1, 2, 3]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # 색칠하기 (PMTC가 무조건 최우선!)
            if is_pmtc:
                cell.fill = pmtc_fill
                cell.font = pmtc_font
            elif weekday >= 5: 
                cell.fill = weekend_fill

        current_date += timedelta(days=1)

    filename = f"배포용_운영일정_{start_date_str}_{end_date_str}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    wb.save(response)

    return response

import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.db import transaction
from django.db.models import Q
from .models import Quiz, TestResult, StudentLog
from accounts.models import Profile # (Profile 위치에 맞게 import)

@login_required
@require_GET
def get_manual_exam_targets(request):
    """응시 대상자 필터링 (수정 불가 락 & 면담 미완료 락 적용)"""
    cohort_id = request.GET.get('cohort_id')
    quiz_id = request.GET.get('quiz_id')
    attempt = int(request.GET.get('attempt', 1))

    base_profiles = Profile.objects.filter(
        cohort_id=cohort_id, is_manager=False, is_pl=False, status__in=['attending', 'caution', 'counseling']
    ).select_related('user', 'company', 'process')

    # ★★★ [핵심 추가] 매니저는 '본인 공정' 학생만 봅니다! (최고관리자는 전체 다 봄) ★★★
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.process:
            base_profiles = base_profiles.filter(process=request.user.profile.process)

    data = []
    
    for p in base_profiles:
        existing_result = TestResult.objects.filter(user=p.user, quiz_id=quiz_id, attempt_number=attempt).first()
        is_already_entered = existing_result is not None
        existing_score = existing_result.score if existing_result else None

        if attempt == 1:
            data.append({
                'id': p.id, 'user_id': p.user.id, 'name': p.name, 'company': p.company.name if p.company else '-',
                'is_already_entered': is_already_entered, 'existing_score': existing_score, 'is_locked': False
            })
        else:
            prev_attempt = attempt - 1
            prev_result = TestResult.objects.filter(user=p.user, quiz_id=quiz_id, attempt_number=prev_attempt).first()
            
            if prev_result and not prev_result.is_pass:
                unresolved_log = StudentLog.objects.filter(
                    profile=p, related_quiz_id=quiz_id, log_type='exam_fail', stage=prev_attempt, is_resolved=False
                ).exists()
                
                data.append({
                    'id': p.id, 'user_id': p.user.id, 'name': p.name, 'company': p.company.name if p.company else '-',
                    'is_already_entered': is_already_entered, 'existing_score': existing_score, 'is_locked': unresolved_log
                })

    return JsonResponse({'status': 'success', 'data': data})

@login_required
@require_POST
def submit_manual_exam_scores(request):
    """점수 저장 및 누락자 자동 색출/알림 발송 (Bulk 처리 최적화)"""
    data = json.loads(request.body)
    cohort_id = data.get('cohort_id')
    quiz_id = data.get('quiz_id')
    attempt = int(data.get('attempt', 1))
    results = data.get('results', []) 

    quiz = Quiz.objects.get(id=quiz_id)
    pass_score = quiz.pass_score

    # ★ 최적화: 한 번에 DB에 밀어넣을 바구니(List) 준비
    test_results_to_create = []
    student_logs_to_create = []
    notifications_to_create = []
    profiles_to_update = []
    users_to_update = []

    # N+1 쿼리 방지: 필요한 학생들 정보 한방에 미리 다 가져오기
    user_ids = [res['user_id'] for res in results if res.get('score')]
    existing_results = set(TestResult.objects.filter(
        user_id__in=user_ids, quiz=quiz, attempt_number=attempt
    ).values_list('user_id', flat=True))
    
    profile_ids = [res['profile_id'] for res in results if res.get('score')]
    profiles_dict = {p.id: p for p in Profile.objects.filter(id__in=profile_ids).select_related('user', 'process')}
    
    superusers = list(User.objects.filter(is_superuser=True))
    managers_by_process = {}

    with transaction.atomic():
        for res in results:
            score_str = res.get('score')
            if not score_str: continue 
            
            user_id = int(res['user_id'])
            profile_id = int(res['profile_id'])
            score = float(score_str)

            if user_id in existing_results: continue

            is_pass = score >= pass_score
            profile = profiles_dict.get(profile_id)

            # 1. 성적 바구니에 담기
            test_results_to_create.append(
                TestResult(user_id=user_id, quiz=quiz, attempt_number=attempt, score=score, is_pass=is_pass)
            )

            if not is_pass:
                if attempt < 3:
                    reason_msg = f"[{quiz.title}] {attempt}차 평가 불합격 - 재응시 잠금"
                    if attempt == 2: reason_msg += " (PL 면담 필요)"

                    # 2. 학생 로그 바구니에 담기
                    student_logs_to_create.append(
                        StudentLog(profile=profile, recorder=request.user, log_type='exam_fail',
                                   reason=reason_msg, related_quiz=quiz, stage=attempt, is_resolved=False)
                    )
                    
                    # 수신자(매니저) 리스트 캐싱
                    receivers = set(superusers)
                    if profile.process:
                        if profile.process.id not in managers_by_process:
                            managers_by_process[profile.process.id] = list(User.objects.filter(is_staff=True, profile__is_manager=True, profile__process=profile.process))
                        receivers.update(managers_by_process[profile.process.id])
                        
                    # 3. 알림 바구니에 담기
                    target_url = f"/quiz/manager/trainees/{profile.id}/logs/"
                    for recv in receivers:
                        notifications_to_create.append(
                            Notification(recipient=recv, sender=request.user, notification_type='counseling', related_url=target_url,
                                         message=f"🚨 {profile.name}님 '{quiz.title}' 불합격! 면담(잠금 해제) 기록이 필요합니다.")
                        )
                elif attempt == 3:
                    profile.status = 'dropout'
                    profiles_to_update.append(profile)

                    profile.user.is_active = False
                    users_to_update.append(profile.user)

                    student_logs_to_create.append(
                        StudentLog(profile=profile, recorder=request.user, log_type='exam_fail',
                                   reason=f"{quiz.title} 3차 수기 시험 과락으로 인한 자동 퇴소 및 계정 정지 처리",
                                   related_quiz=quiz, stage=3, is_resolved=False)
                    )

        # ★★★ [하이라이트] 바구니에 모은 데이터 한방에 DB로 전송 (벌크 인서트/업데이트) ★★★
        if test_results_to_create: TestResult.objects.bulk_create(test_results_to_create)
        if student_logs_to_create: StudentLog.objects.bulk_create(student_logs_to_create)
        if notifications_to_create: Notification.objects.bulk_create(notifications_to_create)
        if profiles_to_update: Profile.objects.bulk_update(profiles_to_update, ['status'])
        if users_to_update: User.objects.bulk_update(users_to_update, ['is_active'])

        # 누락자 스캔 및 알림도 최적화
        if cohort_id:
            if attempt == 1:
                target_users = Profile.objects.filter(cohort_id=cohort_id, is_manager=False, is_pl=False, status__in=['attending', 'caution', 'counseling']).values_list('user_id', flat=True)
            else:
                prev_attempt = attempt - 1
                target_users = TestResult.objects.filter(quiz=quiz, attempt_number=prev_attempt, is_pass=False, user__profile__cohort_id=cohort_id).values_list('user_id', flat=True)
            
            entered_users = TestResult.objects.filter(quiz=quiz, attempt_number=attempt).values_list('user_id', flat=True)
            missing_users = set(target_users) - set(entered_users)
            
            # 누락자 알림 바구니에 담기
            missing_notifications = []
            if missing_users:
                missing_profiles = Profile.objects.filter(user_id__in=missing_users)
                for missing_profile in missing_profiles:
                    missing_notifications.append(
                        Notification(recipient=request.user, sender=request.user, notification_type='exam',
                                     message=f"⚠️ [점수 누락] '{missing_profile.name}' 교육생의 {attempt}차 점수가 입력되지 않았습니다!",
                                     related_url=f"/quiz/manager/trainees/{missing_profile.id}/")
                    )
                if missing_notifications: Notification.objects.bulk_create(missing_notifications)

    return JsonResponse({'status': 'success', 'message': '채점 점수가 안전하게 저장되었습니다. (누락자는 알림 센터로 발송됨)'})